from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import tarfile
import zipfile
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import numpy as np

from k3_primary_evidence import K3_EVIDENCE_PATH, K3_PARAMETER_SOURCE, K3_TOTAL_B
from artifact_paths import portable_path
from docx import Document
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "outputs/019f6c42-2d53-7743-ab07-6293e2618dd7"
EPOCH = ROOT / "sources/epoch_all_ai_models_2026-07-31.csv"
ECI_REPRODUCED = ROOT / "sources/epoch_eci_reproduced_scores_2026-07-31.csv"
ECI_RELEASE_CROSSCHECK = OUT / "epoch_eci_reproduction_crosscheck_2026-07-31.csv"
ECI_COMPONENT = ROOT / "sources/epoch_eci_benchmarks_2026-07-31.csv"
ECI = ROOT / "sources/input_eci_parameter_regression_workbook_2026-07-17.xlsx"
ARCHIVE = ROOT / "sources/input_epoch_ai_models_archive_2026-07-17.zip"
OBS = OUT / "unified_model_observations_compute_enriched_2026-07-17.csv"
MEAS = OUT / "unified_model_measurements_long_compute_enriched_2026-07-17.csv"
AA_AUDIT = OUT / "aa_epoch_match_audit_compute_enriched_2026-07-17.csv"
VIEW_AUDIT = OUT / "epoch_archive_view_match_audit_2026-07-17.csv"
AA_SOURCE = ROOT / "sources/input_artificial_analysis_leaderboard_2026-07-17.txt"
METR_SOURCE = ROOT / "sources/metr_horizon_official_signals_2026-07-18.csv"
METR_RAW = ROOT / "sources/metr_benchmark_results_1_1_2026-07-18.yaml"
METR_METADATA = ROOT / "sources/metr_horizon_official_metadata_2026-07-18.json"
METR_AUDIT = OUT / "metr_primary_source_audit_2026-07-18.json"
METR_LEGACY = ROOT / "sources/metr_horizon_user_snapshot_2026-07-17.csv"
LATEX_SOURCE = ROOT / "sources/input_no_cot_arxiv_2606.07157v3_source.tar.gz"
FORECAST_LEDGER = ROOT / "sources/human_parameter_forecasts_2026-07-17.csv"
REGISTRY = OUT / "frontier_parameter_prediction_registry_v2.1_2026-07-17.docx"
WORKBOOK = OUT / "frontier_parameter_model_crowd_50pct_2026-07-17.xlsx"
INSPECT = OUT / "frontier_parameter_model_crowd_50pct_2026-07-17.xlsx.inspect.ndjson"
RESULT = OUT / "codex_independent_reaudit_metrics_2026-07-17.json"
NO_COT_EXACT_DATES = OUT / "no_cot_exact_date_model_audit_2026-07-18.csv"
NO_COT_DATE_OVERRIDES = ROOT / "sources/no_cot_exact_date_overrides_2026-07-18.csv"
NO_COT_DATE_METADATA = ROOT / "sources/no_cot_exact_date_collection_metadata_2026-07-18.json"
NO_COT_ARCHITECTURE_AUDIT = OUT / "no_cot_architecture_elasticity_audit_2026-07-18.json"
NO_COT_ARCHITECTURE_PREDICTIONS = OUT / "no_cot_architecture_elasticity_predictions_2026-07-18.csv"
IKP_CALIBRATION = ROOT / "sources/ikp_upstream_calibration_2026-07-18.json"
IKP_SUMMARY = ROOT / "sources/ikp_upstream_evaluation_summary_2026-07-18.json"
IKP_CONFIG = ROOT / "sources/ikp_upstream_models_2026-07-18.json"
IKP_METADATA = ROOT / "sources/ikp_source_metadata_2026-07-18.json"
IKP_AUDIT = OUT / "ikp_parameter_signal_audit_2026-07-18.json"
IKP_PREDICTIONS = OUT / "ikp_parameter_chronological_predictions_2026-07-18.csv"
IKP_OVERLAP = OUT / "ikp_parameter_incremental_overlap_2026-07-18.csv"
IKP_DENSING = ROOT / "sources/ikp_upstream_densing_analysis_2026-07-18.csv"
IKP_BENCHMARK_SCORES = ROOT / "sources/ikp_upstream_benchmark_scores_2026-07-18.csv"
IKP_CONDITIONAL_AUDIT = OUT / "ikp_conditional_benchmark_signal_audit_2026-07-18.json"
IKP_CONDITIONAL_PREDICTIONS = OUT / "ikp_conditional_benchmark_predictions_2026-07-18.csv"
OPUS5_EVIDENCE = ROOT / "sources/claude_opus_5_evidence_2026-07-31.json"


def read_csv(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def read_csv_bytes(data: bytes):
    return list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"), newline="")))


def sha(path: Path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def f(value):
    if value in (None, ""):
        return None
    return float(value)


def fit_line(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    slope, intercept = np.polyfit(x, y, 1)
    return float(slope), float(intercept), intercept + slope * x


def fit_ols(columns, y):
    matrix = np.column_stack([np.ones(len(y)), *[np.asarray(column, dtype=float) for column in columns]])
    coefficients, *_ = np.linalg.lstsq(matrix, np.asarray(y, dtype=float), rcond=None)
    return coefficients, matrix @ coefficients


def loocv_rmse_factor(columns, y):
    y = np.asarray(y, dtype=float)
    predictions = []
    for held_out in range(len(y)):
        keep = np.arange(len(y)) != held_out
        train_columns = [np.asarray(column, dtype=float)[keep] for column in columns]
        coefficients, _ = fit_ols(train_columns, y[keep])
        held_out_row = np.asarray([1, *[np.asarray(column, dtype=float)[held_out] for column in columns]])
        predictions.append(float(held_out_row @ coefficients))
    return rmse_factor(y - np.asarray(predictions))


def r2(y, fitted):
    y = np.asarray(y, dtype=float)
    fitted = np.asarray(fitted, dtype=float)
    return float(np.corrcoef(y, fitted)[0, 1] ** 2)


def rmse_factor(residual):
    residual = np.asarray(residual, dtype=float)
    return float(10 ** np.sqrt(np.mean(residual**2)))


def years_since_2024(value: str):
    d = date.fromisoformat(value[:10])
    return (d - date(2024, 1, 1)).days / 365.25


def geomean(values):
    return math.exp(sum(math.log(x) for x in values) / len(values))


def multiplicative_metrics(rows, prediction_field):
    errors = np.asarray(
        [math.log10(float(row[prediction_field]) / float(row["actual_b"])) for row in rows],
        dtype=float,
    )
    absolute = np.abs(errors)
    return {
        "n": len(rows),
        "families": len({row["family"] for row in rows}),
        "mean_absolute_log10_error": float(np.mean(absolute)),
        "median_multiplicative_error": float(10 ** np.median(absolute)),
        "rmse_log10": float(np.sqrt(np.mean(errors**2))),
        "within_2x": float(np.mean(absolute <= math.log10(2))),
    }


def independent_family_bootstrap(rows, delta_field, samples=20000):
    by_family = defaultdict(list)
    for row in rows:
        by_family[row["family"]].append(float(row[delta_field]))
    family_means = np.asarray([np.mean(values) for values in by_family.values()], dtype=float)
    rng = np.random.default_rng(20260718)
    draws = rng.choice(
        family_means,
        size=(samples, len(family_means)),
        replace=True,
    ).mean(axis=1)
    return {
        "observed_delta": float(family_means.mean()),
        "ci_90": [float(value) for value in np.quantile(draws, [0.05, 0.95])],
        "probability_blend_better": float(np.mean(draws < 0)),
        "families": len(family_means),
        "samples": samples,
    }


def identity_key(row):
    return tuple((row.get(field, "") or "").strip() for field in ("Model", "Organization", "Publication date"))


def inspect_table(sheet: str):
    for line in INSPECT.read_text(encoding="utf-8").splitlines():
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        if obj.get("kind") == "table" and obj.get("sheet") == sheet:
            return obj
    raise RuntimeError(f"Sheet not found in inspect output: {sheet}")


def table_with_header(sheet: str, first_header: str):
    values = inspect_table(sheet)["values"]
    for index, row in enumerate(values):
        if row and row[0] == first_header:
            headers = {name: column for column, name in enumerate(row) if name is not None}
            body = []
            for candidate in values[index + 1:]:
                if not candidate or not candidate[0]:
                    break
                body.append(candidate)
            return headers, body
    raise RuntimeError(f"Header {first_header!r} not found on sheet {sheet!r}")


epoch_rows = read_csv(EPOCH)
epoch_header = list(epoch_rows[0])

archive_result = {}
archive_views = {}
with zipfile.ZipFile(ARCHIVE) as zf:
    archived_all = zf.read("all_ai_models.csv")
    archive_result["all_ai_byte_identical"] = archived_all == EPOCH.read_bytes()
    archive_result["all_ai_sha256"] = hashlib.sha256(archived_all).hexdigest()
    master_by_key = defaultdict(list)
    for idx, row in enumerate(epoch_rows, start=2):
        master_by_key[identity_key(row)].append((idx, row))

    for name in ("frontier_ai_models.csv", "notable_ai_models.csv", "large_scale_ai_models.csv"):
        view = read_csv_bytes(zf.read(name))
        archive_views[name] = view
        common = sorted(set(view[0]) & set(epoch_header))
        matched = 0
        zero_disagreement = 0
        ambiguous_keys = 0
        unmatched = []
        selected_disagreements = []
        for row in view:
            candidates = master_by_key.get(identity_key(row), [])
            if not candidates:
                unmatched.append(row.get("Model", ""))
                continue
            if len(candidates) > 1:
                ambiguous_keys += 1
            ranked = []
            for source_row, candidate in candidates:
                disagreements = sum((row.get(col, "") or "") != (candidate.get(col, "") or "") for col in common)
                ranked.append((disagreements, source_row, candidate))
            disagreements, _, _ = min(ranked, key=lambda item: (item[0], item[1]))
            selected_disagreements.append(disagreements)
            matched += 1
            zero_disagreement += disagreements == 0
        archive_result[name] = {
            "records": len(view),
            "matched": matched,
            "zero_shared_field_disagreement": zero_disagreement,
            "ambiguous_exact_keys": ambiguous_keys,
            "unmatched": unmatched,
            "max_selected_shared_field_disagreements": max(selected_disagreements, default=0),
        }

observations = read_csv(OBS)
measurements = read_csv(MEAS)
obs_ids = [row["observation_id"] for row in observations]
measurement_ids = [row["measurement_id"] for row in measurements]
obs_id_set = set(obs_ids)

raw_json_errors = []
for row in observations:
    try:
        json.loads(row["source_record_json"])
    except Exception as exc:  # noqa: BLE001
        raw_json_errors.append((row["observation_id"], str(exc)))

source_counts = Counter(row["source"] for row in observations)
orphan_measurements = [row["measurement_id"] for row in measurements if row["observation_id"] not in obs_id_set]
view_sources = {"Epoch Frontier View", "Epoch Notable View", "Epoch Large-Scale View"}
view_included = [row["observation_id"] for row in observations if row["source"] in view_sources and row["model_level_include"] != "false"]

aa_rows = [row for row in observations if row["source"] == "AA"]
aa_by_display = defaultdict(list)
for row in aa_rows:
    aa_by_display[row["source_model_name"]].append(row)
aa_selection_violations = []
for display, rows in aa_by_display.items():
    selected = [row for row in rows if row["model_level_include"] == "true"]
    if len(selected) != 1:
        aa_selection_violations.append({"display": display, "selected": len(selected), "issue": "selected count"})
        continue
    numeric = [f(row["aa_intelligence_index"]) for row in rows if f(row["aa_intelligence_index"]) is not None]
    if numeric and f(selected[0]["aa_intelligence_index"]) != max(numeric):
        aa_selection_violations.append({"display": display, "selected": f(selected[0]["aa_intelligence_index"]), "max": max(numeric), "issue": "not max"})

epoch_names = Counter(row["Model"] for row in epoch_rows)
aa_checkpoint_link_errors = []
for row in aa_rows:
    if row["epoch_link_level"] in {"checkpoint", "checkpoint_system_configuration"}:
        if not row["matched_epoch_model"] or epoch_names[row["matched_epoch_model"]] == 0:
            aa_checkpoint_link_errors.append(row["observation_id"])

mega_result = {
    "observations": len(observations),
    "measurements": len(measurements),
    "source_counts": dict(sorted(source_counts.items())),
    "duplicate_observation_ids": len(obs_ids) - len(obs_id_set),
    "duplicate_measurement_ids": len(measurement_ids) - len(set(measurement_ids)),
    "orphan_measurements": len(orphan_measurements),
    "source_record_json_errors": len(raw_json_errors),
    "correlated_view_rows_included": len(view_included),
    "aa_display_names": len(aa_by_display),
    "aa_selection_violations": aa_selection_violations,
    "aa_checkpoint_link_errors": aa_checkpoint_link_errors,
}

# Independently compare raw records embedded in the megafile to the upstream files.
observations_by_id = {row["observation_id"]: row for row in observations}
epoch_raw_mismatches = []
for source_row, raw in enumerate(epoch_rows, start=2):
    observation = observations_by_id.get(f"epoch:{source_row:05d}")
    if observation is None or json.loads(observation["source_record_json"]) != raw:
        epoch_raw_mismatches.append(source_row)

view_specs = {
    "frontier_ai_models.csv": ("epoch_frontier_view", "Epoch Frontier View"),
    "notable_ai_models.csv": ("epoch_notable_view", "Epoch Notable View"),
    "large_scale_ai_models.csv": ("epoch_large_scale_view", "Epoch Large-Scale View"),
}
view_raw_mismatches = {}
for filename, (prefix, _) in view_specs.items():
    failures = []
    for source_row, raw in enumerate(archive_views[filename], start=2):
        observation = observations_by_id.get(f"{prefix}:{source_row:05d}")
        if observation is None or json.loads(observation["source_record_json"]) != raw:
            failures.append(source_row)
    view_raw_mismatches[filename] = failures

aa_lines = [line.strip() for line in AA_SOURCE.read_text(encoding="utf-8").splitlines()]
aa_start = aa_lines.index("Claude Fable 5 (with fallback)")
aa_end = aa_lines.index("Key definitions")
aa_source_records = []
for offset in range(aa_start, aa_end, 11):
    cells = aa_lines[offset:offset + 11]
    aa_source_records.append({
        "source_row": offset + 1,
        "raw": {
            "display_name": cells[0], "context_window": cells[1], "creator": cells[2], "provider": cells[3],
            "intelligence_index": cells[4], "blended_price_usd_per_million_tokens": cells[5],
            "output_speed_tokens_per_second": cells[6], "latency_first_chunk_seconds": cells[7],
            "total_response_seconds": cells[8], "footer_token_1": cells[9], "footer_token_2": cells[10],
        },
    })
aa_raw_mismatches = []
for source in aa_source_records:
    observation = observations_by_id.get(f"aa:{source['source_row']:04d}")
    if observation is None or json.loads(observation["source_record_json"]) != source["raw"]:
        aa_raw_mismatches.append(source["source_row"])

def xlsx_value(value):
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        serial = to_excel(value)
        return int(serial) if float(serial).is_integer() else float(serial)
    return "" if value is None else value


eci_workbook = load_workbook(ECI, read_only=True, data_only=True)
eci_sheet = eci_workbook["ECI Graph Data"]
eci_headers = [xlsx_value(cell.value) for cell in next(eci_sheet.iter_rows(min_row=1, max_row=1))[:12]]
eci_graph_rows = []
for excel_row, cells in enumerate(eci_sheet.iter_rows(min_row=2), start=2):
    values = [xlsx_value(cell.value) for cell in cells[:12]]
    values.extend([""] * (len(eci_headers) - len(values)))
    if not values[0]:
        continue
    eci_graph_rows.append((excel_row, dict(zip(eci_headers, values))))
regression_sheet = eci_workbook["Regression Data"]
regression_headers = [xlsx_value(cell.value) for cell in next(regression_sheet.iter_rows(min_row=1, max_row=1))[:14]]
regression_by_name = {}
for excel_row, cells in enumerate(regression_sheet.iter_rows(min_row=2), start=2):
    values = [xlsx_value(cell.value) for cell in cells[:14]]
    values.extend([""] * (len(regression_headers) - len(values)))
    if values[0]:
        regression_by_name[values[0]] = (excel_row, dict(zip(regression_headers, values)))

def objects_equal(left, right):
    if left.keys() != right.keys():
        return False
    for key in left:
        a, b = left[key], right[key]
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            if not math.isclose(float(a), float(b), rel_tol=1e-12, abs_tol=1e-12):
                return False
        elif a != b:
            return False
    return True


eci_current_by_name = {
    row["source_model_name"]: row for row in observations if row["source"] == "ECI"
}
eci_legacy_by_name = {
    row["source_model_name"]: row
    for row in observations
    if row["source"] == "ECI Legacy View"
}
eci_raw_mismatches = []
legacy_only_names = []
for excel_row, graph_raw in eci_graph_rows:
    model = graph_raw["Model"]
    observation = eci_current_by_name.get(model)
    if observation is None:
        observation = eci_legacy_by_name.get(model)
        legacy_only_names.append(model)
    if observation is None:
        eci_raw_mismatches.append({"row": excel_row, "model": model, "issue": "missing current or legacy-view observation"})
        continue
    embedded = json.loads(observation["source_record_json"])
    failures = []
    embedded_graph = embedded.get("graph_data")
    if observation["source"] == "ECI":
        embedded_graph = embedded_graph.get("legacy_graph_data") if isinstance(embedded_graph, dict) else None
    if not isinstance(embedded_graph, dict) or not objects_equal(embedded_graph, graph_raw):
        failures.append("graph_data")
    expected_regression = regression_by_name.get(model, (None, None))[1]
    embedded_regression = embedded.get("regression_data")
    if expected_regression is None:
        if embedded_regression is not None:
            failures.append("unexpected regression_data")
    elif not isinstance(embedded_regression, dict) or not objects_equal(embedded_regression, expected_regression):
        failures.append("regression_data")
    if failures:
        eci_raw_mismatches.append({"row": excel_row, "model": model, "issues": failures})

expected_legacy_only_names = {"DeepSeek-V3.1", "Kimi K2 (Sep 2025)"}
if set(legacy_only_names) != expected_legacy_only_names:
    eci_raw_mismatches.append({
        "issue": "retired legacy inventory mismatch",
        "expected": sorted(expected_legacy_only_names),
        "actual": sorted(set(legacy_only_names)),
    })

eci_reproduced_rows = read_csv(ECI_REPRODUCED)
eci_release_rows = read_csv(ECI_RELEASE_CROSSCHECK)
eci_release_by_name = {row["model"]: row for row in eci_release_rows}
eci_current_raw_mismatches = []
for source_row, reproduced_raw in enumerate(eci_reproduced_rows, start=2):
    model = reproduced_raw["Model"]
    observation = eci_current_by_name.get(model)
    if observation is None:
        eci_current_raw_mismatches.append({"row": source_row, "model": model, "issue": "missing current ECI observation"})
        continue
    embedded = json.loads(observation["source_record_json"])
    graph_data = embedded.get("graph_data")
    failures = []
    if not isinstance(graph_data, dict) or graph_data.get("reproduced_eci") != reproduced_raw:
        failures.append("reproduced_eci")
    expected_release = eci_release_by_name.get(model)
    if expected_release is None or graph_data.get("release_date_crosscheck") != expected_release:
        failures.append("release_date_crosscheck")
    if failures:
        eci_current_raw_mismatches.append({"row": source_row, "model": model, "issues": failures})
if set(eci_current_by_name) != {row["Model"] for row in eci_reproduced_rows}:
    eci_current_raw_mismatches.append({"issue": "current ECI model-set mismatch"})

metr_source_rows = read_csv(METR_SOURCE)
metr_observations = {
    row["source_model_name"]: row for row in observations if row["source"] == "METR" and row["record_type"] == "model"
}
metr_raw_mismatches = []
for source in metr_source_rows:
    observation = metr_observations.get(source["source_id"])
    if observation is None:
        metr_raw_mismatches.append({"source_id": source["source_id"], "issue": "missing observation"})
        continue
    raw = json.loads(observation["source_record_json"])
    failures = [field for field, expected in source.items() if raw.get(field) != expected]
    try:
        if raw.get("parsed_scaffolds") != json.loads(source["scaffolds_json"]):
            failures.append("parsed_scaffolds")
    except json.JSONDecodeError:
        failures.append("scaffolds_json_invalid")
    if failures:
        metr_raw_mismatches.append({"source_id": source["source_id"], "fields": failures})
metr_metadata = json.loads(METR_METADATA.read_text(encoding="utf-8"))
metr_primary_audit = json.loads(METR_AUDIT.read_text(encoding="utf-8"))
trend = metr_metadata["trend"]
expected_metr_law = {
    "allTimeDoublingDays": float(trend["all_time_stitched_point_estimate_days"]),
    "from2023DoublingDays": float(trend["from_2023_on_point_estimate_days"]),
    "from2023CiLowDays": float(trend["from_2023_on_ci_low_days"]),
    "from2023CiHighDays": float(trend["from_2023_on_ci_high_days"]),
    "exclusionRule": trend["exclusion_rule"],
    "long_tasks_version": trend["long_tasks_version"],
    "swaa_version": trend["swaa_version"],
}
metr_law_observation = observations_by_id.get("metr:law")
if metr_law_observation is None or json.loads(metr_law_observation["source_record_json"]) != expected_metr_law:
    metr_raw_mismatches.append({"source_id": "__trend_law__", "issue": "derived law mismatch"})
if metr_primary_audit.get("status") != "PASS" or metr_primary_audit["legacy_exact_crosscheck"]["exact_rows"] != 26 or metr_primary_audit["legacy_exact_crosscheck"]["mismatch_count"] != 0:
    metr_raw_mismatches.append({"source_id": "__primary_audit__", "issue": "official/legacy reconciliation is not an exact pass"})
for path in (METR_SOURCE, METR_RAW, METR_METADATA, METR_LEGACY):
    relative_path = str(path.relative_to(ROOT))
    if metr_primary_audit["files"].get(relative_path, {}).get("sha256") != sha(path):
        metr_raw_mismatches.append({"source_id": "__primary_audit__", "issue": f"hash mismatch for {relative_path}"})

with tarfile.open(LATEX_SOURCE, "r:gz") as archive:
    tex = archive.extractfile("arxiv_version.tex").read().decode("utf-8")

def latex_table(label):
    label_index = tex.index(f"\\label{{{label}}}")
    start = tex.rfind("\\begin{table", 0, label_index)
    end = tex.index("\\end{table}", label_index) + len("\\end{table}")
    return tex[start:end]


def normalize_tex(value):
    value = re.sub(r"\\textbf\{([^{}]*)\}", r"\1", value)
    value = re.sub(r"\\textit\{([^{}]*)\}", r"\1", value)
    value = value.replace(r"\text{k}", "k").replace(r"\,", "").replace("~", " ")
    value = value.replace("$", "").replace("{", "").replace("}", "").replace(r"\%", "%")
    return re.sub(r"\s+", " ", value).strip()


def latex_data_lines(block):
    return [
        line.strip() for line in block.splitlines()
        if "&" in line and line.strip().endswith(r"\\") and "multicolumn" not in line
    ]


def latex_cells(line):
    return [normalize_tex(cell) for cell in re.sub(r"\\\\\s*$", "", line).split("&")]


months = {name: index + 1 for index, name in enumerate(
    ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
)}
def month_iso(value):
    month, year = normalize_tex(value).split()
    return f"{int(year):04d}-{months[month]:02d}-01"


frontier_release = {}
for line in latex_data_lines(latex_table("tab:frontier-models")):
    cells = latex_cells(line)
    if len(cells) == 2 and cells[0] != "Model":
        frontier_release[cells[0]] = month_iso(cells[1])
frontier_release.update({
    "Sonnet 3.7": frontier_release["Claude 3.7 Sonnet"],
    "Opus 4": frontier_release["Claude Opus 4"],
    "Opus 4.1": frontier_release["Claude Opus 4.1"],
    "Opus 4.5": frontier_release["Claude Opus 4.5"],
    "Opus 4.6": frontier_release["Claude Opus 4.6"],
    "Opus 4.7": frontier_release["Claude Opus 4.7"],
})
nocot_expected = {}
for line in latex_data_lines(latex_table("tab:horizons-per-model")):
    raw_cells = [cell.strip() for cell in re.sub(r"\\\\\s*$", "", line).split("&")]
    if len(raw_cells) != 3 or normalize_tex(raw_cells[0]) == "Model":
        continue
    model = normalize_tex(raw_cells[0])
    nocot_expected[model] = {
        "model": model,
        "release_month": frontier_release[model],
        "time_horizon": normalize_tex(raw_cells[1]),
        "token_horizon": normalize_tex(raw_cells[2]),
    }

def parameter_billions(value):
    match = re.fullmatch(r"([0-9.]+)([BT])", normalize_tex(value).replace(" ", ""), re.I)
    return float(match.group(1)) * (1000 if match.group(2).upper() == "T" else 1)


open_meta = {}
developer = ""
for line in latex_table("tab:open-source-models").splitlines():
    group = re.search(r"\\textit\{([^}]+)\}", line)
    if "multicolumn" in line and group:
        developer = re.sub(r"\s*\([^)]*\)\s*", "", group.group(1)).strip()
        continue
    if "&" not in line or not line.strip().endswith(r"\\"):
        continue
    cells = latex_cells(line.strip())
    if len(cells) != 7 or cells[0] == "Model":
        continue
    parameter_parts = [part.strip() for part in cells[2].split("/")]
    total = parameter_billions(parameter_parts[0])
    active = parameter_billions(parameter_parts[1]) if len(parameter_parts) == 2 else total
    open_meta[cells[0]] = {
        "developer": developer, "model": cells[0], "releaseDate": month_iso(cells[1]),
        "totalB": int(total) if total.is_integer() else total,
        "activeB": int(active) if active.is_integer() else active,
        "layers": int(cells[3]), "architecture": cells[4], "reasoning": cells[5], "modality": cells[6],
    }
open_horizons = {}
for line in latex_data_lines(latex_table("tab:open-weight-precise-time-horizons")):
    cells = latex_cells(line)
    if len(cells) == 5 and re.fullmatch(r"[0-9.]+", cells[1]):
        open_horizons[cells[0]] = {"point": float(cells[1]), "median": float(cells[2]), "low": float(cells[3]), "high": float(cells[4])}
for model, meta in open_meta.items():
    nocot_expected[model] = {**meta, **open_horizons[model]}

scale_match = re.search(
    r"Doubling the 50\\% TH requires a \$([0-9.]+)\\times\$ increase in total parameters, a \$([0-9.]+)\\times\$ increase in active parameters, a \$([0-9.]+)\\times\$ increase in the layer count, or a \$([0-9.]+)\\times\$ increase in pretraining FLOPs",
    tex,
)
def macro_number(name):
    return float(re.search(rf"\\newcommand\{{\\{name}\}}\{{([0-9.]+)\}}", tex).group(1))
nocot_expected["__scaling_law__"] = {
    "totalParametersMultiplier": float(scale_match.group(1)),
    "activeParametersMultiplier": float(scale_match.group(2)),
    "layerMultiplier": float(scale_match.group(3)),
    "pretrainingFlopsMultiplier": float(scale_match.group(4)),
    "timeHorizonDoublingDays": int(macro_number("FINALDOUBLINGTIMETIME")),
    "tokenHorizonDoublingDays": int(macro_number("FINALDOUBLINGTIMETOKENS")),
}
nocot_observations = {row["source_model_name"]: row for row in observations if row["source"] == "No-CoT"}
nocot_raw_mismatches = []
for model, expected in nocot_expected.items():
    observation = nocot_observations.get(model)
    if observation is None or json.loads(observation["source_record_json"]) != expected:
        nocot_raw_mismatches.append(model)

eci_component_source_rows = read_csv(ECI_COMPONENT)
eci_component_raw_mismatches = []
for source_row, expected in enumerate(eci_component_source_rows, start=2):
    observation = observations_by_id.get(f"eci_component:{source_row:05d}")
    if observation is None or json.loads(observation["source_record_json"]) != expected:
        eci_component_raw_mismatches.append(str(source_row))

source_preservation = {
    "epoch_primary_rows_checked": len(epoch_rows),
    "epoch_primary_raw_mismatches": epoch_raw_mismatches,
    "epoch_view_rows_checked": sum(len(rows) for rows in archive_views.values()),
    "epoch_view_raw_mismatches": view_raw_mismatches,
    "aa_rows_checked": len(aa_source_records),
    "aa_raw_mismatches": aa_raw_mismatches,
    "eci_rows_checked": len(eci_graph_rows),
    "eci_raw_mismatches": eci_raw_mismatches,
    "eci_current_rows_checked": len(eci_reproduced_rows),
    "eci_current_raw_mismatches": eci_current_raw_mismatches,
    "eci_legacy_source_view_rows": len(eci_legacy_by_name),
    "eci_component_rows_checked": len(eci_component_source_rows),
    "eci_component_raw_mismatches": eci_component_raw_mismatches,
    "metr_rows_checked": len(metr_source_rows) + 1,
    "metr_raw_mismatches": metr_raw_mismatches,
    "nocot_rows_checked": len(nocot_expected),
    "nocot_raw_mismatches": nocot_raw_mismatches,
}
source_preservation_all_pass = not any([
    epoch_raw_mismatches,
    *view_raw_mismatches.values(),
    aa_raw_mismatches,
    eci_raw_mismatches,
    eci_current_raw_mismatches,
    eci_component_raw_mismatches,
    metr_raw_mismatches,
    nocot_raw_mismatches,
])

date_audit = {}
for source in ("Epoch", "ECI", "ECI Component", "AA", "No-CoT", "METR"):
    rows = [row for row in observations if row["source"] == source]
    date_audit[source] = {
        "rows": len(rows),
        "canonical_date_missing": sum(not row["canonical_release_date"] for row in rows),
        "date_conflicts_over_31_days": [
            {
                "model": row["source_model_name"], "source_date": row["source_release_date"],
                "canonical_date": row["canonical_release_date"], "delta_days": row["release_date_delta_days"],
                "canonical_source": row["canonical_release_date_source"],
            }
            for row in rows if row["date_conflict_flag"] == "true"
        ],
        "source_date_precision": dict(Counter(row["source_release_date_precision"] for row in rows)),
    }

# Independently reconcile the exact-date overlay against the raw no-CoT model
# inventory, the base unified registry, and the four explicit date-only overrides.
no_cot_exact_rows = read_csv(NO_COT_EXACT_DATES)
no_cot_override_rows = read_csv(NO_COT_DATE_OVERRIDES)
no_cot_date_metadata = json.loads(NO_COT_DATE_METADATA.read_text(encoding="utf-8"))
no_cot_model_names = set(nocot_expected) - {"__scaling_law__"}
no_cot_exact_by_model = {row["model"]: row for row in no_cot_exact_rows}
no_cot_override_by_model = {row["paper_model"]: row for row in no_cot_override_rows}

exact_date_issues = []
if len(no_cot_exact_by_model) != len(no_cot_exact_rows):
    exact_date_issues.append("duplicate model rows in exact-date ledger")
if set(no_cot_exact_by_model) != no_cot_model_names:
    exact_date_issues.append("exact-date ledger model set differs from the parsed paper model set")
if len(no_cot_override_by_model) != len(no_cot_override_rows):
    exact_date_issues.append("duplicate model rows in date-only override ledger")
if set(no_cot_override_by_model) != {
    "GPT-2", "GPT-3", "GPT-3.5", "Qwen 3 30B-A3B (2507)"
}:
    exact_date_issues.append("date-only override model set differs from the four audited exceptions")

for model, row in no_cot_exact_by_model.items():
    try:
        paper_date = date.fromisoformat(row["paper_month_date"])
        exact_date = date.fromisoformat(row["exact_release_date"])
    except ValueError:
        exact_date_issues.append(f"{model}: invalid ISO date")
        continue
    if paper_date.day != 1:
        exact_date_issues.append(f"{model}: paper month is not represented by month-start")
    if (exact_date - paper_date).days != int(row["day_offset_from_month_start"]):
        exact_date_issues.append(f"{model}: day offset does not recompute")
    override = no_cot_override_by_model.get(model)
    if override:
        if row["explicit_override"].lower() != "true":
            exact_date_issues.append(f"{model}: override not marked explicit")
        if row["exact_release_date"] != override["exact_release_date"]:
            exact_date_issues.append(f"{model}: override date mismatch")
        if row["parameter_join_policy"] != "date_only_no_epoch_parameter_join":
            exact_date_issues.append(f"{model}: override permits an unintended parameter join")
    else:
        base = nocot_observations.get(model)
        if base is None or row["exact_release_date"] != base["canonical_release_date"]:
            exact_date_issues.append(f"{model}: exact date differs from the base canonical registry")
        if row["explicit_override"].lower() != "false":
            exact_date_issues.append(f"{model}: non-override marked explicit")

metadata_hash_checks = {}
for relative, entry in no_cot_date_metadata["files"].items():
    path = ROOT / relative
    metadata_hash_checks[relative] = path.exists() and sha(path) == entry["sha256"]
if not all(metadata_hash_checks.values()):
    exact_date_issues.append("one or more exact-date source hashes do not reconcile")

no_cot_exact_date_reaudit = {
    "paper_models": len(no_cot_model_names),
    "exact_date_rows": len(no_cot_exact_rows),
    "unique_exact_date_models": len(no_cot_exact_by_model),
    "explicit_date_only_overrides": len(no_cot_override_rows),
    "remaining_month_only_models": sum(not row["exact_release_date"] for row in no_cot_exact_rows),
    "parameter_join_policies_for_overrides": sorted(
        {no_cot_exact_by_model[model]["parameter_join_policy"] for model in no_cot_override_by_model}
    ),
    "metadata_hash_checks": metadata_hash_checks,
    "issues": exact_date_issues,
    "all_pass": not exact_date_issues,
}

# Independently reproduce the no-CoT dense/MoE Pareto factors and the paired
# held-out point comparisons from the generated prediction ledger.
no_cot_architecture_audit = json.loads(
    NO_COT_ARCHITECTURE_AUDIT.read_text(encoding="utf-8")
)
no_cot_architecture_predictions = read_csv(NO_COT_ARCHITECTURE_PREDICTIONS)
architecture_panel_by_checkpoint = {}
for row in observations:
    if (
        row["source"] != "No-CoT"
        or not row["total_parameters_b"]
        or not row["nocot_time_horizon_median_minutes"]
        or row["canonical_display_name"].startswith("GPT-4")
        or row["architecture"] not in {"Dense", "MoE"}
    ):
        continue
    architecture_panel_by_checkpoint.setdefault(row["canonical_checkpoint_id"], row)
architecture_panel = list(architecture_panel_by_checkpoint.values())


def independent_pareto_factor(rows, parameter_field="total_parameters_b"):
    points = [
        {
            "model": row["canonical_display_name"],
            "parameters": float(row.get(parameter_field) or row["total_parameters_b"]),
            "horizon": float(row["nocot_time_horizon_median_minutes"]),
        }
        for row in rows
    ]
    frontier = [
        point
        for point in points
        if not any(
            candidate["parameters"] <= point["parameters"]
            and candidate["horizon"] >= point["horizon"]
            and (
                candidate["parameters"] < point["parameters"]
                or candidate["horizon"] > point["horizon"]
            )
            for candidate in points
        )
    ]
    slope, _, _ = fit_line(
        [math.log(point["parameters"]) for point in frontier],
        [math.log(point["horizon"]) for point in frontier],
    )
    return 2 ** (1 / slope), sorted(point["model"] for point in frontier)


independent_architecture_factors = {}
independent_canonical_architecture_factors = {}
for label, subset in {
    "pooled": architecture_panel,
    "dense": [row for row in architecture_panel if row["architecture"] == "Dense"],
    "moe": [row for row in architecture_panel if row["architecture"] == "MoE"],
}.items():
    factor, models = independent_pareto_factor(
        subset, parameter_field="raw_total_parameters_b"
    )
    independent_architecture_factors[label] = {
        "factor_per_horizon_doubling": factor,
        "pareto_models": models,
        "pareto_n": len(models),
    }
    canonical_factor, canonical_models = independent_pareto_factor(subset)
    independent_canonical_architecture_factors[label] = {
        "factor_per_horizon_doubling": canonical_factor,
        "pareto_models": canonical_models,
        "pareto_n": len(canonical_models),
    }


def independent_equal_family_delta(baseline_spec, candidate_spec):
    selected = [
        row
        for row in no_cot_architecture_predictions
        if row["split"] == "chronological_developer_holdout"
        and row["specification"] in {baseline_spec, candidate_spec}
    ]
    by_spec = {
        specification: {
            row["model"]: row
            for row in selected
            if row["specification"] == specification
        }
        for specification in (baseline_spec, candidate_spec)
    }
    common = sorted(set(by_spec[baseline_spec]) & set(by_spec[candidate_spec]))
    grouped = defaultdict(list)
    for model in common:
        baseline = by_spec[baseline_spec][model]
        candidate = by_spec[candidate_spec][model]
        if baseline["family"] != candidate["family"]:
            raise ValueError(f"Architecture prediction family mismatch: {model}")
        grouped[baseline["family"]].append(
            float(candidate["absolute_log_error"])
            - float(baseline["absolute_log_error"])
        )
    return {
        "n": len(common),
        "families": len(grouped),
        "observed_delta": float(
            np.mean([np.mean(values) for values in grouped.values()])
        ),
    }


independent_direct_delta = independent_equal_family_delta(
    "direct_pooled", "direct_architecture"
)
independent_pareto_delta = independent_equal_family_delta(
    "training_pareto_pooled", "training_pareto_architecture"
)
architecture_issues = []
if len(architecture_panel) != 35:
    architecture_issues.append("open-weight architecture panel is not 35 unique checkpoints")
if Counter(row["architecture"] for row in architecture_panel) != {"Dense": 19, "MoE": 16}:
    architecture_issues.append("dense/MoE inventory changed")
for label, recomputed in independent_architecture_factors.items():
    reported = no_cot_architecture_audit["paper_relationship_reproduction"][label]
    if not math.isclose(
        recomputed["factor_per_horizon_doubling"],
        reported["deterministic_bootstrap_median_reproduction"],
        rel_tol=0,
        abs_tol=1e-12,
    ):
        architecture_issues.append(f"{label} Pareto factor does not recompute")
    if recomputed["pareto_n"] != reported["pareto_n"]:
        architecture_issues.append(f"{label} Pareto membership count differs")
    canonical_reported = no_cot_architecture_audit[
        "canonical_parameter_relationship_sensitivity"
    ][label]
    canonical_recomputed = independent_canonical_architecture_factors[label]
    if not math.isclose(
        canonical_recomputed["factor_per_horizon_doubling"],
        canonical_reported["factor_per_horizon_doubling"],
        rel_tol=0,
        abs_tol=1e-12,
    ):
        architecture_issues.append(
            f"{label} canonical-parameter Pareto factor does not recompute"
        )
    if canonical_recomputed["pareto_n"] != canonical_reported["pareto_n"]:
        architecture_issues.append(
            f"{label} canonical-parameter Pareto membership count differs"
        )
for label, recomputed, key in (
    (
        "direct",
        independent_direct_delta,
        "direct_architecture_minus_pooled",
    ),
    (
        "training_pareto",
        independent_pareto_delta,
        "training_pareto_architecture_minus_pooled",
    ),
):
    reported = no_cot_architecture_audit["paired_comparisons"][
        "chronological_developer_holdout"
    ][key]
    if recomputed["n"] != reported["n"] or recomputed["families"] != reported["families"]:
        architecture_issues.append(f"{label} paired inventory does not recompute")
    if not math.isclose(
        recomputed["observed_delta"], reported["observed_delta"], rel_tol=0, abs_tol=1e-12
    ):
        architecture_issues.append(f"{label} paired delta does not recompute")
strict_predictions = [
    row
    for row in no_cot_architecture_predictions
    if row["split"] == "chronological_developer_holdout"
]
if any(
    row["test_family_excluded"] != "True" or row["strictly_earlier"] != "True"
    for row in strict_predictions
):
    architecture_issues.append("a primary architecture prediction violates chronology/family holdout")
for relative, expected in no_cot_architecture_audit["source_hashes"].items():
    if sha(ROOT / relative) != expected:
        architecture_issues.append(f"source hash mismatch: {relative}")
if no_cot_architecture_audit["decision"]["replace_pooled_live_elasticity_with_moe_specific"]:
    architecture_issues.append("MoE-specific elasticity was unexpectedly promoted")

no_cot_architecture_reaudit = {
    "panel_models": len(architecture_panel),
    "architecture_counts": dict(Counter(row["architecture"] for row in architecture_panel)),
    "independent_pareto_factors": independent_architecture_factors,
    "independent_canonical_parameter_pareto_factors": independent_canonical_architecture_factors,
    "independent_direct_delta": independent_direct_delta,
    "independent_training_pareto_delta": independent_pareto_delta,
    "strict_prediction_rows": len(strict_predictions),
    "issues": architecture_issues,
    "all_pass": not architecture_issues,
}

# Independently reconstruct the compute calibration samples and sequential regressions.
primary_epoch_compute = {
    row["canonical_checkpoint_id"]: row
    for row in observations
    if row["source"] == "Epoch" and row["epoch_training_compute_flop"] and row["epoch_parameters_b"]
}
stage1_map = {}
for row in observations:
    if row["source"] != "AA" or row["model_level_include"] != "true" or not row["aa_intelligence_index"]:
        continue
    if row["epoch_link_level"] not in {"checkpoint", "checkpoint_system_configuration"}:
        continue
    epoch = primary_epoch_compute.get(row["canonical_checkpoint_id"])
    if not epoch:
        continue
    current = stage1_map.get(row["canonical_checkpoint_id"])
    if current is None or f(row["aa_intelligence_index"]) > f(current[0]["aa_intelligence_index"]):
        stage1_map[row["canonical_checkpoint_id"]] = (row, epoch)
stage1 = sorted(stage1_map.values(), key=lambda pair: (pair[0]["canonical_release_date"], pair[0]["source_model_name"]))
s1_score = [f(aa["aa_intelligence_index"]) for aa, _ in stage1]
s1_date = [years_since_2024(aa["canonical_release_date"]) for aa, _ in stage1]
s1_logc = [math.log10(f(epoch["epoch_training_compute_flop"])) for _, epoch in stage1]
s1_score_slope, s1_score_intercept, s1_score_fit = fit_line(s1_score, s1_logc)
s1_score_residual = np.asarray(s1_logc) - s1_score_fit
s1_date_slope, s1_date_intercept, s1_date_fit = fit_line(s1_date, s1_score_residual)
s1_fit = s1_score_fit + s1_date_fit
s1_residual = np.asarray(s1_logc) - s1_fit
s1_joint_coefficients, s1_joint_fit = fit_ols([s1_score, s1_date], s1_logc)
s1_joint_residual = np.asarray(s1_logc) - s1_joint_fit

stage2 = []
for row in observations:
    if row["source"] != "Epoch Frontier View" or not row["epoch_training_compute_flop"] or not row["total_parameters_b"]:
        continue
    raw = json.loads(row["source_record_json"])
    if raw.get("Confidence") not in {"Confident", "Likely"}:
        continue
    if row["canonical_release_date"] < "2020-01-01":
        continue
    if "Language" not in str(raw.get("Domain", "")) or "Language modeling/generation" not in str(raw.get("Task", "")):
        continue
    stage2.append(row)
stage2.sort(key=lambda row: (row["canonical_release_date"], row["source_model_name"]))
s2_date = [years_since_2024(row["canonical_release_date"]) for row in stage2]
s2_logc = [math.log10(f(row["epoch_training_compute_flop"])) for row in stage2]
s2_logp = [math.log10(f(row["total_parameters_b"])) for row in stage2]
s2_compute_slope, s2_compute_intercept, s2_compute_fit = fit_line(s2_logc, s2_logp)
s2_compute_residual = np.asarray(s2_logp) - s2_compute_fit
s2_date_slope, s2_date_intercept, s2_date_fit = fit_line(s2_date, s2_compute_residual)
s2_fit = s2_compute_fit + s2_date_fit
s2_residual = np.asarray(s2_logp) - s2_fit
s2_joint_coefficients, s2_joint_fit = fit_ols([s2_logc, s2_date], s2_logp)
s2_joint_residual = np.asarray(s2_logp) - s2_joint_fit

target_names = {
    "Claude Fable 5": "Claude Fable 5 (with fallback)",
    "GPT-5.6 Sol": "GPT-5.6 Sol (max)",
    "Kimi K3": "Kimi K3",
    "Claude Opus 4.7 / 4.8 shared base": "Claude Opus 4.8 (max)",
    "GPT-5.5": "GPT-5.5 (xhigh)",
    "GPT-5.6 Terra": "GPT-5.6 Terra (max)",
    "Claude Sonnet 5": "Claude Sonnet 5 (max)",
    "GPT-5.6 Luna": "GPT-5.6 Luna (max)",
    "Grok 4.5": "Grok 4.5 (high)",
    "Claude Opus 5": "Claude Opus 5 (max)",
}
selected_aa = {row["source_model_name"]: row for row in aa_rows if row["model_level_include"] == "true"}
opus5_evidence = json.loads(OPUS5_EVIDENCE.read_text(encoding="utf-8"))
selected_aa["Claude Opus 5 (max)"] = {
    "source_model_name": "Claude Opus 5 (max)",
    "aa_intelligence_index": str(opus5_evidence["artificial_analysis"]["selected"]["score"]),
    "canonical_release_date": opus5_evidence["identity"]["release_date"],
}

raw_params = {}
raw_params_joint = {}
for target, aa_name in target_names.items():
    row = selected_aa[aa_name]
    score = f(row["aa_intelligence_index"])
    t = years_since_2024(row["canonical_release_date"])
    predicted_logc = s1_score_intercept + s1_score_slope * score + s1_date_intercept + s1_date_slope * t
    predicted_logp = s2_compute_intercept + s2_compute_slope * predicted_logc + s2_date_intercept + s2_date_slope * t
    raw_params[target] = 10**predicted_logp
    predicted_logc_joint = s1_joint_coefficients[0] + s1_joint_coefficients[1] * score + s1_joint_coefficients[2] * t
    predicted_logp_joint = s2_joint_coefficients[0] + s2_joint_coefficients[1] * predicted_logc_joint + s2_joint_coefficients[2] * t
    raw_params_joint[target] = 10**predicted_logp_joint
calibration_factor = math.sqrt(K3_TOTAL_B / raw_params["Kimi K3"] * 1500 / raw_params["Grok 4.5"])
compute_priors_t = {target: value * calibration_factor / 1000 for target, value in raw_params.items()}
joint_calibration_factor = math.sqrt(
    K3_TOTAL_B / raw_params_joint["Kimi K3"]
    * 1500
    / raw_params_joint["Grok 4.5"]
)
joint_compute_priors_t = {target: value * joint_calibration_factor / 1000 for target, value in raw_params_joint.items()}

horizon_headers, horizon_rows = table_with_header("Horizon Estimates", "Model / base")
horizon_by_model = {row[0]: row for row in horizon_rows}
existing = {name: float(row[horizon_headers["Existing central (T)"]]) for name, row in horizon_by_model.items()}
model_posterior = {name: float(row[horizon_headers["Posterior (T)"]]) for name, row in horizon_by_model.items()}
compute_corr = float(np.corrcoef(
    [math.log(compute_priors_t[name]) for name in target_names],
    [math.log(existing[name]) for name in target_names],
)[0, 1])

compute_result = {
    "k3_anchor_total_b": K3_TOTAL_B,
    "k3_anchor_source": K3_PARAMETER_SOURCE,
    "stage1_rows": len(stage1),
    "stage1_model_names": [aa["source_model_name"] for aa, _ in stage1],
    "stage1_score_slope": s1_score_slope,
    "stage1_score_intercept": s1_score_intercept,
    "stage1_date_residual_slope": s1_date_slope,
    "stage1_date_residual_intercept": s1_date_intercept,
    "stage1_r2": r2(s1_logc, s1_fit),
    "stage1_rmse_factor": rmse_factor(s1_residual),
    "stage1_loocv_rmse_factor_joint": loocv_rmse_factor([s1_score, s1_date], s1_logc),
    "stage1_joint_coefficients_intercept_score_date": s1_joint_coefficients.tolist(),
    "stage1_joint_r2": r2(s1_logc, s1_joint_fit),
    "stage1_joint_rmse_factor": rmse_factor(s1_joint_residual),
    "stage2_rows": len(stage2),
    "stage2_model_names": [row["source_model_name"] for row in stage2],
    "stage2_compute_slope": s2_compute_slope,
    "stage2_compute_intercept": s2_compute_intercept,
    "stage2_date_residual_slope": s2_date_slope,
    "stage2_date_residual_intercept": s2_date_intercept,
    "stage2_r2": r2(s2_logp, s2_fit),
    "stage2_rmse_factor": rmse_factor(s2_residual),
    "stage2_loocv_rmse_factor_joint": loocv_rmse_factor([s2_logc, s2_date], s2_logp),
    "stage2_joint_coefficients_intercept_compute_date": s2_joint_coefficients.tolist(),
    "stage2_joint_r2": r2(s2_logp, s2_joint_fit),
    "stage2_joint_rmse_factor": rmse_factor(s2_joint_residual),
    "calibration_factor": calibration_factor,
    "compute_priors_t": compute_priors_t,
    "joint_calibration_factor": joint_calibration_factor,
    "joint_compute_priors_t": joint_compute_priors_t,
    "joint_vs_workbook_prior_ratio": {target: joint_compute_priors_t[target] / compute_priors_t[target] for target in target_names},
    "compute_existing_log_correlation": compute_corr,
}

# Independently read crowd points from the final registry rather than the builder source.
doc = Document(REGISTRY)
forecast_table = doc.tables[0]
crowd_points = {"Fable": [], "Sol": []}
contributors = []
for row in forecast_table.rows[1:]:
    cells = [cell.text.strip() for cell in row.cells]
    contributors.append(cells[0])
    for key, index in (("Fable", 4), ("Sol", 5)):
        match = re.search(r"([0-9]+(?:\.[0-9]+)?)T", cells[index])
        if match:
            crowd_points[key].append(float(match.group(1)))
crowd_centers_from_displayed_points = {key: geomean(values) for key, values in crowd_points.items()}

# Exact centers from the normalized ledger. Superseded records are retained in the file but excluded here.
forecast_rows = read_csv(FORECAST_LEDGER)
forecast_ids = {row["forecast_id"] for row in forecast_rows}
superseded_ids = {row["supersedes"] for row in forecast_rows if row["supersedes"]}
assert superseded_ids <= forecast_ids
active_forecasts = [row for row in forecast_rows if row["forecast_id"] not in superseded_ids]

def crowd_point(row):
    low = float(row["low_t"])
    high = float(row["high_t"])
    central_text = (row.get("central_t") or "").strip()
    central = float(central_text) if central_text else None
    assert central is None or low <= central <= high
    return central if central is not None else math.sqrt(low * high)

fable_exact_points = [
    crowd_point(row)
    for row in active_forecasts
    if row["model"] == "Claude Fable 5"
]
sol_exact_points = [
    crowd_point(row)
    for row in active_forecasts
    if row["model"] == "GPT-5.6 Sol"
]
crowd_exact = {"Fable": geomean(fable_exact_points), "Sol": geomean(sol_exact_points)}
final_exact = {
    "Fable": math.sqrt(model_posterior["Claude Fable 5"] * crowd_exact["Fable"]),
    "Sol": math.sqrt(model_posterior["GPT-5.6 Sol"] * crowd_exact["Sol"]),
}

final_headers, final_rows = table_with_header("Final Ensemble", "Model / base")
final_by_model = {row[0]: row for row in final_rows}
workbook_final = {
    "Fable": float(final_by_model["Claude Fable 5"][final_headers["Final forecast (T)"]]),
    "Sol": float(final_by_model["GPT-5.6 Sol"][final_headers["Final forecast (T)"]]),
}

crowd_result = {
    "contributors": contributors,
    "contributor_count": len(contributors),
    "fable_n": len(crowd_points["Fable"]),
    "sol_n": len(crowd_points["Sol"]),
    "centers_from_displayed_one_decimal_points": crowd_centers_from_displayed_points,
    "exact_centers_from_stated_forecasts": crowd_exact,
    "model_posterior": {"Fable": model_posterior["Claude Fable 5"], "Sol": model_posterior["GPT-5.6 Sol"]},
    "recomputed_final_50pct_log_weight": final_exact,
    "workbook_final": workbook_final,
    "absolute_final_difference": {key: abs(final_exact[key] - workbook_final[key]) for key in final_exact},
}

# Validate workbook manifest hashes against the actual files listed in the cached inspection.
manifest_headers, manifest_rows = table_with_header("Source Manifest", "Source")
manifest_checks = []
for row in manifest_rows:
    path = Path(row[manifest_headers["Path"]])
    if not path.is_absolute():
        path = ROOT / path
    manifest_checks.append({
        "source": row[manifest_headers["Source"]],
        "path": portable_path(path),
        "exists": path.exists(),
        "size_matches": path.exists() and path.stat().st_size == int(row[manifest_headers["Bytes"]]),
        "sha_matches": path.exists() and sha(path) == row[manifest_headers["SHA-256"]],
    })

inspect_text = INSPECT.read_text(encoding="utf-8")
formula_error_tokens = sorted(set(re.findall(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A)", inspect_text)))
manifest_all_pass = all(
    row["exists"] and row["size_matches"] and row["sha_matches"]
    for row in manifest_checks
)

# Independently reproduce the gated IKP factual-capacity branch.  This does
# not call the production analyzer: source hashes, published fit, serving-base
# collapse, strict target fit, incremental overlap, and workbook integration are
# all recomputed from the frozen contracts and ledgers.
ikp_metadata = json.loads(IKP_METADATA.read_text(encoding="utf-8"))
ikp_source_hash_checks = []
for record in ikp_metadata["files"]:
    path = ROOT / record["local_path"]
    ikp_source_hash_checks.append(
        {
            "label": record["label"],
            "path": portable_path(path),
            "exists": path.exists(),
            "bytes_match": path.exists() and path.stat().st_size == record["bytes"],
            "sha256_match": path.exists() and sha(path) == record["sha256"],
        }
    )

ikp_audit = json.loads(IKP_AUDIT.read_text(encoding="utf-8"))
ikp_published = ikp_audit["published_reproduction"]
ikp_incremental = ikp_audit["incremental_overlap"]
ikp_target = ikp_audit["target_signal"]["fable"]
ikp_decision = ikp_audit["decision"]
ikp_calibration = json.loads(IKP_CALIBRATION.read_text(encoding="utf-8"))
ikp_configs = json.loads(IKP_CONFIG.read_text(encoding="utf-8"))["models"]
ikp_summary = {
    row["model"]: row for row in json.loads(IKP_SUMMARY.read_text(encoding="utf-8"))
}

calibration_points = ikp_calibration["calibration_points"]
ikp_log_params = np.log10([float(row["params_B"]) for row in calibration_points])
ikp_accuracy = np.asarray([float(row["accuracy"]) for row in calibration_points])
ikp_slope, ikp_intercept = np.polyfit(ikp_log_params, ikp_accuracy, 1)
ikp_fitted = ikp_intercept + ikp_slope * ikp_log_params
ikp_r2 = 1 - float(np.sum((ikp_accuracy - ikp_fitted) ** 2)) / float(
    np.sum((ikp_accuracy - np.mean(ikp_accuracy)) ** 2)
)

ikp_base_groups = defaultdict(list)
for row in calibration_points:
    model = row["model"]
    config = ikp_configs[model]
    ikp_base_groups[model.removesuffix("-think")].append(
        {
            "model": model,
            "params_b": float(row["params_B"]),
            "accuracy": float(row["accuracy"]),
            "family": row["family"],
            "vendor": row["vendor"],
            "release_date": config["release_date"],
        }
    )

ikp_collapsed = []
ikp_collapse_issues = []
for base_key, variants in sorted(ikp_base_groups.items()):
    params = {row["params_b"] for row in variants}
    families = {row["family"] for row in variants}
    vendors = {row["vendor"] for row in variants}
    if len(params) != 1 or len(families) != 1 or len(vendors) != 1:
        ikp_collapse_issues.append(base_key)
        continue
    ikp_collapsed.append(
        {
            "base_key": base_key,
            "params_b": next(iter(params)),
            "accuracy": float(np.mean([row["accuracy"] for row in variants])),
            "family": next(iter(families)),
            "vendor": next(iter(vendors)),
            "release_date": min(row["release_date"] for row in variants),
            "variants": [row["model"] for row in variants],
        }
    )

fable_config = ikp_configs["claude-fable-5"]
fable_accuracy = float(ikp_summary["claude-fable-5"]["accuracy"])
ikp_strict_train = [
    row
    for row in ikp_collapsed
    if row["release_date"] < fable_config["release_date"]
    and row["vendor"] != fable_config["vendor"]
]
strict_log_params = np.log10([row["params_b"] for row in ikp_strict_train])
strict_accuracy = np.asarray([row["accuracy"] for row in ikp_strict_train])
strict_slope, strict_intercept = np.polyfit(strict_log_params, strict_accuracy, 1)
strict_fable_b = float(10 ** ((fable_accuracy - strict_intercept) / strict_slope))

ikp_overlap = read_csv(IKP_OVERLAP)
for row in ikp_overlap:
    for field in (
        "actual_b",
        "existing_predicted_b",
        "ikp_predicted_b",
        "blended_predicted_b",
        "blend_minus_existing_abs_log10_error",
        "blend_weight",
    ):
        row[field] = float(row[field])
full_bootstrap = independent_family_bootstrap(
    ikp_overlap, "blend_minus_existing_abs_log10_error"
)

chronological_protocol = ikp_incremental["chronological_fixed_weight_subset"][
    "protocol"
]
chronological_thresholds = re.search(
    r"at least (\d+) earlier overlap rows from (\d+) other families",
    chronological_protocol,
)
if chronological_thresholds is None:
    raise ValueError("Could not parse IKP chronological-subset protocol")
minimum_prior_rows, minimum_prior_families = map(
    int, chronological_thresholds.groups()
)
chronological_overlap = []
for row in sorted(ikp_overlap, key=lambda item: (item["release_date"], item["family"])):
    earlier = [
        other
        for other in ikp_overlap
        if other["release_date"] < row["release_date"]
        and other["family"] != row["family"]
    ]
    if len(earlier) >= minimum_prior_rows and len(
        {other["family"] for other in earlier}
    ) >= minimum_prior_families:
        chronological_overlap.append(row)
chronological_bootstrap = independent_family_bootstrap(
    chronological_overlap, "blend_minus_existing_abs_log10_error"
)

fable_horizon_row = horizon_by_model["Claude Fable 5"]
fable_branch_values = {
    "existing": float(fable_horizon_row[horizon_headers["Existing central (T)"]]),
    "horizon": float(fable_horizon_row[horizon_headers["Horizon prior (T)"]]),
    "compute": float(
        fable_horizon_row[horizon_headers["Compute-structured prior (T)"]]
    ),
}
fable_workbook_weights = {
    "existing": float(fable_horizon_row[horizon_headers["Existing weight"]]),
    "horizon": float(fable_horizon_row[horizon_headers["Horizon weight"]]),
    "compute": float(fable_horizon_row[horizon_headers["Compute weight"]]),
}
ikp_evidence_weight = float(ikp_decision["incremental_evidence_weight"])
workbook_branch_weight = sum(fable_workbook_weights.values())
workbook_implied_ikp_weight = 1 - workbook_branch_weight
normalized_branch_weights = {
    key: value / workbook_branch_weight
    for key, value in fable_workbook_weights.items()
}
expected_policy_branch_weights = {
    key: (1 - ikp_evidence_weight) * value
    for key, value in normalized_branch_weights.items()
}
expected_policy_posterior = math.prod(
    fable_branch_values[key] ** expected_policy_branch_weights[key]
    for key in fable_branch_values
) * (strict_fable_b / 1000) ** ikp_evidence_weight
recomputed_workbook_posterior = math.prod(
    fable_branch_values[key] ** fable_workbook_weights[key]
    for key in fable_branch_values
) * (strict_fable_b / 1000) ** workbook_implied_ikp_weight
workbook_fable_with_ikp = float(fable_horizon_row[horizon_headers["Posterior (T)"]])

ikp_gates = ikp_decision["evidence_gates"]
ikp_nested = ikp_incremental["nested_leave_one_family_out_weight_learning"]
recomputed_promotion = all(
    (
        len(ikp_overlap) >= ikp_gates["minimum_overlap_models"],
        len({row["family"] for row in ikp_overlap})
        >= ikp_gates["minimum_overlap_families"],
        full_bootstrap["ci_90"][1]
        < ikp_gates["maximum_bootstrap_ci90_upper"],
        full_bootstrap["probability_blend_better"]
        >= ikp_gates["minimum_bootstrap_probability_better"],
        len(chronological_overlap)
        >= ikp_gates["minimum_chronological_subset_models"],
        len({row["family"] for row in chronological_overlap})
        >= ikp_gates["minimum_chronological_subset_families"],
        chronological_bootstrap["ci_90"][1]
        < ikp_gates["maximum_bootstrap_ci90_upper"],
        chronological_bootstrap["probability_blend_better"]
        >= ikp_gates["minimum_bootstrap_probability_better"],
        ikp_nested["all_test_families_excluded"],
    )
)
tested_ikp_weights = {float(row["blend_weight"]) for row in ikp_overlap}
recomputed_evidence_weight = (
    next(iter(tested_ikp_weights)) if recomputed_promotion else 0.0
)
ikp_reaudit_issues = []

def ikp_check(condition, message):
    if not condition:
        ikp_reaudit_issues.append(message)


ikp_check(all(row["exists"] and row["bytes_match"] and row["sha256_match"] for row in ikp_source_hash_checks), "pinned source hash mismatch")
ikp_check(not ikp_collapse_issues, f"serving-base collapse inconsistencies: {ikp_collapse_issues}")
ikp_check(len(calibration_points) == 93, "unexpected calibration configuration count")
ikp_check(len(ikp_collapsed) == 87, "unexpected distinct weight-base count")
ikp_check(abs(float(ikp_slope) - ikp_published["slope"]) < 1e-12, "published slope did not reproduce")
ikp_check(abs(float(ikp_intercept) - ikp_published["intercept"]) < 1e-12, "published intercept did not reproduce")
ikp_check(abs(ikp_r2 - ikp_published["r_squared"]) < 1e-12, "published R-squared did not reproduce")
ikp_check(len(ikp_strict_train) == 86, "unexpected strict Fable training row count")
ikp_check(all(row["vendor"] != "anthropic" for row in ikp_strict_train), "Anthropic leaked into strict Fable training")
ikp_check(max(row["release_date"] for row in ikp_strict_train) < fable_config["release_date"], "non-earlier row leaked into strict Fable training")
ikp_check(
    abs(strict_fable_b - ikp_target["strict_open_only_release_and_vendor_holdout"]["mean"]["estimates"]["forward_inverse"]["estimated_b"]) < 1e-8,
    "strict Fable estimate did not reproduce",
)
ikp_check(len(ikp_overlap) == ikp_incremental["models"], "overlap model count mismatch")
ikp_check(
    len({row["normalized_model"] for row in ikp_overlap}) == len(ikp_overlap),
    "duplicate overlap model identity",
)
ikp_check(
    len({row["family"] for row in ikp_overlap}) == ikp_incremental["families"],
    "overlap family count mismatch",
)
ikp_check(abs(multiplicative_metrics(ikp_overlap, "existing_predicted_b")["mean_absolute_log10_error"] - ikp_incremental["existing"]["mean_absolute_log10_error"]) < 1e-12, "existing overlap metric mismatch")
ikp_check(abs(multiplicative_metrics(ikp_overlap, "ikp_predicted_b")["mean_absolute_log10_error"] - ikp_incremental["ikp"]["mean_absolute_log10_error"]) < 1e-12, "IKP overlap metric mismatch")
ikp_check(abs(multiplicative_metrics(ikp_overlap, "blended_predicted_b")["mean_absolute_log10_error"] - ikp_incremental["blend_10pct"]["mean_absolute_log10_error"]) < 1e-12, "blended overlap metric mismatch")
ikp_check(np.allclose(full_bootstrap["ci_90"], ikp_incremental["family_bootstrap"]["ci_90"], atol=1e-12), "full-overlap bootstrap mismatch")
ikp_check(
    abs(
        full_bootstrap["probability_blend_better"]
        - ikp_incremental["family_bootstrap"]["probability_blend_better"]
    )
    < 1e-12,
    "full-overlap bootstrap probability mismatch",
)
ikp_check(
    len(chronological_overlap)
    == ikp_incremental["chronological_fixed_weight_subset"]["models"],
    "chronological overlap count mismatch",
)
ikp_check(
    len({row["family"] for row in chronological_overlap})
    == ikp_incremental["chronological_fixed_weight_subset"]["families"],
    "chronological overlap family count mismatch",
)
ikp_check(np.allclose(chronological_bootstrap["ci_90"], ikp_incremental["chronological_fixed_weight_subset"]["family_bootstrap"]["ci_90"], atol=1e-12), "chronological bootstrap mismatch")
ikp_check(
    abs(
        chronological_bootstrap["probability_blend_better"]
        - ikp_incremental["chronological_fixed_weight_subset"]["family_bootstrap"][
            "probability_blend_better"
        ]
    )
    < 1e-12,
    "chronological bootstrap probability mismatch",
)
ikp_check(len(tested_ikp_weights) == 1, "overlap uses inconsistent test weights")
ikp_check(
    ikp_decision["promote_incremental_ikp_weight"] == recomputed_promotion,
    "IKP promotion flag does not match declared evidence gates",
)
ikp_check(
    abs(ikp_evidence_weight - recomputed_evidence_weight) < 1e-15,
    "IKP evidence weight does not match promotion arithmetic",
)
ikp_check(
    abs(
        float(ikp_decision["incremental_final_weight_when_crowd_is_50pct"])
        - recomputed_evidence_weight * 0.5
    )
    < 1e-15,
    "IKP final weight does not match 50% crowd-mixture arithmetic",
)
ikp_check(
    ikp_decision["change_fable_center"] == recomputed_promotion,
    "IKP Fable-center decision does not match promotion",
)
ikp_check(
    ikp_decision["change_sol_center"] is False,
    "IKP unexpectedly changes unobserved Sol",
)
ikp_check(
    abs(recomputed_workbook_posterior - workbook_fable_with_ikp) < 1e-10,
    "workbook IKP posterior formula mismatch",
)
ikp_check(
    abs(workbook_implied_ikp_weight - ikp_evidence_weight) < 1e-15,
    "workbook IKP weight does not match the primary audit decision",
)
ikp_check(
    all(
        abs(fable_workbook_weights[key] - expected_policy_branch_weights[key])
        < 1e-15
        for key in fable_workbook_weights
    ),
    "workbook base-branch weights do not match the primary IKP policy",
)
ikp_check(
    abs(expected_policy_posterior - workbook_fable_with_ikp) < 1e-10,
    "workbook Fable posterior does not match the primary IKP policy",
)

ikp_reaudit = {
    "all_pass": not ikp_reaudit_issues,
    "issues": ikp_reaudit_issues,
    "source_hash_checks": ikp_source_hash_checks,
    "published_fit_recomputation": {
        "configurations": len(calibration_points),
        "slope": float(ikp_slope),
        "intercept": float(ikp_intercept),
        "r_squared": ikp_r2,
    },
    "serving_base_collapse": {
        "distinct_weight_bases": len(ikp_collapsed),
        "variants_collapsed": len(calibration_points) - len(ikp_collapsed),
        "issues": ikp_collapse_issues,
    },
    "strict_fable_recomputation": {
        "train_rows": len(ikp_strict_train),
        "train_families": len({row["family"] for row in ikp_strict_train}),
        "train_vendors": len({row["vendor"] for row in ikp_strict_train}),
        "train_max_date": max(row["release_date"] for row in ikp_strict_train),
        "anthropic_excluded": all(row["vendor"] != "anthropic" for row in ikp_strict_train),
        "estimated_b": strict_fable_b,
    },
    "incremental_overlap_recomputation": {
        "models": len(ikp_overlap),
        "families": len({row["family"] for row in ikp_overlap}),
        "existing": multiplicative_metrics(ikp_overlap, "existing_predicted_b"),
        "ikp": multiplicative_metrics(ikp_overlap, "ikp_predicted_b"),
        "blend_10pct": multiplicative_metrics(ikp_overlap, "blended_predicted_b"),
        "family_bootstrap": full_bootstrap,
        "chronological_models": len(chronological_overlap),
        "chronological_families": len({row["family"] for row in chronological_overlap}),
        "chronological_family_bootstrap": chronological_bootstrap,
        "chronological_minimum_prior_rows": minimum_prior_rows,
        "chronological_minimum_prior_families": minimum_prior_families,
        "declared_evidence_gates": ikp_gates,
        "recomputed_promotion": recomputed_promotion,
        "published_promotion": ikp_decision["promote_incremental_ikp_weight"],
    },
    "workbook_integration": {
        "branch_values_t": fable_branch_values,
        "workbook_branch_weights": fable_workbook_weights,
        "workbook_implied_ikp_weight": workbook_implied_ikp_weight,
        "expected_policy_branch_weights": expected_policy_branch_weights,
        "strict_ikp_fable_t": strict_fable_b / 1000,
        "ikp_evidence_weight": ikp_evidence_weight,
        "expected_policy_weighted_blend_t": expected_policy_posterior,
        "recomputed_workbook_posterior_t": recomputed_workbook_posterior,
        "workbook_posterior_t": workbook_fable_with_ikp,
        "formula_absolute_difference": abs(
            recomputed_workbook_posterior - workbook_fable_with_ikp
        ),
        "policy_absolute_difference": abs(
            expected_policy_posterior - workbook_fable_with_ikp
        ),
    },
}

# Independently reconstruct the conditional IKP-vs-standard-benchmark audit.
# Primary row-equal predictions are refit from the pinned source tables; all
# sensitivity ledgers, vendor bootstraps, chronology checks, and workbook rows
# are then reconciled without calling the production analyzer.
conditional_audit = json.loads(IKP_CONDITIONAL_AUDIT.read_text(encoding="utf-8"))
conditional_predictions = read_csv(IKP_CONDITIONAL_PREDICTIONS)
for row in conditional_predictions:
    for field in (
        "actual_b",
        "baseline_predicted_b",
        "candidate_predicted_b",
        "baseline_abs_log10_error",
        "candidate_abs_log10_error",
        "candidate_minus_baseline_abs_log10_error",
    ):
        row[field] = float(row[field])

conditional_excludes = {
    "minimax-m1-think",
    "hunyuan-a13b",
    "hunyuan-a13b-think",
    "hermes-3-405b",
    "ling-2.6-flash",
    "deepseek-v3.1-nex-n1",
    "intellect-3-think",
}
conditional_benchmarks = ("mmlu", "mmlu_pro", "gpqa_diamond", "simpleqa")
conditional_densing = read_csv(IKP_DENSING)
conditional_scores = read_csv(IKP_BENCHMARK_SCORES)
conditional_scores_by_model = {row["model"]: row for row in conditional_scores}
conditional_groups = defaultdict(list)
for source in conditional_densing:
    if source["model"] in conditional_excludes:
        continue
    conditional_groups[source["model"].removesuffix("-think")].append(source)

conditional_bases = []
conditional_collapse_issues = []
for base_key, variants in sorted(conditional_groups.items()):
    for field in ("vendor", "family", "arch", "params_B", "active_B", "release_date"):
        if len({row[field] for row in variants}) != 1:
            conditional_collapse_issues.append(f"{base_key}:{field}")
    benchmark_values = {}
    for benchmark in conditional_benchmarks:
        values = [
            float(conditional_scores_by_model[row["model"]][benchmark])
            for row in variants
            if row["model"] in conditional_scores_by_model
            and conditional_scores_by_model[row["model"]][benchmark]
        ]
        if values and max(values) - min(values) > 1e-12:
            conditional_collapse_issues.append(f"{base_key}:{benchmark}")
        benchmark_values[benchmark] = values[0] if values else None
    release_date = variants[0]["release_date"]
    conditional_bases.append(
        {
            "base_key": base_key,
            "vendor": variants[0]["vendor"],
            "family": variants[0]["family"],
            "params_b": float(variants[0]["params_B"]),
            "release_date": release_date,
            "release_day_years": (
                date.fromisoformat(release_date) - date(2024, 1, 1)
            ).days
            / 365.25,
            "ikp_score": float(
                np.mean([float(row["pen_acc"]) * 100 for row in variants])
            ),
            **benchmark_values,
        }
    )

conditional_primary_recomputed = []
for benchmark in conditional_benchmarks:
    panel = [row for row in conditional_bases if row[benchmark] is not None]
    for test in sorted(panel, key=lambda row: (row["release_date"], row["base_key"])):
        train = [
            row
            for row in panel
            if row["release_date"] < test["release_date"]
            and row["vendor"] != test["vendor"]
        ]
        if len(train) < 10 or len({row["vendor"] for row in train}) < 5:
            continue
        target = np.log10([row["params_b"] for row in train])
        baseline_coefficients, _ = fit_ols(
            [[row[benchmark] for row in train], [row["release_day_years"] for row in train]],
            target,
        )
        candidate_coefficients, _ = fit_ols(
            [
                [row[benchmark] for row in train],
                [row["release_day_years"] for row in train],
                [row["ikp_score"] for row in train],
            ],
            target,
        )
        baseline_log10 = float(
            np.asarray([1, test[benchmark], test["release_day_years"]])
            @ baseline_coefficients
        )
        candidate_log10 = float(
            np.asarray(
                [1, test[benchmark], test["release_day_years"], test["ikp_score"]]
            )
            @ candidate_coefficients
        )
        conditional_primary_recomputed.append(
            {
                "benchmark": benchmark,
                "base_key": test["base_key"],
                "baseline_predicted_b": float(10**baseline_log10),
                "candidate_predicted_b": float(10**candidate_log10),
            }
        )

conditional_primary_ledger = {
    (row["benchmark"], row["base_key"]): row
    for row in conditional_predictions
    if row["specification"] == "score_date"
    and row["training_weighting"] == "row_equal"
}
conditional_primary_differences = []
for reproduced in conditional_primary_recomputed:
    ledger = conditional_primary_ledger.get(
        (reproduced["benchmark"], reproduced["base_key"])
    )
    conditional_primary_differences.append(
        {
            "benchmark": reproduced["benchmark"],
            "base_key": reproduced["base_key"],
            "baseline_absolute_difference_b": (
                abs(reproduced["baseline_predicted_b"] - ledger["baseline_predicted_b"])
                if ledger
                else None
            ),
            "candidate_absolute_difference_b": (
                abs(reproduced["candidate_predicted_b"] - ledger["candidate_predicted_b"])
                if ledger
                else None
            ),
        }
    )


def conditional_vendor_bootstrap(rows):
    by_vendor = defaultdict(list)
    for row in rows:
        by_vendor[row["vendor"]].append(
            row["candidate_minus_baseline_abs_log10_error"]
        )
    means = np.asarray([np.mean(values) for values in by_vendor.values()])
    rng = np.random.default_rng(20260718)
    draws = rng.choice(means, size=(20000, len(means)), replace=True).mean(axis=1)
    return {
        "observed_delta": float(means.mean()),
        "ci_90": [float(value) for value in np.quantile(draws, [0.05, 0.95])],
        "probability_candidate_better": float(np.mean(draws < 0)),
        "vendors": len(means),
    }


conditional_sensitivity_checks = []
for benchmark in conditional_benchmarks:
    for weighting in ("row_equal", "vendor_equal"):
        for specification in ("score_date", "score_date_arch"):
            selected = [
                row
                for row in conditional_predictions
                if row["benchmark"] == benchmark
                and row["training_weighting"] == weighting
                and row["specification"] == specification
            ]
            if not selected:
                continue
            published = conditional_audit["heldout_results"][benchmark][
                "specifications"
            ][f"{weighting}__{specification}"]
            recomputed_bootstrap = conditional_vendor_bootstrap(selected)
            conditional_sensitivity_checks.append(
                {
                    "benchmark": benchmark,
                    "weighting": weighting,
                    "specification": specification,
                    "rows": len(selected),
                    "vendors": len({row["vendor"] for row in selected}),
                    "baseline_median_x": multiplicative_metrics(
                        selected, "baseline_predicted_b"
                    )["median_multiplicative_error"],
                    "candidate_median_x": multiplicative_metrics(
                        selected, "candidate_predicted_b"
                    )["median_multiplicative_error"],
                    "bootstrap": recomputed_bootstrap,
                    "published": published,
                }
            )

conditional_workbook_headers, conditional_workbook_rows = table_with_header(
    "IKP Signal Audit", "Benchmark"
)
conditional_narrative_headers, conditional_narrative_rows = table_with_header(
    "IKP Signal Audit", "Claim"
)
conditional_reaudit_issues = []


def conditional_check(condition, message):
    if not condition:
        conditional_reaudit_issues.append(message)


conditional_check(len(conditional_densing) == 100, "unexpected conditional source configuration count")
conditional_check(len(conditional_scores) == 81, "unexpected conditional benchmark row count")
conditional_check(len(conditional_groups) == 87, "unexpected conditional weight-base count")
conditional_check(not conditional_collapse_issues, f"conditional base-collapse issues: {conditional_collapse_issues}")
conditional_check(len(conditional_predictions) == 196, "unexpected conditional prediction count")
conditional_check(
    len(
        {
            (
                row["benchmark"],
                row["specification"],
                row["training_weighting"],
                row["base_key"],
            )
            for row in conditional_predictions
        }
    )
    == len(conditional_predictions),
    "duplicate conditional prediction key",
)
conditional_check(
    all(row["train_max_date"] < row["release_date"] for row in conditional_predictions),
    "conditional chronology violation",
)
conditional_check(
    all(row["test_vendor_excluded"] == "True" for row in conditional_predictions),
    "conditional vendor-holdout violation",
)
conditional_check(
    len(conditional_primary_recomputed) == len(conditional_primary_ledger) == 49,
    "primary conditional refit count mismatch",
)
conditional_check(
    all(
        row["baseline_absolute_difference_b"] is not None
        and row["baseline_absolute_difference_b"] < 1e-8
        and row["candidate_absolute_difference_b"] < 1e-8
        for row in conditional_primary_differences
    ),
    "primary conditional predictions did not refit",
)
for check in conditional_sensitivity_checks:
    published = check["published"]
    conditional_check(
        abs(check["baseline_median_x"] - published["baseline"]["median_multiplicative_error"]) < 1e-12,
        f"conditional baseline metric mismatch: {check['benchmark']} {check['weighting']} {check['specification']}",
    )
    conditional_check(
        abs(check["candidate_median_x"] - published["candidate_with_ikp"]["median_multiplicative_error"]) < 1e-12,
        f"conditional candidate metric mismatch: {check['benchmark']} {check['weighting']} {check['specification']}",
    )
    conditional_check(
        np.allclose(check["bootstrap"]["ci_90"], published["paired_vendor_bootstrap"]["ci_90"], atol=1e-12),
        f"conditional vendor bootstrap mismatch: {check['benchmark']} {check['weighting']} {check['specification']}",
    )
conditional_check(
    conditional_audit["heldout_results"]["gpqa_diamond"]["passing_specifications"] == 4,
    "GPQA conditional gate mismatch",
)
conditional_check(
    conditional_audit["heldout_results"]["mmlu"]["passing_specifications"] == 3,
    "MMLU conditional sensitivity count mismatch",
)
conditional_check(
    conditional_audit["decision"]["change_live_ikp_weight"] is False,
    "conditional audit unexpectedly changes IKP weight",
)
conditional_check(
    conditional_audit["decision"]["primary_parameter_signal_promoted"]
    == ikp_decision["promote_incremental_ikp_weight"],
    "conditional audit has stale primary IKP promotion state",
)
conditional_check(
    abs(
        float(
            conditional_audit["decision"][
                "retain_current_final_fable_ikp_weight"
            ]
        )
        - float(ikp_decision["incremental_final_weight_when_crowd_is_50pct"])
    )
    < 1e-15,
    "conditional audit has stale retained IKP weight",
)
conditional_check(
    conditional_audit["source_files"].get(str(IKP_AUDIT.relative_to(ROOT)))
    == sha(IKP_AUDIT),
    "conditional audit does not hash-pin the primary IKP decision",
)
conditional_check(len(conditional_workbook_rows) == 4, "conditional workbook table row count mismatch")
conditional_check(len(conditional_narrative_rows) == 6, "conditional workbook narrative row count mismatch")
conditional_check(
    all(row[conditional_narrative_headers["Status"]] == "STALE" for row in conditional_narrative_rows),
    "stale upstream narrative claims are not preserved in workbook",
)

conditional_reaudit = {
    "all_pass": not conditional_reaudit_issues,
    "issues": conditional_reaudit_issues,
    "source_inventory": {
        "raw_configurations": len(conditional_densing),
        "benchmark_rows": len(conditional_scores),
        "weight_bases": len(conditional_groups),
        "base_collapse_issues": conditional_collapse_issues,
    },
    "prediction_integrity": {
        "rows": len(conditional_predictions),
        "primary_refits": len(conditional_primary_recomputed),
        "maximum_primary_baseline_difference_b": max(
            row["baseline_absolute_difference_b"]
            for row in conditional_primary_differences
            if row["baseline_absolute_difference_b"] is not None
        ),
        "maximum_primary_candidate_difference_b": max(
            row["candidate_absolute_difference_b"]
            for row in conditional_primary_differences
            if row["candidate_absolute_difference_b"] is not None
        ),
        "all_strictly_chronological": all(
            row["train_max_date"] < row["release_date"]
            for row in conditional_predictions
        ),
        "all_test_vendors_excluded": all(
            row["test_vendor_excluded"] == "True" for row in conditional_predictions
        ),
    },
    "sensitivity_recomputations": conditional_sensitivity_checks,
    "policy_propagation": {
        "primary_parameter_signal_promoted": ikp_decision[
            "promote_incremental_ikp_weight"
        ],
        "primary_final_fable_ikp_weight": ikp_decision[
            "incremental_final_weight_when_crowd_is_50pct"
        ],
        "conditional_retained_final_fable_ikp_weight": conditional_audit[
            "decision"
        ]["retain_current_final_fable_ikp_weight"],
    },
    "workbook_integration": {
        "conditional_rows": len(conditional_workbook_rows),
        "narrative_rows": len(conditional_narrative_rows),
        "stale_rows": sum(
            row[conditional_narrative_headers["Status"]] == "STALE"
            for row in conditional_narrative_rows
        ),
    },
}

result = {
    "inputs": {
        "epoch_sha256": sha(EPOCH),
        "eci_component_sha256": sha(ECI_COMPONENT),
        "archive_sha256": sha(ARCHIVE),
        "observations_sha256": sha(OBS),
        "measurements_sha256": sha(MEAS),
        "aa_source_sha256": sha(AA_SOURCE),
        "metr_source_sha256": sha(METR_SOURCE),
        "metr_raw_sha256": sha(METR_RAW),
        "metr_metadata_sha256": sha(METR_METADATA),
        "metr_audit_sha256": sha(METR_AUDIT),
        "metr_legacy_crosscheck_sha256": sha(METR_LEGACY),
        "registry_sha256": sha(REGISTRY),
        "workbook_sha256": sha(WORKBOOK),
        "ikp_audit_sha256": sha(IKP_AUDIT),
        "ikp_predictions_sha256": sha(IKP_PREDICTIONS),
        "ikp_overlap_sha256": sha(IKP_OVERLAP),
        "ikp_conditional_audit_sha256": sha(IKP_CONDITIONAL_AUDIT),
        "ikp_conditional_predictions_sha256": sha(IKP_CONDITIONAL_PREDICTIONS),
        "k3_primary_evidence_sha256": sha(K3_EVIDENCE_PATH),
    },
    "epoch_archive_reconciliation": archive_result,
    "megafile_integrity": mega_result,
    "source_record_preservation": source_preservation,
    "source_record_preservation_all_pass": source_preservation_all_pass,
    "date_audit": date_audit,
    "no_cot_exact_date_reaudit": no_cot_exact_date_reaudit,
    "no_cot_architecture_reaudit": no_cot_architecture_reaudit,
    "ikp_parameter_signal_reaudit": ikp_reaudit,
    "ikp_conditional_benchmark_reaudit": conditional_reaudit,
    "compute_recomputation": compute_result,
    "crowd_and_final_recomputation": crowd_result,
    "workbook_manifest_checks": manifest_checks,
    "manifest_all_pass": manifest_all_pass,
    "workbook_formula_error_tokens": formula_error_tokens,
    "audit_scope_limits": [
        "Structural reconciliation and formulas were independently recomputed.",
        "No-CoT dense/MoE Pareto factors and paired held-out point deltas were independently recomputed from the unified panel and prediction ledger.",
        "IKP source hashes, published fit, duplicate serving-base collapse, strict target fit, overlap deltas, family bootstraps, and workbook blend were independently recomputed.",
        "IKP conditional benchmark predictions were refit from source rows; exact-date vendor holdouts, all sensitivity bootstraps, stale narrative claims, and workbook integration were independently reconciled.",
        "Semantic correctness of every one of the 274 AA-to-Epoch adjudications cannot be proven mechanically.",
        "Source truth can still contain upstream errors or estimates; exact preservation is not equivalent to factual correctness.",
    ],
}

RESULT.write_text(json.dumps(result, indent=2, sort_keys=True), encoding="utf-8")
print(json.dumps({
    "result": portable_path(RESULT),
    "archive": archive_result,
    "megafile": mega_result,
    "source_preservation": source_preservation,
    "source_preservation_all_pass": source_preservation_all_pass,
    "compute": {
        "stage1_rows": len(stage1),
        "stage1_r2": compute_result["stage1_r2"],
        "stage1_rmse_factor": compute_result["stage1_rmse_factor"],
        "stage2_rows": len(stage2),
        "stage2_r2": compute_result["stage2_r2"],
        "stage2_rmse_factor": compute_result["stage2_rmse_factor"],
        "stage1_loocv_rmse_factor_joint": compute_result["stage1_loocv_rmse_factor_joint"],
        "stage2_loocv_rmse_factor_joint": compute_result["stage2_loocv_rmse_factor_joint"],
        "joint_vs_workbook_prior_ratio": compute_result["joint_vs_workbook_prior_ratio"],
        "correlation": compute_corr,
    },
    "crowd_final": crowd_result,
    "manifest_all_pass": manifest_all_pass,
    "no_cot_exact_date_reaudit": no_cot_exact_date_reaudit,
    "no_cot_architecture_reaudit": no_cot_architecture_reaudit,
    "ikp_parameter_signal_reaudit": ikp_reaudit,
    "ikp_conditional_benchmark_reaudit": conditional_reaudit,
    "formula_error_tokens": formula_error_tokens,
}, indent=2))
if not manifest_all_pass:
    failed = [row["source"] for row in manifest_checks if not (row["exists"] and row["size_matches"] and row["sha_matches"])]
    raise SystemExit(f"Workbook source manifest validation failed: {failed}")
if not source_preservation_all_pass:
    raise SystemExit("Independent source-record preservation audit failed")
if formula_error_tokens:
    raise SystemExit(f"Workbook formula errors found: {formula_error_tokens}")
if not no_cot_exact_date_reaudit["all_pass"]:
    raise SystemExit(f"No-CoT exact-date re-audit failed: {exact_date_issues}")
if not no_cot_architecture_reaudit["all_pass"]:
    raise SystemExit(f"No-CoT architecture re-audit failed: {architecture_issues}")
if not ikp_reaudit["all_pass"]:
    raise SystemExit(f"IKP parameter-signal re-audit failed: {ikp_reaudit_issues}")
if not conditional_reaudit["all_pass"]:
    raise SystemExit(f"IKP conditional benchmark re-audit failed: {conditional_reaudit_issues}")
