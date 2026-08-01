#!/usr/bin/env python3
"""Integrity tests for the chronological parameter backtest artifact."""

from __future__ import annotations

import json
import hashlib
import math
import unittest
from pathlib import Path

from openpyxl import load_workbook

import run_parameter_backtest as backtest
from artifact_paths import portable_path
from aa_parameter_label_availability import (
    LEDGER_PATH,
    load_parameter_label_availability,
)
from aa_score_availability import (
    LEDGER_PATH as SCORE_LEDGER_PATH,
    RAW_PATH as SCORE_RAW_PATH,
)


WORK_DIR = Path(__file__).resolve().parent
RESULT_PATH = (
    WORK_DIR
    / "outputs"
    / "019f6c42-2d53-7743-ab07-6293e2618dd7"
    / "frontier_parameter_chronological_backtest_2026-07-17.json"
)
WORKBOOK_PATH = (
    WORK_DIR
    / "outputs"
    / "019f6c42-2d53-7743-ab07-6293e2618dd7"
    / "frontier_parameter_chronological_backtest_2026-07-17.xlsx"
)


class ParameterBacktestTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.data = json.loads(RESULT_PATH.read_text())

    def test_inventory_is_exact(self) -> None:
        self.assertEqual(self.data["inventory"]["AA"]["rows"], 50)
        self.assertEqual(self.data["inventory"]["ECI"]["rows"], 89)
        self.assertEqual(self.data["inventory"]["No-CoT"]["rows"], 35)
        self.assertEqual(self.data["inventory"]["Compute"]["rows"], 312)

    def test_every_prediction_is_strictly_chronological(self) -> None:
        for row in self.data["predictions"]:
            self.assertLess(
                row["train_max_date"],
                row.get("prediction_information_date") or row["release_date"],
                msg=f"date leakage in {row['panel']} / {row['model']}",
            )

    def test_preferred_predictions_hold_out_the_family(self) -> None:
        for row in self.data["predictions"]:
            self.assertEqual(row["split"], "chronological_family_holdout")
            self.assertTrue(
                row["test_family_excluded"],
                msg=f"family leakage in {row['panel']} / {row['model']}",
            )

    def test_prediction_arithmetic_recomputes(self) -> None:
        for row in self.data["predictions"]:
            expected_log_error = math.log10(row["predicted_b"] / row["actual_b"])
            expected_factor = 10 ** abs(expected_log_error)
            self.assertAlmostEqual(row["log10_error"], expected_log_error, places=11)
            self.assertAlmostEqual(row["multiplicative_error"], expected_factor, places=11)

    def test_ensemble_uses_one_canonical_parameter_truth(self) -> None:
        for row in self.data["ensemble_predictions"]:
            self.assertTrue(row["actual_canonical_checkpoint_id"])
            self.assertTrue(row["actual_parameter_source_family"])
            self.assertTrue(row["actual_parameter_source"])
            self.assertEqual(set(row["component_actuals_b"]), {
                component["panel"] for component in row["components"]
            })
            for component_actual in row["component_actuals_b"].values():
                ratio = max(row["actual_b"], component_actual) / min(
                    row["actual_b"], component_actual
                )
                self.assertLess(ratio, backtest.PARAMETER_TRUTH_MATCH_RATIO)

    def test_k3_uses_exact_primary_truth_instead_of_component_average(self) -> None:
        k3 = next(
            row
            for row in self.data["ensemble_predictions"]
            if row["normalized_model"] == "kimik3"
        )
        self.assertEqual(k3["actual_b"], 2780.0)
        self.assertEqual(
            k3["component_actuals_b"],
            {"AA": 2780.0, "ECI": 2780.0, "Compute": 2800.0},
        )
        self.assertNotEqual(
            k3["actual_b"],
            sum(k3["component_actuals_b"].values())
            / len(k3["component_actuals_b"]),
        )
        self.assertEqual(k3["actual_parameter_source_family"], "Primary evidence")
        self.assertEqual(k3["actual_parameter_source"], backtest.K3_PARAMETER_SOURCE)
        aa = next(component for component in k3["components"] if component["panel"] == "AA")
        external = next(
            row
            for row in self.data["external_checks"]
            if row["anchor"] == "Kimi K3" and "all Kimi held out" in row["method"]
        )
        self.assertAlmostEqual(aa["predicted_b"], external["predicted_b"], places=12)
        aa_row = next(
            row
            for row in self.data["predictions"]
            if row["model"] == "Kimi K3" and row["panel"] == "AA"
        )
        self.assertEqual(aa_row["prediction_information_date"], "2026-07-16")
        self.assertEqual(aa_row["train_max_date"], "2026-07-15")
        self.assertEqual(aa_row["train_n"], 46)
        self.assertEqual(aa_row["train_family_n"], 21)
        self.assertEqual(aa_row["exclusion_policy"], "all Kimi lineages held out")
        expected_log = sum(
            component["weight"] * math.log10(component["predicted_b"])
            for component in k3["components"]
        ) / sum(component["weight"] for component in k3["components"])
        self.assertAlmostEqual(k3["predicted_b"], 10**expected_log, places=11)

    def test_ensemble_has_unique_normalized_checkpoints(self) -> None:
        keys = [row["normalized_model"] for row in self.data["ensemble_predictions"]]
        self.assertEqual(len(keys), len(set(keys)))
        self.assertGreaterEqual(len(keys), 30)

    def test_ensemble_architecture_flags_are_consistent_when_observed(self) -> None:
        reasoning_conflicts = []
        for row in self.data["ensemble_predictions"]:
            self.assertFalse(row["moe_flag_conflict"], row["model"])
            if row["reasoning_flag_conflict"]:
                reasoning_conflicts.append(row["model"])
                self.assertIsNone(row["reasoning"])
            self.assertIn(row["moe"], (0, 1, None))
            self.assertIn(row["reasoning"], (0, 1, None))
        self.assertEqual(reasoning_conflicts, ["Gemma 4 31B IT"])

    def test_ensemble_improves_on_every_individual_current_component(self) -> None:
        metrics = self.data["current_like_metrics"]
        ensemble = metrics["Available-components ensemble"]["median_multiplicative_error"]
        for panel in ("AA", "ECI", "No-CoT", "Compute"):
            self.assertLess(ensemble, metrics[panel]["median_multiplicative_error"])

    def test_external_anchor_checks_are_not_anchor_locked(self) -> None:
        checks = self.data["external_checks"]
        self.assertEqual(len(checks), 5)
        k3 = [row for row in checks if row["anchor"] == "Kimi K3"]
        self.assertEqual(len(k3), 2)
        self.assertTrue(all(row["actual_b"] == 2780.0 for row in k3))
        for row in k3:
            expected_error = max(
                row["predicted_b"] / row["actual_b"],
                row["actual_b"] / row["predicted_b"],
            )
            self.assertTrue(
                math.isclose(
                    row["multiplicative_error"],
                    expected_error,
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                )
            )
            self.assertFalse(
                math.isclose(row["predicted_b"], row["actual_b"], rel_tol=0, abs_tol=1e-9),
                "An independently held-out check must not be overwritten by the anchor",
            )
        self.assertLess(k3[1]["train_n"], k3[0]["train_n"])
        self.assertIn("all Kimi held out", k3[1]["method"])

        grok = [row for row in checks if row["anchor"] == "Grok 4.5"]
        self.assertEqual(
            {row["method"] for row in grok},
            {
                "AA direct chronological panel fit",
                "ECI direct chronological 60/40 panel fit",
                "AA/ECI direct geometric ensemble before anchor lock",
            },
        )
        self.assertTrue(all(row["train_max_date"] < "2026-07-08" for row in grok))
        self.assertFalse(any("compute" in row["method"].lower() for row in grok))
        self.assertTrue(all("site" not in row["note"].lower() for row in grok[:2]))

        panels, _ = backtest._load_panels()
        recomputed = backtest._external_checks(panels)
        self.assertEqual(
            [(row["method"], row["predicted_b"]) for row in checks],
            [(row["method"], row["predicted_b"]) for row in recomputed],
        )

    def test_source_hashes_are_present(self) -> None:
        hashes = self.data["source_files"]
        timing = load_parameter_label_availability()
        expected_evidence = sum(
            len(record["local_evidence"]) for record in timing["records"]
        )
        self.assertEqual(len(hashes), 8 + expected_evidence)
        self.assertIn(portable_path(LEDGER_PATH), hashes)
        self.assertIn(portable_path(SCORE_LEDGER_PATH), hashes)
        self.assertIn(portable_path(SCORE_RAW_PATH), hashes)
        self.assertIn(portable_path(backtest.OPEN_MODEL_PARAMETER_TRUTH_PATH), hashes)
        self.assertIn(portable_path(backtest.AA_DETAILED_PATH), hashes)
        for path, digest in hashes.items():
            self.assertRegex(digest, r"^[0-9a-f]{64}$")
            source = Path(path)
            if not source.is_absolute():
                source = WORK_DIR / source
            self.assertEqual(digest, hashlib.sha256(source.read_bytes()).hexdigest())
        self.assertNotIn("source_data_contracts", self.data)

    def test_backtest_has_no_generated_site_dependency(self) -> None:
        source = Path(backtest.__file__).read_text(encoding="utf-8")
        self.assertNotIn("SITE_DATA_PATH", source)
        self.assertNotIn("site/public/data/forecast-model.json", source)

    def test_workbook_contains_corrected_three_component_k3_row(self) -> None:
        workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)
        sheet = workbook["OOS Predictions"]
        headers = {cell.value: index for index, cell in enumerate(sheet[5])}
        rows = [
            row
            for row in sheet.iter_rows(min_row=6, values_only=True)
            if row[headers["Panel"]] == "Available-components ensemble"
            and row[headers["Model"]] == "Kimi K3"
        ]
        self.assertEqual(len(rows), 1)
        row = rows[0]
        source = next(
            item
            for item in self.data["ensemble_predictions"]
            if item["model"] == "Kimi K3"
        )
        self.assertAlmostEqual(row[headers["Predicted (B)"]], source["predicted_b"])
        self.assertAlmostEqual(row[headers["Factor error"]], source["multiplicative_error"])
        self.assertEqual(row[headers["Components"]], 3)


if __name__ == "__main__":
    unittest.main(verbosity=2)
