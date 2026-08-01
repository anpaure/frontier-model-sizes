from __future__ import annotations

import json
import math
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "019f6c42-2d53-7743-ab07-6293e2618dd7"
WORKBOOK = OUT / "k3_calibrated_frontier_parameter_crosscheck_2026-07-17.xlsx"
BACKTEST = OUT / "frontier_parameter_chronological_backtest_2026-07-17.json"
REGRESSION = ROOT / "regression_results.json"


class K3CrosscheckCurrentDataTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.backtest = json.loads(BACKTEST.read_text(encoding="utf-8"))
        cls.regression = json.loads(REGRESSION.read_text(encoding="utf-8"))
        cls.workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)

    def test_revised_estimate_uses_current_exact_eci(self) -> None:
        expected = next(
            row
            for row in self.regression["eci"]["open_models"]
            if row["model"] == "Kimi K3"
        )
        sheet = self.workbook["Revised Estimates"]
        headers = {cell.value: index + 1 for index, cell in enumerate(sheet[5])}
        row_number = next(
            index
            for index in range(6, 100)
            if sheet.cell(index, headers["Model"]).value == "Kimi K3"
        )
        self.assertEqual(
            sheet.cell(row_number, headers["Release"]).value.date().isoformat(),
            expected["release_date"],
        )
        self.assertTrue(
            math.isclose(
                sheet.cell(row_number, headers["ECI"]).value,
                expected["score"],
                rel_tol=0,
                abs_tol=1e-12,
            )
        )
        self.assertEqual(sheet.cell(row_number, headers["Revised total (B)"]).value, 2780)
        self.assertEqual(
            sheet.cell(row_number, headers["ECI source / note"]).value,
            expected["source"],
        )

    def test_validation_uses_generated_aa_checks_and_frozen_eci_check(self) -> None:
        checks = {
            row["method"]: row
            for row in self.backtest["external_checks"]
            if row["anchor"] == "Kimi K3"
        }
        sheet = self.workbook["K3 Validation"]
        values = {
            sheet.cell(row, 1).value: sheet.cell(row, 2).value
            for row in range(7, 21)
        }
        self.assertTrue(
            math.isclose(
                values["Current family-balanced AA prediction (B)"],
                checks["AA expanding fit"]["predicted_b"],
                rel_tol=0,
                abs_tol=1e-9,
            )
        )
        self.assertTrue(
            math.isclose(
                values["Leave-entire-Kimi-family-out AA prediction (B)"],
                checks["AA expanding fit; all Kimi held out"]["predicted_b"],
                rel_tol=0,
                abs_tol=1e-9,
            )
        )

        eci_row = next(
            row
            for row in self.regression["eci"]["open_models"]
            if row["model"] == "Kimi K3"
        )
        self.assertTrue(
            math.isclose(
                values["Current reproduced Epoch ECI"],
                eci_row["score"],
                rel_tol=0,
                abs_tol=1e-12,
            )
        )
        frozen_prediction = values["Frozen ECI 60/40 prediction (B)"]
        expected_error = max(frozen_prediction / 2780, 2780 / frozen_prediction)
        self.assertTrue(
            math.isclose(
                values["Frozen ECI multiplicative error"],
                expected_error,
                rel_tol=0,
                abs_tol=1e-12,
            )
        )

    def test_stale_unavailable_claim_and_copied_predictions_are_absent(self) -> None:
        workbook_text = "\n".join(
            str(cell.value)
            for sheet in self.workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertNotIn("No K3 ECI", workbook_text)
        self.assertNotIn("ECI implied by 2.78T", workbook_text)
        self.assertNotIn("2.608T", workbook_text)
        self.assertNotIn("2.374T", workbook_text)


if __name__ == "__main__":
    unittest.main(verbosity=2)
