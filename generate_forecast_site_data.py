#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import math
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "outputs/019f6c42-2d53-7743-ab07-6293e2618dd7"
FINAL_WORKBOOK = OUT / "frontier_parameter_model_crowd_50pct_2026-07-17.xlsx"
BRANCH_WORKBOOK = OUT / "k3_calibrated_frontier_parameter_crosscheck_2026-07-17.xlsx"
LEDGER = ROOT / "sources/human_parameter_forecasts_2026-07-17.csv"
SITE_DATA = ROOT / "site/public/data/forecast-model.json"
OPENROUTER_RESULT = OUT / "openrouter_parameter_signal_backtest_2026-07-18.json"
OPENROUTER_MODELS = ROOT / "sources/openrouter_model_signals_2026-07-18.csv"
OPENROUTER_TIERS = ROOT / "sources/openrouter_endpoint_tier_signals_2026-07-18.csv"
OPENROUTER_DAILY = ROOT / "sources/openrouter_throughput_daily_2026-07-18.csv"
OPENROUTER_COLLECTION_AUDIT = OUT / "openrouter_collection_audit_2026-07-18.json"
OPENROUTER_HISTORY_MANIFEST = ROOT / "sources/openrouter_snapshot_history_manifest_2026-07-18.csv"
OPENROUTER_TEMPORAL_RESULT = OUT / "openrouter_temporal_stability_audit_2026-07-18.json"
OPENROUTER_REQUEST_WEIGHTED_RESULT = OUT / "openrouter_request_weighted_operational_audit_2026-07-18.json"
OPENROUTER_OFFICIAL_PRICES = ROOT / "sources/openrouter_official_endpoint_prices_2026-07-18.csv"
OPENROUTER_OFFICIAL_AUDIT = OUT / "openrouter_official_endpoint_audit_2026-07-18.json"
ECI_REPRODUCED_SCORES = ROOT / "sources/epoch_eci_reproduced_scores_2026-07-31.csv"
ECI_REPRODUCTION_CROSSCHECK = OUT / "epoch_eci_reproduction_crosscheck_2026-07-31.csv"
ECI_REPRODUCTION_AUDIT = OUT / "epoch_eci_reproduction_audit_2026-07-31.json"
EPOCH_SNAPSHOT_MANIFEST = ROOT / "sources/epoch_snapshot_manifest_2026-07-31.json"
ECI_COMPONENT_RESULT = OUT / "eci_component_extended_audit_2026-07-18.json"
ECI_COMPONENT_COMPARISON = OUT / "eci_component_active_incremental_comparison_2026-07-18.csv"
ECI_FIT_TOURNAMENT = OUT / "eci_fit_tournament_2026-07-18.json"
ECI_FIT_PREDICTIONS = OUT / "eci_fit_tournament_predictions_2026-07-18.csv"
ECI_FIT_TARGETS = OUT / "eci_fit_tournament_frontier_sensitivity_2026-07-18.csv"
ECI_HISTORICAL_SCORES = ROOT / "sources/epoch_eci_historical_model_scores_2026-07-18.csv"
ECI_HISTORICAL_METADATA = ROOT / "sources/epoch_eci_historical_fit_metadata_2026-07-18.json"
ECI_MULTIVARIATE_RESULT = OUT / "eci_multivariate_component_audit_2026-07-18.json"
ECI_MULTIVARIATE_PREDICTIONS = OUT / "eci_multivariate_component_predictions_2026-07-18.csv"
ECI_MULTIVARIATE_NARROW_CI_PREDICTIONS = OUT / "eci_multivariate_component_narrow_eci_ci_predictions_2026-07-18.csv"
ECI_MULTIVARIATE_TARGETS = OUT / "eci_multivariate_component_targets_2026-07-18.csv"
ECI_MULTIVARIATE_COVERAGE = OUT / "eci_multivariate_component_coverage_2026-07-18.csv"
POSTTRAINING_LINEAGE_RESULT = OUT / "posttraining_lineage_audit_2026-07-18.json"
POSTTRAINING_LINEAGE_EDGES = OUT / "posttraining_lineage_edges_2026-07-18.csv"
POSTTRAINING_LINEAGE_MEASUREMENTS = OUT / "posttraining_lineage_measurements_2026-07-18.csv"
POSTTRAINING_LINEAGE_PREDICTIONS = OUT / "posttraining_lineage_predictions_2026-07-18.csv"
FRONTIER_SHARED_BASE_SENSITIVITY = OUT / "frontier_shared_base_sensitivity_2026-07-18.csv"
FRONTIER_LINEAGE_EVIDENCE = OUT / "frontier_lineage_evidence_2026-07-18.csv"
AA_EXPANDED_RESULT = OUT / "aa_expanded_parameter_audit_2026-07-18.json"
AA_EXPANDED_PANEL = OUT / "aa_expanded_parameter_panel_2026-07-18.csv"
AA_EXPANDED_PREDICTIONS = OUT / "aa_expanded_parameter_predictions_2026-07-18.csv"
AA_EXPANDED_OVERLAPS = OUT / "aa_expanded_parameter_overlap_audit_2026-07-18.csv"
AA_DETAILED_RAW = ROOT / "sources/aa_detailed_snapshot_2026-07-31.html.gz"
AA_DETAILED_MODELS = ROOT / "sources/aa_detailed_model_signals_2026-07-31.csv"
AA_DETAILED_METADATA = ROOT / "sources/aa_detailed_collection_metadata_2026-07-31.json"
AA_CALIBRATION_OVERRIDES = ROOT / "sources/aa_calibration_primary_overrides_2026-07-31.json"
AA_INFERENCE_RESULT = OUT / "aa_inference_budget_audit_2026-07-18.json"
AA_DETAILED_PANEL = OUT / "aa_detailed_parameter_panel_2026-07-18.csv"
AA_REASONING_PAIRS = OUT / "aa_reasoning_pair_audit_2026-07-18.csv"
AA_DETAILED_CROSSCHECK = OUT / "aa_detailed_epoch_crosscheck_2026-07-18.csv"
AA_INFERENCE_PREDICTIONS = OUT / "aa_inference_budget_predictions_2026-07-18.csv"
AA_OPERATIONAL_RESULT = OUT / "aa_operational_signal_audit_2026-07-18.json"
AA_OPERATIONAL_PANEL = OUT / "aa_operational_parameter_panel_2026-07-18.csv"
AA_OPERATIONAL_PREDICTIONS = OUT / "aa_operational_backtest_predictions_2026-07-18.csv"
AA_OPENROUTER_CROSSCHECK = OUT / "aa_openrouter_operational_crosscheck_2026-07-18.csv"
ACTIVE_TRANSPORT_RESULT = OUT / "active_parameter_transport_audit_2026-07-18.json"
ACTIVE_TRANSPORT_PREDICTIONS = OUT / "active_parameter_transport_predictions_2026-07-18.csv"
ACTIVE_TRANSPORT_TARGETS = OUT / "active_parameter_transport_targets_2026-07-18.csv"
K3_RELEASE_EVIDENCE = ROOT / "sources/kimi_k3_release_evidence_2026-07-31.json"
OPENROUTER_ACTIVE_PRICE_RESULT = OUT / "openrouter_active_price_audit_2026-07-18.json"
OPENROUTER_ACTIVE_PRICE_MATCHES = OUT / "openrouter_active_parameter_match_audit_2026-07-18.csv"
OPENROUTER_ACTIVE_PRICE_PREDICTIONS = OUT / "openrouter_active_price_predictions_2026-07-18.csv"
OPENROUTER_ACTIVE_PRICE_TARGETS = OUT / "openrouter_active_price_targets_2026-07-18.csv"
OPENROUTER_HISTORICAL_PRICE_RAW = ROOT / "sources/openrouter_historical_price_ledger_2026-07-18.json.gz"
OPENROUTER_HISTORICAL_PRICE_POINTS = ROOT / "sources/openrouter_historical_price_change_points_2026-07-18.csv"
OPENROUTER_HISTORICAL_PRICE_METADATA = ROOT / "sources/openrouter_historical_price_collection_metadata_2026-07-18.json"
OPENROUTER_HISTORICAL_PRICE_RESULT = OUT / "openrouter_historical_price_audit_2026-07-18.json"
OPENROUTER_HISTORICAL_PRICE_MATCHES = OUT / "openrouter_historical_price_match_audit_2026-07-18.csv"
OPENROUTER_HISTORICAL_PRICE_PREDICTIONS = OUT / "openrouter_historical_price_backtest_predictions_2026-07-18.csv"
OPENROUTER_HISTORICAL_PRICE_TARGETS = OUT / "openrouter_historical_price_frontier_targets_2026-07-18.csv"
HF_ARCHITECTURE_RAW = ROOT / "sources/huggingface_architecture_config_snapshot_2026-07-18.json.gz"
HF_ARCHITECTURE_SIGNALS = ROOT / "sources/huggingface_architecture_config_signals_2026-07-18.csv"
HF_ARCHITECTURE_AUDIT = OUT / "huggingface_architecture_config_collection_audit_2026-07-18.json"
NO_COT_EXACT_DATE_AUDIT = OUT / "no_cot_exact_date_audit_2026-07-18.json"
NO_COT_EXACT_DATE_MODELS = OUT / "no_cot_exact_date_model_audit_2026-07-18.csv"
NO_COT_ARCHITECTURE_AUDIT = OUT / "no_cot_architecture_elasticity_audit_2026-07-18.json"
NO_COT_ARCHITECTURE_PREDICTIONS = OUT / "no_cot_architecture_elasticity_predictions_2026-07-18.csv"
FRONTIER_PRIMARY_EVIDENCE = ROOT / "sources/frontier_primary_evidence_2026-07-18.csv"
FRONTIER_PRIMARY_METADATA = ROOT / "sources/frontier_primary_evidence_collection_metadata_2026-07-18.json"
FRONTIER_PRIMARY_AUDIT = OUT / "frontier_primary_evidence_audit_2026-07-18.json"
FRONTIER_PRIMARY_CONTROLS = OUT / "frontier_primary_evidence_controls_2026-07-18.csv"
METR_PRIMARY_SIGNALS = ROOT / "sources/metr_horizon_official_signals_2026-07-18.csv"
METR_PRIMARY_RAW = ROOT / "sources/metr_benchmark_results_1_1_2026-07-18.yaml"
METR_PRIMARY_METADATA = ROOT / "sources/metr_horizon_official_metadata_2026-07-18.json"
METR_PRIMARY_AUDIT = OUT / "metr_primary_source_audit_2026-07-18.json"
IKP_RESULT = OUT / "ikp_parameter_signal_audit_2026-07-18.json"
IKP_CONDITIONAL_RESULT = OUT / "ikp_conditional_benchmark_signal_audit_2026-07-18.json"
OPUS_5_EVIDENCE = ROOT / "sources/claude_opus_5_evidence_2026-07-31.json"
K3_EFFICIENCY_PRIOR = OUT / "k3_efficiency_prior_2026-08-01.json"

BASELINE_FINAL_WEIGHTS = {
    "aa": 9.5625,
    "eci": 9.5625,
    "price": 3.375,
    "horizon": 25.0,
    "compute": 2.5,
    "crowd": 50.0,
}
CROWD_ENSEMBLE_MODELS = {"Claude Fable 5", "GPT-5.6 Sol"}
OPUS_5_MODEL = "Claude Opus 5"


def clean(value):
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return value


def sha256(path: Path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def geomean(values):
    values = list(values)
    return math.exp(sum(math.log(value) for value in values) / len(values))


def default_weights_from_ikp(ikp):
    decision = ikp["decision"]
    promoted = bool(decision["promote_incremental_ikp_weight"])
    evidence_weight = float(decision["incremental_evidence_weight"])
    final_weight = float(decision["incremental_final_weight_when_crowd_is_50pct"])
    if not all(math.isfinite(value) for value in (evidence_weight, final_weight)):
        raise ValueError("IKP decision weights must be finite")
    if not 0 <= evidence_weight <= 1 or not 0 <= final_weight <= 0.5:
        raise ValueError("IKP decision weights are outside their admissible range")
    if not promoted and (evidence_weight != 0 or final_weight != 0):
        raise ValueError("A non-promoted IKP signal must have zero live weight")
    if promoted and evidence_weight <= 0:
        raise ValueError("A promoted IKP signal must have positive live weight")
    if not math.isclose(final_weight, 0.5 * evidence_weight, rel_tol=0, abs_tol=1e-12):
        raise ValueError("IKP final weight must equal half its evidence weight under the 50% crowd policy")

    # The five baseline factors occupy the 50% evidence half. Any promoted IKP
    # allocation proportionally shrinks that block; a failed gate restores it in full.
    baseline_scale = 1 - evidence_weight
    weights = {
        factor: round(weight * baseline_scale, 12)
        for factor, weight in BASELINE_FINAL_WEIGHTS.items()
        if factor != "crowd"
    }
    weights["ikp"] = round(100 * final_weight, 12)
    weights["crowd"] = BASELINE_FINAL_WEIGHTS["crowd"]
    if not math.isclose(sum(weights.values()), 100.0, rel_tol=0, abs_tol=1e-10):
        raise ValueError(f"Decision-derived site weights do not sum to 100: {weights}")
    return weights


def crowd_point(row):
    low = float(row["low_t"])
    high = float(row["high_t"])
    central_text = (row.get("central_t") or "").strip()
    central = float(central_text) if central_text else None
    if central is not None and not low <= central <= high:
        raise ValueError(f"Stated central outside bounds in {row['forecast_id']}: {central}")
    return central if central is not None else math.sqrt(low * high)


def active_forecasts():
    with LEDGER.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    superseded = {row["supersedes"] for row in rows if row["supersedes"]}
    active = [row for row in rows if row["forecast_id"] not in superseded]
    keys = [(row["contributor"], row["model"]) for row in active]
    if len(keys) != len(set(keys)):
        raise ValueError("Multiple active forecasts for one contributor/model pair")
    return active


def rows_by_header(workbook, sheet_name, header_row, key_header):
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[header_row]]
    index = {header: column for column, header in enumerate(headers) if header is not None}
    rows = {}
    started = False
    for values in sheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
        key = values[index[key_header]]
        if key in (None, ""):
            if started:
                break
            continue
        started = True
        rows[str(key)] = {header: clean(values[column]) for header, column in index.items()}
    return rows


def forecast_from_weights(model, weights):
    if model["lockedAnchor"]:
        return model["disclosedT"]
    evidence_keys = [key for key in weights if key != "crowd"]
    available_evidence = {
        key: model["factors"].get(key)
        for key in evidence_keys
        if model["factors"].get(key) is not None and weights.get(key, 0) > 0
    }
    requested_evidence_weight = sum(max(0.0, weights.get(key, 0)) for key in evidence_keys)
    available_evidence_weight = sum(weights[key] for key in available_evidence)
    effective = {}
    if available_evidence_weight > 0 and requested_evidence_weight > 0:
        for key, value in available_evidence.items():
            effective[key] = (
                value,
                requested_evidence_weight * weights[key] / available_evidence_weight,
            )
    crowd_value = model["factors"].get("crowd")
    if crowd_value is not None and weights.get("crowd", 0) > 0:
        effective["crowd"] = (crowd_value, weights["crowd"])
    total_weight = sum(weight for _, weight in effective.values())
    if not effective or total_weight <= 0:
        raise ValueError(f"No usable factors for {model['name']}")
    return math.exp(
        sum(weight / total_weight * math.log(value) for value, weight in effective.values())
    )


def main():
    epoch_snapshot = json.loads(EPOCH_SNAPSHOT_MANIFEST.read_text(encoding="utf-8"))
    aa_detailed_metadata = json.loads(AA_DETAILED_METADATA.read_text(encoding="utf-8"))
    k3_evidence = json.loads(K3_RELEASE_EVIDENCE.read_text(encoding="utf-8"))
    k3 = k3_evidence["kimi_k3"]
    k3_efficiency = json.loads(K3_EFFICIENCY_PRIOR.read_text(encoding="utf-8"))
    if epoch_snapshot["snapshot_as_of"] != "2026-07-31":
        raise ValueError("Site build requires the installed 2026-07-31 Epoch snapshot")
    if aa_detailed_metadata["snapshot_date"] != "2026-07-31":
        raise ValueError("Site build requires the installed 2026-07-31 AA snapshot")
    if (
        k3["total_parameters_b_exact"] != 2780.0
        or k3["activated_parameters_b_exact"] != 104.2
        or not k3["parameter_count_disclosed"]
        or not k3["activated_parameter_count_disclosed"]
    ):
        raise ValueError("K3 official evidence must retain exact 2.78T total / 104.2B active")
    if (
        k3_efficiency["anchor"]["total_parameters_t"] != 2.780
        or not k3_efficiency["decision"]["apply_center_preserving_upper_tail_projection"]
        or k3_efficiency["decision"]["change_point_centers"]
        or k3_efficiency["decision"]["incremental_point_center_weight"] != 0
        or k3_efficiency["decision"]["change_crowd_weight"]
        or k3_efficiency["decision"]["crowd_weight_for_fable_and_sol"] != 0.5
        or k3_efficiency["decision"]["rejected_nonlinear_eci_weight"] != 0
        or k3_efficiency["decision"]["literal_constraint_enforced_when_reference_below_center"]
        or k3_efficiency["method"]["nonlinear_forms_used"]
    ):
        raise ValueError("K3 efficiency-projection contract is inconsistent with the live model")

    opus_5_evidence = json.loads(OPUS_5_EVIDENCE.read_text(encoding="utf-8"))
    opus_5_identity = opus_5_evidence["identity"]
    opus_5_aa = opus_5_evidence["artificial_analysis"]["selected"]
    opus_5_epoch = opus_5_evidence["epoch"]
    opus_5_availability = opus_5_evidence["availability"]
    if (
        opus_5_identity["canonical_name"] != OPUS_5_MODEL
        or opus_5_identity["release_date"] != "2026-07-24"
        or opus_5_identity["parameter_disclosed"]
        or opus_5_identity["same_weight_identity_disclosed"]
        or opus_5_identity["base_identity_policy"] != "unique_base"
    ):
        raise ValueError("Opus 5 identity evidence violates the distinct undisclosed-base policy")
    if any(opus_5_availability[key] for key in ("metr", "no_cot", "ikp")):
        raise ValueError("Opus 5 must not gain an unavailable direct measurement")

    openrouter = json.loads(OPENROUTER_RESULT.read_text(encoding="utf-8"))
    openrouter_temporal = json.loads(
        OPENROUTER_TEMPORAL_RESULT.read_text(encoding="utf-8")
    )
    openrouter_request_weighted = json.loads(
        OPENROUTER_REQUEST_WEIGHTED_RESULT.read_text(encoding="utf-8")
    )
    openrouter_collection = json.loads(
        OPENROUTER_COLLECTION_AUDIT.read_text(encoding="utf-8")
    )
    openrouter_official = json.loads(
        OPENROUTER_OFFICIAL_AUDIT.read_text(encoding="utf-8")
    )
    eci_reproduction = json.loads(
        ECI_REPRODUCTION_AUDIT.read_text(encoding="utf-8")
    )
    eci_component = json.loads(ECI_COMPONENT_RESULT.read_text(encoding="utf-8"))
    eci_fit = json.loads(ECI_FIT_TOURNAMENT.read_text(encoding="utf-8"))
    eci_multivariate = json.loads(
        ECI_MULTIVARIATE_RESULT.read_text(encoding="utf-8")
    )
    posttraining_lineage = json.loads(
        POSTTRAINING_LINEAGE_RESULT.read_text(encoding="utf-8")
    )
    aa_expanded = json.loads(AA_EXPANDED_RESULT.read_text(encoding="utf-8"))
    aa_inference = json.loads(AA_INFERENCE_RESULT.read_text(encoding="utf-8"))
    aa_operational = json.loads(AA_OPERATIONAL_RESULT.read_text(encoding="utf-8"))
    active_transport = json.loads(ACTIVE_TRANSPORT_RESULT.read_text(encoding="utf-8"))
    active_price = json.loads(OPENROUTER_ACTIVE_PRICE_RESULT.read_text(encoding="utf-8"))
    historical_price = json.loads(
        OPENROUTER_HISTORICAL_PRICE_RESULT.read_text(encoding="utf-8")
    )
    historical_price_metadata = json.loads(
        OPENROUTER_HISTORICAL_PRICE_METADATA.read_text(encoding="utf-8")
    )
    hf_architecture = json.loads(HF_ARCHITECTURE_AUDIT.read_text(encoding="utf-8"))
    no_cot_exact_dates = json.loads(
        NO_COT_EXACT_DATE_AUDIT.read_text(encoding="utf-8")
    )
    no_cot_architecture = json.loads(
        NO_COT_ARCHITECTURE_AUDIT.read_text(encoding="utf-8")
    )
    frontier_primary = json.loads(
        FRONTIER_PRIMARY_AUDIT.read_text(encoding="utf-8")
    )
    metr_primary = json.loads(METR_PRIMARY_AUDIT.read_text(encoding="utf-8"))
    ikp = json.loads(IKP_RESULT.read_text(encoding="utf-8"))
    ikp_conditional = json.loads(IKP_CONDITIONAL_RESULT.read_text(encoding="utf-8"))
    default_weights = default_weights_from_ikp(ikp)
    if (
        not ikp_conditional["decision"]["conditional_incremental_signal_corroborated"]
        or ikp_conditional["decision"]["change_live_ikp_weight"]
    ):
        raise ValueError("IKP conditional benchmark decision does not match the live-weight policy")
    if (
        eci_reproduction["reproduction"]["input_rows"]
        != epoch_snapshot["inventory"]["component_rows"]
        or eci_reproduction["reproduction"]["input_models"]
        != epoch_snapshot["inventory"]["models"]
        or eci_reproduction["reproduction"]["input_benchmarks"]
        != epoch_snapshot["inventory"]["benchmarks"]
    ):
        raise ValueError("ECI reproduction inventory disagrees with the Epoch snapshot manifest")
    aa_primary_metadata_overrides = aa_inference["data_audit"].get(
        "primary_metadata_overrides", []
    )
    aa_calibration_overrides = json.loads(
        AA_CALIBRATION_OVERRIDES.read_text(encoding="utf-8")
    )
    expected_override_ids = sorted(
        row["override_id"] for row in aa_calibration_overrides.get("overrides", [])
    )
    actual_override_ids = sorted(
        row["override_id"] for row in aa_primary_metadata_overrides
    )
    if (
        aa_inference["data_audit"]["raw_models"] != aa_detailed_metadata["models"]
        or aa_calibration_overrides.get("schema_version") != "1.0"
        or aa_calibration_overrides.get("snapshot_date") != "2026-07-31"
        or not expected_override_ids
        or len(set(expected_override_ids)) != len(expected_override_ids)
        or actual_override_ids != expected_override_ids
        or aa_inference["data_audit"]["open_weight_parameter_score_date_configurations"]
        != aa_detailed_metadata["open_weight_parameter_score_date_rows"]
        + len(aa_primary_metadata_overrides)
    ):
        raise ValueError(
            "AA calibration view does not reconcile to the July 31 raw snapshot plus primary-source overrides"
        )
    k3_transport = active_transport["kimi_k3_external_architecture_check"]
    if (
        not math.isclose(k3_transport["k3_disclosed_total_b"], k3["total_parameters_b_exact"])
        or not math.isclose(k3_transport["k3_disclosed_active_b"], k3["activated_parameters_b_exact"])
    ):
        raise ValueError("Active-parameter audit disagrees with official K3 counts")
    ikp_fable_t = (
        ikp["target_signal"]["fable"]["strict_open_only_release_and_vendor_holdout"]
        ["mean"]["estimates"]["forward_inverse"]["estimated_b"]
        / 1000
    )
    final = load_workbook(FINAL_WORKBOOK, data_only=True, read_only=True)
    branch = load_workbook(BRANCH_WORKBOOK, data_only=True, read_only=True)
    frontier = rows_by_header(final, "Frontier Estimates", 5, "Model / base")
    horizon = rows_by_header(final, "Horizon Estimates", 5, "Model / base")
    final_rows = rows_by_header(final, "Final Ensemble", 5, "Model / base")
    revised = rows_by_header(branch, "Revised Estimates", 5, "Model")
    forecasts = active_forecasts()
    crowd_by_model = {}
    for row in forecasts:
        crowd_by_model.setdefault(row["model"], []).append(row)

    branch_aliases = {
        "Claude Fable 5": ["Claude Fable 5"],
        "GPT-5.6 Sol": ["GPT-5.6 Sol"],
        OPUS_5_MODEL: [OPUS_5_MODEL],
        "Kimi K3": ["Kimi K3"],
        "Claude Opus 4.7 / 4.8 shared base": ["Claude Opus 4.7", "Claude Opus 4.8"],
        "GPT-5.5": ["GPT-5.5"],
        "GPT-5.6 Terra": ["GPT-5.6 Terra"],
        "Claude Sonnet 5": ["Claude Sonnet 5"],
        "GPT-5.6 Luna": ["GPT-5.6 Luna"],
        "Grok 4.5": ["Grok 4.5"],
    }

    model_order = list(final_rows)
    models = []
    for name in model_order:
        final_row = final_rows[name]
        horizon_row = horizon[name]
        frontier_row = frontier[name]
        branch_rows = [revised[alias] for alias in branch_aliases[name]]
        aa_values = [float(row["AA direct, K3-anchored (B)"]) / 1000 for row in branch_rows if row["AA direct, K3-anchored (B)"] is not None]
        eci_values = [float(row["ECI 60/40 blend (B)"]) / 1000 for row in branch_rows if row["ECI 60/40 blend (B)"] is not None]
        crowd_rows = crowd_by_model.get(name, [])
        crowd_points = [crowd_point(row) for row in crowd_rows]
        locked = final_row["Final status"] == "Disclosed anchor"
        factors = {
            "aa": geomean(aa_values) if aa_values else None,
            "eci": geomean(eci_values) if eci_values else None,
            "price": float(frontier_row["LOBO price-implied (T)"]) if frontier_row["LOBO price-implied (T)"] is not None else None,
            "horizon": float(horizon_row["Horizon prior (T)"]),
            "compute": float(horizon_row.get("Compute-structured prior (T)", horizon_row.get("Compute prior (T)"))),
            "ikp": ikp_fable_t if name == "Claude Fable 5" else None,
            "crowd": geomean(crowd_points) if crowd_points and name in CROWD_ENSEMBLE_MODELS else None,
        }
        model = {
            "id": name.lower().replace(" / ", "-").replace(" ", "-").replace(".", "").replace("/", "-"),
            "name": "Claude Opus 4.8" if name == "Claude Opus 4.7 / 4.8 shared base" else name,
            "shortName": {
                "Claude Fable 5": "Claude Fable 5",
                "GPT-5.6 Sol": "GPT-5.6 Sol",
                OPUS_5_MODEL: "Claude Opus 5",
                "Claude Opus 4.7 / 4.8 shared base": "Claude Opus 4.8",
            }.get(name, name),
            "provider": frontier_row["Provider"],
            "releaseDate": frontier_row["Release"],
            "aaScore": frontier_row["AA v4.1"],
            "eciScore": frontier_row["ECI"],
            "eciCi90": (
                [opus_5_epoch["eci_ci_low"], opus_5_epoch["eci_ci_high"]]
                if name == OPUS_5_MODEL
                else None
            ),
            "aaConfiguration": opus_5_aa["configuration"] if name == OPUS_5_MODEL else None,
            "aaFallbackModel": opus_5_aa["fallback_model"] if name == OPUS_5_MODEL else None,
            "lockedAnchor": locked,
            "disclosedT": float(final_row["Final forecast (T)"]) if locked else None,
            "currentEvidenceT": float(final_row["Evidence model (T)"]),
            "currentFinalT": float(final_row["Final forecast (T)"]),
            "factors": factors,
            "crowd": {
                "n": len(crowd_rows),
                "pooled": name in CROWD_ENSEMBLE_MODELS,
                "contributors": [row["contributor"] for row in crowd_rows],
                "forecasts": [row["forecast_text"] for row in crowd_rows],
            },
            "methodNote": horizon_row["Method note"],
        }
        if name == OPUS_5_MODEL:
            if model["releaseDate"] != opus_5_identity["release_date"]:
                raise ValueError(
                    f"Opus 5 release-date mismatch: {model['releaseDate']}"
                )
            if not math.isclose(
                float(model["aaScore"]),
                float(opus_5_aa["score"]),
                rel_tol=0,
                abs_tol=1e-12,
            ):
                raise ValueError(f"Opus 5 AA-score mismatch: {model['aaScore']}")
            if not math.isclose(
                float(model["eciScore"]),
                float(opus_5_epoch["eci_exact"]),
                rel_tol=0,
                abs_tol=1e-12,
            ):
                raise ValueError(f"Opus 5 ECI-score mismatch: {model['eciScore']}")
            if model["lockedAnchor"] or model["disclosedT"] is not None:
                raise ValueError("Opus 5 must remain an undisclosed regression target")
            if model["factors"]["crowd"] is not None or model["factors"]["ikp"] is not None:
                raise ValueError("Opus 5 has no crowd or IKP model-level measurement")
        recomputed = forecast_from_weights(model, default_weights)
        if not math.isclose(recomputed, model["currentFinalT"], rel_tol=0, abs_tol=1e-9):
            raise ValueError(f"Default site mix disagrees with workbook for {name}: {recomputed} vs {model['currentFinalT']}")
        models.append(model)

    data = {
        "schemaVersion": 1,
        "snapshotDate": "2026-07-31",
        "title": "Frontier estimates",
        "unit": "trillion total base parameters",
        "defaultWeights": default_weights,
        "factors": [
            {"id": "aa", "label": "Artificial Analysis Intelligence Index", "shortLabel": "Artificial Analysis Intelligence Index", "description": "K3-anchored, family-balanced direct parameter regression."},
            {"id": "eci", "label": "Epoch Capabilities Index", "shortLabel": "Epoch Capabilities Index", "description": "60/40 blend of no-date and exact-date direct regressions."},
            {"id": "price", "label": "API price", "shortLabel": "Price", "description": "API price should correlate with model size because larger models generally cost more to serve; we use it as a low-weight cross-check, not a primary signal."},
            {"id": "horizon", "label": "No-CoT Time Horizon", "shortLabel": "No-CoT Time Horizon", "description": "No-CoT measures how long a task a model can complete without explicit chain-of-thought reasoning; we use that task horizon as our strongest capability-to-parameter scaling signal."},
            {"id": "compute", "label": "Compute-structured AA", "shortLabel": "Compute", "description": "Training compute should rise with model scale; we map AA and release date through Epoch's compute estimates and use the result as a low-weight structural cross-check."},
            {"id": "ikp", "label": "Knowledge capacity", "shortLabel": "IKP", "description": f"Incompressible Knowledge Probes: a direct Fable factual-capacity sensitivity. Its live weight is {100 * ikp['decision']['incremental_evidence_weight']:.0f}% because promotion requires every predeclared validation gate to pass."},
            {"id": "crowd", "label": "Public estimate", "shortLabel": "Public estimate", "description": "Geometric mean of parameter-count forecasts from a 20-person poll of researchers and engineers; used only for Fable and Sol."},
        ],
        "presets": [
            {"id": "current", "label": "Current posterior", "weights": default_weights},
            {"id": "regression", "label": "Regression only", "weights": {"aa": 50, "eci": 50, "price": 0, "horizon": 0, "compute": 0, "ikp": 0, "crowd": 0}},
            {"id": "capability", "label": "Capability heavy", "weights": {"aa": 15, "eci": 15, "price": 0, "horizon": 60, "compute": 10, "ikp": 0, "crowd": 0}},
            {"id": "crowd", "label": "Crowd heavy", "weights": {"aa": 5, "eci": 5, "price": 0, "horizon": 15, "compute": 5, "ikp": 0, "crowd": 70}},
        ],
        "models": models,
        "humanForecasts": {
            "activeRecords": len(forecasts),
            "contributors": len({row["contributor"] for row in forecasts}),
        },
        "method": {
            "combination": "Factors are combined in log space. Missing signals are renormalized inside the evidence block, so adding an Fable-only measurement does not silently change Sol's 50% crowd weight.",
            "anchors": "Disclosed Kimi K3 and Grok 4.5 totals remain fixed under every weighting scenario.",
            "currentMix": f"The default reproduces the audited workbook exactly: 50% crowd for Fable/Sol. IKP receives {100 * ikp['decision']['incremental_evidence_weight']:.0f}% inside Fable's evidence half under the live audit decision, so the AA/ECI/price/no-CoT/compute block receives the remaining {100 * (1 - ikp['decision']['incremental_evidence_weight']):.0f}%. Sol has no IKP observation.",
            "ikp": f"IKP direct-capacity audit: {ikp['source_inventory']['calibration_configurations']} source configurations collapse to {ikp['source_inventory']['calibration_weight_bases']} distinct weight bases. On {ikp['incremental_overlap']['models']} exact models across {ikp['incremental_overlap']['families']} families, the existing median error is {ikp['incremental_overlap']['existing']['median_multiplicative_error']:.2f}× versus {ikp['incremental_overlap']['ikp']['median_multiplicative_error']:.2f}× for IKP. A fixed 10% diagnostic blend improves both bootstrap checks, but the later chronological subset has {ikp['incremental_overlap']['chronological_fixed_weight_subset']['models']} models versus the predeclared minimum of {ikp['decision']['evidence_gates']['minimum_chronological_subset_models']}; promotion therefore fails and live IKP weight is {100 * ikp['decision']['incremental_evidence_weight']:.0f}%. The strict pre-Fable, Anthropic-excluded sensitivity is {ikp_fable_t:.1f}T. The separate conditional test supports GPQA and MMLU but cannot override the failed primary coverage gate. Sol remains unobserved.",
            "noCotDates": f"No-CoT date audit: all {no_cot_exact_dates['inventory']['no_cot_models']} paper checkpoints now have day-level dates, with {no_cot_exact_dates['inventory']['explicit_date_only_overrides']} date-only overrides and {no_cot_exact_dates['inventory']['parameter_identities_added_by_overrides']} parameter identities added. Applying the method-matched exact-date/month-date Pareto slope ratio changes the published time law from {no_cot_exact_dates['time_horizon']['adjusted_reported_law']['paper_reported_point_days']:.1f} to {no_cot_exact_dates['time_horizon']['adjusted_reported_law']['adjusted_point_days']:.1f} days and the token law from {no_cot_exact_dates['token_horizon']['adjusted_reported_law']['paper_reported_point_days']:.1f} to {no_cot_exact_dates['token_horizon']['adjusted_reported_law']['adjusted_point_days']:.1f} days. The branch weight remains unchanged.",
            "noCotArchitecture": f"No-CoT architecture audit: the paper's 8.1x MoE relationship is exactly reproducible at {no_cot_architecture['paper_relationship_reproduction']['moe']['deterministic_bootstrap_median_reproduction']:.2f}x on {no_cot_architecture['paper_relationship_reproduction']['moe']['pareto_n']} MoE Pareto models. It remains descriptive rather than a promoted inverse parameter estimator: the separately learned architecture slope worsens strictly chronological developer-held-out error, while the training-fold-only Pareto comparison has a 90% family interval crossing zero. The pooled live mapping and headline forecasts remain unchanged.",
            "primaryEvidence": f"First-party frontier audit: OpenAI reports a {frontier_primary['official_measurements']['gpt_5_6_sol_nocot_minutes']:.1f}-minute no-CoT horizon for Sol versus {frontier_primary['official_measurements']['gpt_5_5_comparator_minutes']:.1f} minutes for GPT-5.5. The fully identified chronological developer-holdout has only {frontier_primary['inventory']['chronological_developer_holdout_predictions']} predictions across {frontier_primary['inventory']['heldout_developers']} held-out developers, and the developer-balanced interval crosses zero. Plausible mappings of the same point span {frontier_primary['sol_mapping_sensitivity']['direct_model_level_horizon_regression_t']:.1f}T to {frontier_primary['sol_mapping_sensitivity']['gpt_5_5_suite_rebased_pooled_elasticity_t']:.1f}T, so its incremental live weight is 0%. Anthropic's system card independently establishes that Fable 5 and Mythos 5 share underlying weights; Opus 4.8 fallback remains a serving-system caveat, not a second base.",
            "metrPrimary": f"METR primary-source audit: all {metr_primary['official_asset']['result_rows']} official model rows, {metr_primary['losslessness']['full_scaffold_entries']} scaffold entries, confidence intervals, source versions, and trend-law fields are retained directly from METR's YAML. The earlier copy matches {metr_primary['legacy_exact_crosscheck']['exact_rows']}/{metr_primary['legacy_exact_crosscheck']['legacy_rows']} rows across every common field with {metr_primary['legacy_exact_crosscheck']['mismatch_count']} mismatches. This strengthens provenance and losslessness without changing the numerical forecast.",
            "operational": f"Operational cross-check: OpenRouter price remains weakly predictive; all {openrouter_official['model_count_requested']} documented endpoint-API requests succeeded and {100 * openrouter_official['official_price_row_exact_share']:.1f}% of official price rows reconcile exactly to the frontend extraction. The independent AA audit confirms provider-median price on 87 checkpoints and shows 0.83 cross-source price rank correlation, while first-party frontier price, tok/s, and latency add no supported weight.",
            "operationalTemporal": f"OpenRouter temporal cross-check: every refresh is archived byte-for-byte, with default, priority, and flex throughput retained separately. Across {openrouter_temporal['inventory']['immutable_snapshots']} immutable snapshots and {openrouter_temporal['inventory']['history_daily_rows']:,} daily rows, default-tier tok/s still fails after price. In a separate {openrouter_request_weighted['inventory']['complete_checkpoints']}-checkpoint request-supported audit, no throughput, latency, joint, or tail-spread candidate passes the family-interval plus chronological-direction gate. Tok/s and latency remain at 0%.",
            "eciReproduction": f"ECI reproduction cross-check: the pinned official Epoch implementation reproduces all {eci_reproduction['reproduction']['reproduced_models']} scores from {eci_reproduction['reproduction']['input_rows']:,} benchmark rows spanning {eci_reproduction['reproduction']['input_benchmarks']} benchmarks. All {eci_reproduction['published_score_crosscheck']['published_nonblank_scores']} published nonblank scores agree within two-decimal display rounding. The audit preserves {eci_reproduction['release_date_crosscheck']['published_vs_input_date_disagreements']} input-date/release-date disagreements and uses the published checkpoint release date for chronological regression.",
            "eciFit": f"ECI fit tournament: {eci_fit['inventory']['historical_snapshots']} hash-pinned vintages yield {eci_fit['inventory']['first_observed_outer_targets']} first-observed tests, including {eci_fit['inventory']['interval_prospective_targets']} interval-prospective Kimi checkpoints. The flexible ridge form cuts the 25-row selection median error, but fails the inverse-CI prospective veto and spans {eci_fit['decision']['frontier_sensitivity']['Claude Fable 5']['max_over_min']:.2f}x for Fable and {eci_fit['decision']['frontier_sensitivity']['GPT-5.6 Sol']['max_over_min']:.2f}x for Sol across weighting/base-collapse sensitivities. Because both targets are 8.8–10.2 ECI points beyond the open-weight calibration range, the live linear blend is retained.",
            "k3Efficiency": f"K3 efficiency stress test: the statement that a target is at least as parameter-efficient as disclosed 2.780T Kimi K3 is represented by a center-preserving upper-tail projection, not another point-estimate factor or literal conditioning. K3-relative AA and retained log-linear ECI mappings are judgmentally pooled in log space, with target and K3 dates held equal. The default applies winsorization to {100 * k3_efficiency['decision']['default_projection_strength']:.0f}% of upper-tail draws; a below-center reference is overridden by the center. Point centers, lower tails, the exact 50% Fable/Sol crowd blend, and the rejected nonlinear ECI form remain unchanged.",
            "activePrice": f"Active-price cross-check: {active_price['inventory']['active_parameter_matches']} labels—{active_price['inventory']['aa_disclosed_active_parameter_matches']} disclosed active counts plus {active_price['inventory']['dense_config_active_equals_total_controls']} primary-config dense controls—produce {active_price['inventory']['release_ordered_predictions']} release-ordered developer-held-out tests. The config audit parsed {hf_architecture['successful_json_configs']} repositories and retains {hf_architecture['architecture_classification_counts']['unavailable']} gated repositories as unresolved. Active-to-total MoE transport improves point metrics on 16 high-sparsity models, but its seven-developer interval crosses zero and target prices extrapolate beyond training support. It remains a 0%-weight sensitivity.",
            "historicalPrice": f"Historical-price cross-check: a hash-pinned ledger rebuilt from {historical_price_metadata['source']['full_git_history_rebuild_snapshot_count']:,} committed official OpenRouter catalog snapshots retains {historical_price['inventory']['historical_ledger_models']} model IDs and {historical_price['inventory']['historical_change_points']:,} price changes from 21 Sep 2024. All {historical_price['inventory']['calibration_checkpoints_audited']} calibration aliases match exactly; {historical_price['inventory']['eligible_total_rows_by_window']['1']} total-size and {historical_price['inventory']['eligible_active_rows_by_window']['1']} active-size labels have eligible first-day prices. Launch-vintage price beats date alone in every predeclared 1–90 day window, but no active-size window adds a developer-cluster interval wholly below zero beyond AA score plus date. The first-day K3-anchored sensitivities are {historical_price['headline_crosscheck']['fable_k3_anchored_score_date_first_day_price_t']:.1f}T for Fable and {historical_price['headline_crosscheck']['sol_k3_anchored_score_date_first_day_price_t']:.1f}T for Sol; incremental weight remains 0%.",
            "components": f"ECI component cross-check: no individual benchmark survives familywise correction. A separately nested multivariate model moves the {eci_multivariate['inventory']['outer_predictions']}-fold median error from {eci_multivariate['backtest']['total']['all']['baseline']['median_multiplicative_error']:.2f}× to {eci_multivariate['backtest']['total']['all']['candidate']['median_multiplicative_error']:.2f}×, but its family interval crosses zero; the narrow-ECI-CI replication also crosses zero. The split is aggregate-score uncertainty, not parameter-disclosure status. Incremental component weight remains 0%.",
            "posttraining": f"Post-training lineage cross-check: {posttraining_lineage['inventory']['base_model_links']} Epoch base links narrow to {posttraining_lineage['inventory']['candidate_open_language_same_parameter_links']} candidate open-language, unchanged-size links and {posttraining_lineage['inventory']['admitted_measured_lineage_edges']} measured edges across {posttraining_lineage['inventory']['admitted_measured_lineage_bases']} bases. Later descendants have a median {posttraining_lineage['lineage_backtests']['eci']['median_implied_child_over_parent_parameter_ratio']:.2f}× ECI-implied scale ratio despite identical size; ECI collapse is inconclusive and AA has only {posttraining_lineage['inventory']['aa_prediction_edges']} prediction edges. Opus 4.5–4.8 and GPT-5–5.5 remain user-supplied shared-lineage assumptions, not public disclosures. Incremental correction weight remains 0%.",
            "aaExpansion": f"AA expansion cross-check: {aa_expanded['data_audit']['expanded_unique_models']} unique open-weight checkpoints were tested with strict chronological developer holdout. Broad tail error improves, but the {aa_expanded['backtest']['scopes']['frontier_like']['n']} frontier-like tests are neutral, so the expanded-panel AA weight remains 0%.",
            "aaInference": f"AA inference-budget cross-check: the complete {aa_inference['data_audit']['raw_models']}-record snapshot yields {aa_inference['data_audit']['unique_checkpoint_groups']} open-weight checkpoints and {aa_inference['reasoning_configuration_pairs']['all']['pairs']} same-checkpoint reasoning pairs. The measured equal-creator uplift is about six AA points, but token-budget and reasoning-standardized frontier intervals cross zero, so both incremental weights remain 0%.",
            "activeTransport": f"Active-parameter cross-check: K3 discloses exactly {k3['total_parameters_b_exact'] / 1000:.2f}T total and {k3['activated_parameters_b_exact']:.1f}B active parameters. The held-out transport comparison remains inconclusive; K3-anchored sensitivities are 3.8T for Fable and 3.2T for Sol, so incremental weight remains 0%.",
            "computeDependency": "K3 has one explicitly speculative Epoch training-compute estimate, while no live target has a primary-source compute disclosure. The branch otherwise predicts compute from AA and date, so it remains correlated regularization rather than independent target evidence.",
        },
        "operationalSignals": {
            "snapshotDate": openrouter["snapshot_date"],
            "epochCalibrationCheckpoints": openrouter["data_audit"]["unique_epoch_calibration_checkpoints"],
            "developerFamilies": openrouter["data_audit"]["calibration_families"],
            "priceMedianHeldoutErrorX": openrouter["heldout_metrics"]["family"]["date_price"]["median_multiplicative_error"],
            "tokSIncrementalWeight": openrouter["conclusion"]["recommended_incremental_tok_s_weight_in_live_ensemble"],
        },
        "noCotDateSignals": {
            "models": no_cot_exact_dates["inventory"]["no_cot_models"],
            "exactDates": no_cot_exact_dates["inventory"]["models_with_day_level_dates"],
            "remainingMonthOnly": no_cot_exact_dates["inventory"]["models_remaining_month_only"],
            "dateOnlyOverrides": no_cot_exact_dates["inventory"]["explicit_date_only_overrides"],
            "parameterIdentitiesAdded": no_cot_exact_dates["inventory"]["parameter_identities_added_by_overrides"],
            "paperTimeDays": no_cot_exact_dates["time_horizon"]["adjusted_reported_law"]["paper_reported_point_days"],
            "adjustedTimeDays": no_cot_exact_dates["time_horizon"]["adjusted_reported_law"]["adjusted_point_days"],
            "paperTokenDays": no_cot_exact_dates["token_horizon"]["adjusted_reported_law"]["paper_reported_point_days"],
            "adjustedTokenDays": no_cot_exact_dates["token_horizon"]["adjusted_reported_law"]["adjusted_point_days"],
            "weightChanged": no_cot_exact_dates["decision"]["change_live_no_cot_weight"],
        },
        "noCotArchitectureSignals": {
            "models": no_cot_architecture["inventory"]["models"],
            "families": no_cot_architecture["inventory"]["families"],
            "denseModels": no_cot_architecture["inventory"]["dense_models"],
            "moeModels": no_cot_architecture["inventory"]["moe_models"],
            "moeParetoModels": no_cot_architecture["paper_relationship_reproduction"]["moe"]["pareto_n"],
            "moeFactor": no_cot_architecture["paper_relationship_reproduction"]["moe"]["deterministic_bootstrap_median_reproduction"],
            "directDelta": no_cot_architecture["paired_comparisons"]["chronological_developer_holdout"]["direct_architecture_minus_pooled"]["observed_delta"],
            "directCi90": no_cot_architecture["paired_comparisons"]["chronological_developer_holdout"]["direct_architecture_minus_pooled"]["ci_90"],
            "paretoDelta": no_cot_architecture["paired_comparisons"]["chronological_developer_holdout"]["training_pareto_architecture_minus_pooled"]["observed_delta"],
            "paretoCi90": no_cot_architecture["paired_comparisons"]["chronological_developer_holdout"]["training_pareto_architecture_minus_pooled"]["ci_90"],
            "replacePooled": no_cot_architecture["decision"]["replace_pooled_live_elasticity_with_moe_specific"],
        },
        "frontierPrimarySignals": {
            "directSolMinutes": frontier_primary["official_measurements"]["gpt_5_6_sol_nocot_minutes"],
            "gpt55ComparatorMinutes": frontier_primary["official_measurements"]["gpt_5_5_comparator_minutes"],
            "currentProjectedSolMinutes": frontier_primary["current_mapping"]["projected_sol_horizon_minutes"],
            "currentProjectedSolPriorT": frontier_primary["current_mapping"]["projected_sol_horizon_prior_t"],
            "modelLevelSolT": frontier_primary["sol_mapping_sensitivity"]["direct_model_level_horizon_regression_t"],
            "pooledElasticitySolT": frontier_primary["sol_mapping_sensitivity"]["direct_pooled_paper_elasticity_t"],
            "moeElasticitySolT": frontier_primary["sol_mapping_sensitivity"]["direct_moe_paper_elasticity_t"],
            "rebasedPooledSolT": frontier_primary["sol_mapping_sensitivity"]["gpt_5_5_suite_rebased_pooled_elasticity_t"],
            "methodSpread": frontier_primary["sol_mapping_sensitivity"]["nonbaseline_method_max_over_min"],
            "heldoutPredictions": frontier_primary["inventory"]["chronological_developer_holdout_predictions"],
            "heldoutDevelopers": frontier_primary["inventory"]["heldout_developers"],
            "horizonBetterProbability": frontier_primary["heldout_backtest"]["incremental_bootstrap"]["bootstrap_probability_horizon_better"],
            "incrementalCi90": frontier_primary["heldout_backtest"]["incremental_bootstrap"]["ci_90"],
            "incrementalWeight": frontier_primary["decision"]["incremental_live_weight"],
            "fableMythosSharedWeights": frontier_primary["decision"]["apply_fable_mythos_shared_weight_identity"],
            "opusFallbackIsSharedBase": frontier_primary["decision"]["treat_opus_fallback_as_shared_base"],
        },
        "metrPrimarySignals": {
            "officialRows": metr_primary["official_asset"]["result_rows"],
            "uniqueSourceIds": metr_primary["official_asset"]["unique_source_ids"],
            "fullScaffoldEntries": metr_primary["losslessness"]["full_scaffold_entries"],
            "legacyExactRows": metr_primary["legacy_exact_crosscheck"]["exact_rows"],
            "legacyRows": metr_primary["legacy_exact_crosscheck"]["legacy_rows"],
            "mismatchCount": metr_primary["legacy_exact_crosscheck"]["mismatch_count"],
            "from2023DoublingDays": float(metr_primary["trend"]["from_2023_on_point_estimate_days"]),
            "from2023CiLowDays": float(metr_primary["trend"]["from_2023_on_ci_low_days"]),
            "from2023CiHighDays": float(metr_primary["trend"]["from_2023_on_ci_high_days"]),
        },
        "ikpSignals": {
            "calibrationConfigurations": ikp["source_inventory"]["calibration_configurations"],
            "calibrationWeightBases": ikp["source_inventory"]["calibration_weight_bases"],
            "servingVariantsCollapsed": ikp["source_inventory"]["serving_variants_collapsed"],
            "strictPredictionRows": sum(
                metrics[scope]["n"]
                for policy in ikp["heldout_metrics"].values()
                for metrics in policy.values()
                for scope in ("all",)
            ),
            "overlapModels": ikp["incremental_overlap"]["models"],
            "overlapFamilies": ikp["incremental_overlap"]["families"],
            "existingMedianX": ikp["incremental_overlap"]["existing"]["median_multiplicative_error"],
            "ikpMedianX": ikp["incremental_overlap"]["ikp"]["median_multiplicative_error"],
            "blendMedianX": ikp["incremental_overlap"]["blend_10pct"]["median_multiplicative_error"],
            "fullBootstrapCi90": ikp["incremental_overlap"]["family_bootstrap"]["ci_90"],
            "chronologicalSubsetModels": ikp["incremental_overlap"]["chronological_fixed_weight_subset"]["models"],
            "chronologicalSubsetFamilies": ikp["incremental_overlap"]["chronological_fixed_weight_subset"]["families"],
            "chronologicalBootstrapCi90": ikp["incremental_overlap"]["chronological_fixed_weight_subset"]["family_bootstrap"]["ci_90"],
            "familiesImproved": ikp["incremental_overlap"]["families_improved"],
            "signedErrorCorrelation": ikp["incremental_overlap"]["signed_error_correlation_existing_vs_ikp"],
            "fableStrictT": ikp_fable_t,
            "fablePublishedT": ikp["target_signal"]["fable"]["published_lambda0_estimate_b"] / 1000,
            "fableStrictFormRangeT": [
                ikp["target_signal"]["fable"]["strict_open_only_model_form_min_b"] / 1000,
                ikp["target_signal"]["fable"]["strict_open_only_model_form_max_b"] / 1000,
            ],
            "fableSourcePi90T": [
                value / 1000
                for value in ikp["target_signal"]["fable"]["published_pi90_b"]
            ],
            "fableRefusalRate": ikp["target_signal"]["fable"]["refusal_rate"],
            "promoted": ikp["decision"]["promote_incremental_ikp_weight"],
            "evidenceWeight": ikp["decision"]["incremental_evidence_weight"],
            "finalWeight": ikp["decision"]["incremental_final_weight_when_crowd_is_50pct"],
            "minimumChronologicalSubsetModels": ikp["decision"]["evidence_gates"]["minimum_chronological_subset_models"],
            "solObserved": ikp["target_signal"]["sol"]["observed"],
            "conditionalGpqaModels": ikp_conditional["heldout_results"]["gpqa_diamond"]["strict_prediction_models"],
            "conditionalGpqaVendors": ikp_conditional["heldout_results"]["gpqa_diamond"]["strict_prediction_vendors"],
            "conditionalGpqaPassingSpecifications": ikp_conditional["heldout_results"]["gpqa_diamond"]["passing_specifications"],
            "conditionalMmluModels": ikp_conditional["heldout_results"]["mmlu"]["strict_prediction_models"],
            "conditionalMmluPassingSpecifications": ikp_conditional["heldout_results"]["mmlu"]["passing_specifications"],
            "conditionalMmluProPassingSpecifications": ikp_conditional["heldout_results"]["mmlu_pro"]["passing_specifications"],
            "conditionalSignalCorroborated": ikp_conditional["decision"]["conditional_incremental_signal_corroborated"],
            "conditionalWeightChanged": ikp_conditional["decision"]["change_live_ikp_weight"],
            "staleUpstreamNarrativeClaims": ikp_conditional["upstream_reproduction"]["narrative_summary_audit"]["stale_claim_count"],
        },
        "openRouterTemporalSignals": {
            "immutableSnapshots": openrouter_temporal["inventory"]["immutable_snapshots"],
            "currentDailyRows": openrouter_temporal["inventory"]["current_daily_rows"],
            "historyDailyRows": openrouter_temporal["inventory"]["history_daily_rows"],
            "serviceTierRows": openrouter_temporal["inventory"]["service_tier_row_counts"],
            "multiTierEndpointModels": openrouter_temporal["inventory"]["endpoint_models_with_multiple_service_tiers"],
            "unmatchedEndpointModels": openrouter_temporal["inventory"]["current_daily_unmatched_endpoint_models"],
            "familyPriceMae": openrouter_temporal["corrected_default_tier_backtest"]["family"]["date_price"]["mean_absolute_log10_error"],
            "familyPriceTokSMae": openrouter_temporal["corrected_default_tier_backtest"]["family"]["date_price_plus_normalized_tok_s"]["mean_absolute_log10_error"],
            "chronologicalPriceMae": openrouter_temporal["corrected_default_tier_backtest"]["chronological_family"]["date_price"]["mean_absolute_log10_error"],
            "chronologicalPriceTokSMae": openrouter_temporal["corrected_default_tier_backtest"]["chronological_family"]["date_price_plus_normalized_tok_s"]["mean_absolute_log10_error"],
            "medianModelWithinWeekMaxOverMin": openrouter_temporal["temporal_stability"]["model_daily_provider_median_with_at_least_four_days_max_over_min"]["median"],
            "p90ModelWithinWeekMaxOverMin": openrouter_temporal["temporal_stability"]["model_daily_provider_median_with_at_least_four_days_max_over_min"]["p90"],
            "focalModels": openrouter_temporal["temporal_stability"]["focal_models"],
            "tokSIncrementalWeight": openrouter_temporal["decision"]["recommended_incremental_tok_s_weight"],
        },
        "openRouterOfficialSignals": {
            "modelRequests": openrouter_official["model_count_requested"],
            "successfulModelRequests": openrouter_official["model_count_succeeded"],
            "officialEndpointRows": openrouter_official["official_endpoint_rows"],
            "frontendEndpointTierRows": openrouter_official["frontend_endpoint_tier_rows"],
            "officialPriceExactShare": openrouter_official["official_price_row_exact_share"],
            "comparisonGroupCounts": openrouter_official["comparison_group_counts"],
            "endpointTierRows": openrouter_collection["endpoint_tier_row_count"],
            "endpointTierServiceTierCounts": openrouter_collection["endpoint_tier_service_tier_counts"],
            "highContextPriceRows": openrouter_collection["endpoint_tier_rows_with_high_context_price"],
            "focalModels": openrouter_official["focal_models"],
            "incrementalForecastWeight": 0.0,
        },
        "openRouterRequestWeightedSignals": {
            "completeCheckpoints": openrouter_request_weighted["inventory"]["complete_checkpoints"],
            "completeFamilies": openrouter_request_weighted["inventory"]["complete_families"],
            "minimumRequests": 100,
            "supportedCandidates": openrouter_request_weighted["supported_candidates"],
            "familyPriceMedianX": openrouter_request_weighted["heldout_metrics"]["family"]["date_price"]["median_multiplicative_error"],
            "familyLatencyMedianX": openrouter_request_weighted["heldout_metrics"]["family"]["date_price_p50_latency_supported"]["median_multiplicative_error"],
            "familyLatencyCi90": next(
                row["ci_90"]
                for row in openrouter_request_weighted["paired_family_bootstraps"]
                if row["mode"] == "family"
                and row["candidate"] == "date_price_p50_latency_supported"
            ),
            "incrementalWeight": openrouter_request_weighted["decision"]["incremental_live_weight"],
        },
        "eciReproductionSignals": {
            "officialCommit": eci_reproduction["reproduction"]["official_commit"],
            "inputRows": eci_reproduction["reproduction"]["input_rows"],
            "inputModels": eci_reproduction["reproduction"]["input_models"],
            "inputBenchmarks": eci_reproduction["reproduction"]["input_benchmarks"],
            "epochAllModelRows": epoch_snapshot["inventory"]["all_model_rows"],
            "reproducedModels": eci_reproduction["reproduction"]["reproduced_models"],
            "publishedScoreMatches": eci_reproduction["published_score_crosscheck"]["published_nonblank_scores"],
            "maximumPublishedScoreDifference": eci_reproduction["published_score_crosscheck"]["maximum_absolute_score_difference"],
            "publishedReleaseDatesUsed": eci_reproduction["release_date_crosscheck"]["published_release_dates_used"],
            "canonicalInputDateFallbacks": eci_reproduction["release_date_crosscheck"]["canonical_input_date_fallbacks"],
            "preservedDateDisagreements": eci_reproduction["release_date_crosscheck"]["published_vs_input_date_disagreements"],
        },
        "eciFitSignals": {
            "historicalSnapshots": eci_fit["inventory"]["historical_snapshots"],
            "historicalScoreRows": eci_fit["inventory"]["historical_score_rows"],
            "firstObservedTargets": eci_fit["inventory"]["first_observed_outer_targets"],
            "selectionTargets": eci_fit["inventory"]["selection_targets"],
            "prospectiveTargets": eci_fit["inventory"]["interval_prospective_targets"],
            "selectedCandidate": eci_fit["tournament"]["live_inverse_eci_ci"]["selected_challenger"],
            "baselineSelectionMedianX": eci_fit["tournament"]["live_inverse_eci_ci"]["candidates"]["live_60_40_blend"]["selection_first_observed_backfills"]["median_multiplicative_error"],
            "candidateSelectionMedianX": eci_fit["tournament"]["live_inverse_eci_ci"]["candidates"]["ridge_flexible"]["selection_first_observed_backfills"]["median_multiplicative_error"],
            "selectionCi90": eci_fit["tournament"]["live_inverse_eci_ci"]["selected_vs_baseline_selection_bootstrap"]["ci_90"],
            "baselineProspectiveMae": eci_fit["tournament"]["live_inverse_eci_ci"]["candidates"]["live_60_40_blend"]["interval_prospective_validation"]["mean_absolute_log10_error"],
            "candidateProspectiveMae": eci_fit["tournament"]["live_inverse_eci_ci"]["candidates"]["ridge_flexible"]["interval_prospective_validation"]["mean_absolute_log10_error"],
            "fableExtrapolationPoints": eci_fit["frontier_extrapolation"]["Claude Fable 5"],
            "solExtrapolationPoints": eci_fit["frontier_extrapolation"]["GPT-5.6 Sol"],
            "fableSensitivityRatio": eci_fit["decision"]["frontier_sensitivity"]["Claude Fable 5"]["max_over_min"],
            "solSensitivityRatio": eci_fit["decision"]["frontier_sensitivity"]["GPT-5.6 Sol"]["max_over_min"],
            "changeLiveForm": eci_fit["decision"]["change_live_eci_functional_form"],
        },
        "componentSignals": {
            "expandedParameterCheckpoints": eci_component["expanded_total_parameter_panel"]["models"],
            "exactEpochAdditions": eci_component["expanded_total_parameter_panel"]["source_counts"]["exact_epoch_open_extension"],
            "activeParameterCheckpoints": eci_component["active_parameter_component_audit"]["parameter_map_models"],
            "eligibleBenchmarks": eci_component["active_parameter_component_audit"]["eligible_comparisons"],
            "familywiseSupported": len(eci_component["active_parameter_component_audit"]["supported_after_familywise_correction"]),
            "incrementalWeight": eci_component["active_parameter_component_audit"]["decision"]["incremental_component_weight"],
            "bestUncorrectedBenchmark": eci_component["active_parameter_component_audit"]["best_uncorrected_component"]["benchmark"],
            "bestAdjustedP": eci_component["active_parameter_component_audit"]["best_uncorrected_component"]["active_familywise_one_sided_p"],
            "multivariateOuterPredictions": eci_multivariate["inventory"]["outer_predictions"],
            "multivariateFamilies": eci_multivariate["inventory"]["outer_prediction_families"],
            "multivariateBaselineMedianX": eci_multivariate["backtest"]["total"]["all"]["baseline"]["median_multiplicative_error"],
            "multivariateCandidateMedianX": eci_multivariate["backtest"]["total"]["all"]["candidate"]["median_multiplicative_error"],
            "multivariateCi90": eci_multivariate["backtest"]["total"]["all"]["paired_family_bootstrap"]["ci_90"],
            "narrowCiOnlyPredictions": eci_multivariate["inventory"]["narrow_eci_ci_only_outer_predictions"],
            "narrowCiOnlyFamilies": eci_multivariate["inventory"]["narrow_eci_ci_only_outer_prediction_families"],
            "narrowCiOnlyBaselineMedianX": eci_multivariate["narrow_eci_ci_only_training_backtest"]["total"]["baseline"]["median_multiplicative_error"],
            "narrowCiOnlyCandidateMedianX": eci_multivariate["narrow_eci_ci_only_training_backtest"]["total"]["candidate"]["median_multiplicative_error"],
            "narrowCiOnlyCi90": eci_multivariate["narrow_eci_ci_only_training_backtest"]["total"]["paired_family_bootstrap"]["ci_90"],
            "multivariateIncrementalWeight": eci_multivariate["decision"]["incremental_live_weight"],
            "targetSensitivities": eci_multivariate["target_sensitivity"],
        },
        "posttrainingSignals": {
            "epochBaseLinks": posttraining_lineage["inventory"]["base_model_links"],
            "sameParameterBothOpenLinks": posttraining_lineage["inventory"]["same_parameter_both_open_links"],
            "candidateOpenLanguageLinks": posttraining_lineage["inventory"]["candidate_open_language_same_parameter_links"],
            "measuredEdges": posttraining_lineage["inventory"]["admitted_measured_lineage_edges"],
            "measuredBases": posttraining_lineage["inventory"]["admitted_measured_lineage_bases"],
            "measuredDevelopers": posttraining_lineage["inventory"]["admitted_developers"],
            "eciPredictionEdges": posttraining_lineage["inventory"]["eci_prediction_edges"],
            "aaPredictionEdges": posttraining_lineage["inventory"]["aa_prediction_edges"],
            "noCotEdges": posttraining_lineage["inventory"]["nocot_lineage_edges"],
            "metrEdges": posttraining_lineage["inventory"]["metr_lineage_edges"],
            "sameWeightReasoningPairs": posttraining_lineage["hard_same_checkpoint_control"]["open_weight_reasoning_pairs"]["pairs"],
            "sameWeightReasoningCreators": posttraining_lineage["hard_same_checkpoint_control"]["open_weight_reasoning_pairs"]["creators"],
            "sameWeightReasoningMedianUplift": posttraining_lineage["hard_same_checkpoint_control"]["open_weight_reasoning_pairs"]["equal_creator_median_aa_uplift"],
            "eciMedianImpliedRatio": posttraining_lineage["lineage_backtests"]["eci"]["median_implied_child_over_parent_parameter_ratio"],
            "eciCollapseCi90": posttraining_lineage["lineage_backtests"]["eci"]["collapsed_vs_baseline"]["ci_90"],
            "aaCollapseCi90": posttraining_lineage["lineage_backtests"]["aa"]["collapsed_vs_baseline"]["ci_90"],
            "knowledgeMedianUplift": posttraining_lineage["component_posttraining_sensitivity"]["categories"]["knowledge"]["median_component_implied_eci_delta"],
            "otherMedianUplift": posttraining_lineage["component_posttraining_sensitivity"]["categories"]["other"]["median_component_implied_eci_delta"],
            "incrementalWeight": posttraining_lineage["decision"]["incremental_live_weight"],
            "publiclyVerifiedProprietarySharedBase": posttraining_lineage["promotion_gates"]["proprietary_shared_base_claims_publicly_verified"],
        },
        "aaExpansionSignals": {
            "currentCheckpoints": aa_expanded["data_audit"]["current_panel_models"],
            "exactEpochCheckpoints": aa_expanded["data_audit"]["exact_open_epoch_checkpoints"],
            "reconciledOverlaps": aa_expanded["data_audit"]["current_exact_overlaps"],
            "expandedCheckpoints": aa_expanded["data_audit"]["expanded_unique_models"],
            "eligiblePredictions": aa_expanded["backtest"]["eligible_predictions"],
            "frontierLikePredictions": aa_expanded["backtest"]["scopes"]["frontier_like"]["n"],
            "incrementalWeight": aa_expanded["decision"]["incremental_expanded_aa_weight"],
        },
        "aaInferenceSignals": {
            "rawModels": aa_inference["data_audit"]["raw_models"],
            "openWeightConfigurations": aa_inference["data_audit"]["open_weight_parameter_score_date_configurations"],
            "uniqueCheckpoints": aa_inference["data_audit"]["unique_checkpoint_groups"],
            "tokenCoveredCheckpoints": aa_inference["data_audit"]["token_covered_checkpoint_groups"],
            "allReasoningPairs": aa_inference["reasoning_configuration_pairs"]["all"]["pairs"],
            "exactWeightPairs": aa_inference["same_weight_reasoning_pairs"]["pairs"],
            "pairCreators": aa_inference["reasoning_configuration_pairs"]["all"]["creators"],
            "equalCreatorMedianUplift": aa_inference["reasoning_configuration_pairs"]["all"]["equal_creator_median_aa_uplift"],
            "exactEpochCrosschecks": aa_inference["data_audit"]["epoch_exact_crosschecks"],
            "metadataDisagreements": aa_inference["data_audit"]["epoch_crosschecks_with_metadata_disagreement"],
            "portableFrontierBaselineMedianX": aa_inference["reasoning_standardization_backtest"]["portable"]["scopes"]["frontier_like"]["baseline"]["median_multiplicative_error"],
            "portableFrontierCandidateMedianX": aa_inference["reasoning_standardization_backtest"]["portable"]["scopes"]["frontier_like"]["candidate"]["median_multiplicative_error"],
            "incrementalTokenWeight": aa_inference["decision"]["incremental_inference_budget_weight"],
            "incrementalReasoningWeight": aa_inference["decision"]["incremental_reasoning_standardization_weight"],
        },
        "aaOperationalSignals": {
            "priceCheckpoints": aa_operational["data_audit"]["coverage"]["blended_price"]["checkpoints"],
            "speedCheckpoints": aa_operational["data_audit"]["coverage"]["output_speed"]["checkpoints"],
            "providerMedianPriceCheckpoints": aa_operational["backtests"]["price_provider_median"]["eligible_checkpoints"],
            "firstPartyPriceCheckpoints": aa_operational["backtests"]["price_first_party"]["eligible_checkpoints"],
            "exactOpenRouterOverlap": aa_operational["aa_openrouter_exact_crosscheck"]["exact_epoch_checkpoint_intersection"],
            "priceSpearman": aa_operational["aa_openrouter_exact_crosscheck"]["price"]["spearman"],
            "speedSpearman": aa_operational["aa_openrouter_exact_crosscheck"]["raw_speed"]["spearman"],
            "providerMedianFrontierBaselineMedianX": aa_operational["backtests"]["price_provider_median"]["scopes"]["frontier_like"]["baseline"]["median_multiplicative_error"],
            "providerMedianFrontierCandidateMedianX": aa_operational["backtests"]["price_provider_median"]["scopes"]["frontier_like"]["candidate"]["median_multiplicative_error"],
            "firstPartyFrontierBaselineMedianX": aa_operational["backtests"]["price_first_party"]["scopes"]["frontier_like"]["baseline"]["median_multiplicative_error"],
            "firstPartyFrontierCandidateMedianX": aa_operational["backtests"]["price_first_party"]["scopes"]["frontier_like"]["candidate"]["median_multiplicative_error"],
            "incrementalPriceWeight": aa_operational["decision"]["incremental_aa_operational_price_weight"],
            "incrementalSpeedWeight": aa_operational["decision"]["incremental_aa_speed_weight"],
            "incrementalLatencyWeight": aa_operational["decision"]["incremental_aa_latency_weight"],
        },
        "activeParameterSignals": {
            "totalParameterCheckpoints": active_transport["inventory"]["detailed_total_parameter_checkpoints"],
            "activeParameterCheckpoints": active_transport["inventory"]["active_parameter_checkpoints"],
            "developers": active_transport["inventory"]["active_parameter_developers"],
            "chronologicalPredictions": active_transport["inventory"]["chronological_predictions"],
            "frontierLikePredictions": active_transport["inventory"]["frontier_like_predictions"],
            "activeMedianErrorX": active_transport["active_parameter_predictability"]["active_score_date"]["median_multiplicative_error"],
            "samePanelTotalMedianErrorX": active_transport["active_parameter_predictability"]["total_score_date_same_active_checkpoint_panel"]["median_multiplicative_error"],
            "activeComparisonCi90": active_transport["active_parameter_predictability"]["paired_active_vs_same_panel_total"]["ci_90"],
            "highSparsityPredictions": active_transport["inventory"]["high_sparsity_conversion_predictions"],
            "transportCandidateMedianErrorX": active_transport["high_sparsity_total_transport"]["candidate"]["median_multiplicative_error"],
            "transportBaselineMedianErrorX": active_transport["high_sparsity_total_transport"]["direct_total_baseline"]["median_multiplicative_error"],
            "transportCi90": active_transport["high_sparsity_total_transport"]["paired_cluster_bootstrap"]["ci_90"],
            "predictedK3ActiveB": k3_transport["predicted_k3_active_b"],
            "disclosedK3TotalB": k3_transport["k3_disclosed_total_b"],
            "disclosedK3ActiveB": k3_transport["k3_disclosed_active_b"],
            "k3ActivePredictionErrorX": k3_transport["k3_active_prediction_multiplicative_error"],
            "k3ImpliedRatio": k3_transport["k3_disclosed_total_to_active_ratio"],
            "k3DisclosedActiveFraction": k3_transport["k3_disclosed_active_fraction"],
            "targetSensitivities": active_transport["target_sensitivity"],
            "incrementalWeight": active_transport["decision"]["incremental_live_weight"],
            "estimatedTargetComputeModels": active_transport["compute_branch_independence"]["target_models_with_epoch_training_compute_estimate"],
            "disclosedTargetComputeModels": active_transport["compute_branch_independence"]["target_models_with_disclosed_training_compute"],
            "computeIndependent": active_transport["compute_branch_independence"]["independent_target_evidence"],
        },
        "activePriceSignals": {
            "exactActiveMatches": active_price["inventory"]["active_parameter_matches"],
            "disclosedActiveMatches": active_price["inventory"]["aa_disclosed_active_parameter_matches"],
            "denseConfigControls": active_price["inventory"]["dense_config_active_equals_total_controls"],
            "developers": active_price["inventory"]["developers"],
            "chronologicalPredictions": active_price["inventory"]["release_ordered_predictions"],
            "predictionDevelopers": active_price["inventory"]["prediction_developers"],
            "highSparsityPredictions": active_price["inventory"]["high_sparsity_transport_predictions"],
            "highSparsityDevelopers": active_price["inventory"]["high_sparsity_transport_developers"],
            "candidateMedianErrorX": active_price["high_sparsity_total_transport"]["candidate"]["median_multiplicative_error"],
            "baselineMedianErrorX": active_price["high_sparsity_total_transport"]["direct_total_baseline"]["median_multiplicative_error"],
            "transportCi90": active_price["high_sparsity_total_transport"]["paired_cluster_bootstrap"]["ci_90"],
            "performanceGatePassed": active_price["promotion_gates"]["performance_gate_passed"],
            "coverageGatePassed": active_price["promotion_gates"]["coverage_gate_passed"],
            "incrementalWeight": active_price["decision"]["incremental_live_weight"],
            "prospectivePriceBacktest": False,
            "targetSensitivities": active_price["target_sensitivity"],
        },
        "historicalPriceSignals": {
            "ledgerModels": historical_price["inventory"]["historical_ledger_models"],
            "changePoints": historical_price["inventory"]["historical_change_points"],
            "calibrationCheckpoints": historical_price["inventory"]["calibration_checkpoints_audited"],
            "missingCalibrationAliases": historical_price["inventory"]["calibration_aliases_missing_from_history"],
            "oneDayTotalRows": historical_price["inventory"]["eligible_total_rows_by_window"]["1"],
            "oneDayActiveRows": historical_price["inventory"]["eligible_active_rows_by_window"]["1"],
            "oneDayTotalPredictions": historical_price["heldout_metrics"]["1"]["total"]["date_historical_price"]["n"],
            "oneDayTotalMedianErrorX": historical_price["heldout_metrics"]["1"]["total"]["date_historical_price"]["median_multiplicative_error"],
            "oneDayDateOnlyMedianErrorX": historical_price["heldout_metrics"]["1"]["total"]["date_only"]["median_multiplicative_error"],
            "totalRobustAcrossWindows": historical_price["decision"]["launch_vintage_price_predicts_total_beyond_date_robustly_across_all_windows"],
            "activeIncrementalRobust": historical_price["decision"]["launch_vintage_price_adds_robust_information_beyond_score_date_for_active_parameters"],
            "incrementalWeight": historical_price["decision"]["incremental_live_weight_from_this_audit"],
            "targetSensitivities": historical_price["frontier_first_day_price_sensitivity"],
        },
        "sources": {
            "forecastLedger": {"name": LEDGER.name, "sha256": sha256(LEDGER)},
            "finalWorkbook": {"name": FINAL_WORKBOOK.name, "sha256": sha256(FINAL_WORKBOOK)},
            "branchWorkbook": {"name": BRANCH_WORKBOOK.name, "sha256": sha256(BRANCH_WORKBOOK)},
            "openRouterModels": {"name": OPENROUTER_MODELS.name, "sha256": sha256(OPENROUTER_MODELS)},
            "openRouterTiers": {"name": OPENROUTER_TIERS.name, "sha256": sha256(OPENROUTER_TIERS)},
            "openRouterCollectionAudit": {"name": OPENROUTER_COLLECTION_AUDIT.name, "sha256": sha256(OPENROUTER_COLLECTION_AUDIT)},
            "openRouterAudit": {"name": OPENROUTER_RESULT.name, "sha256": sha256(OPENROUTER_RESULT)},
            "openRouterDaily": {"name": OPENROUTER_DAILY.name, "sha256": sha256(OPENROUTER_DAILY)},
            "openRouterHistoryManifest": {"name": OPENROUTER_HISTORY_MANIFEST.name, "sha256": sha256(OPENROUTER_HISTORY_MANIFEST)},
            "openRouterTemporalAudit": {"name": OPENROUTER_TEMPORAL_RESULT.name, "sha256": sha256(OPENROUTER_TEMPORAL_RESULT)},
            "openRouterRequestWeightedAudit": {"name": OPENROUTER_REQUEST_WEIGHTED_RESULT.name, "sha256": sha256(OPENROUTER_REQUEST_WEIGHTED_RESULT)},
            "openRouterOfficialPrices": {"name": OPENROUTER_OFFICIAL_PRICES.name, "sha256": sha256(OPENROUTER_OFFICIAL_PRICES)},
            "openRouterOfficialAudit": {"name": OPENROUTER_OFFICIAL_AUDIT.name, "sha256": sha256(OPENROUTER_OFFICIAL_AUDIT)},
            "eciReproducedScores": {"name": ECI_REPRODUCED_SCORES.name, "sha256": sha256(ECI_REPRODUCED_SCORES)},
            "eciReproductionCrosscheck": {"name": ECI_REPRODUCTION_CROSSCHECK.name, "sha256": sha256(ECI_REPRODUCTION_CROSSCHECK)},
            "eciReproductionAudit": {"name": ECI_REPRODUCTION_AUDIT.name, "sha256": sha256(ECI_REPRODUCTION_AUDIT)},
            "eciComponentAudit": {"name": ECI_COMPONENT_RESULT.name, "sha256": sha256(ECI_COMPONENT_RESULT)},
            "eciComponentComparison": {"name": ECI_COMPONENT_COMPARISON.name, "sha256": sha256(ECI_COMPONENT_COMPARISON)},
            "eciFitTournament": {"name": ECI_FIT_TOURNAMENT.name, "sha256": sha256(ECI_FIT_TOURNAMENT)},
            "eciFitPredictions": {"name": ECI_FIT_PREDICTIONS.name, "sha256": sha256(ECI_FIT_PREDICTIONS)},
            "eciFitTargets": {"name": ECI_FIT_TARGETS.name, "sha256": sha256(ECI_FIT_TARGETS)},
            "eciHistoricalScores": {"name": ECI_HISTORICAL_SCORES.name, "sha256": sha256(ECI_HISTORICAL_SCORES)},
            "eciHistoricalMetadata": {"name": ECI_HISTORICAL_METADATA.name, "sha256": sha256(ECI_HISTORICAL_METADATA)},
            "eciMultivariateAudit": {"name": ECI_MULTIVARIATE_RESULT.name, "sha256": sha256(ECI_MULTIVARIATE_RESULT)},
            "eciMultivariatePredictions": {"name": ECI_MULTIVARIATE_PREDICTIONS.name, "sha256": sha256(ECI_MULTIVARIATE_PREDICTIONS)},
            "eciMultivariateNarrowCiPredictions": {"name": ECI_MULTIVARIATE_NARROW_CI_PREDICTIONS.name, "sha256": sha256(ECI_MULTIVARIATE_NARROW_CI_PREDICTIONS)},
            "eciMultivariateTargets": {"name": ECI_MULTIVARIATE_TARGETS.name, "sha256": sha256(ECI_MULTIVARIATE_TARGETS)},
            "eciMultivariateCoverage": {"name": ECI_MULTIVARIATE_COVERAGE.name, "sha256": sha256(ECI_MULTIVARIATE_COVERAGE)},
            "posttrainingLineageAudit": {"name": POSTTRAINING_LINEAGE_RESULT.name, "sha256": sha256(POSTTRAINING_LINEAGE_RESULT)},
            "posttrainingLineageEdges": {"name": POSTTRAINING_LINEAGE_EDGES.name, "sha256": sha256(POSTTRAINING_LINEAGE_EDGES)},
            "posttrainingLineageMeasurements": {"name": POSTTRAINING_LINEAGE_MEASUREMENTS.name, "sha256": sha256(POSTTRAINING_LINEAGE_MEASUREMENTS)},
            "posttrainingLineagePredictions": {"name": POSTTRAINING_LINEAGE_PREDICTIONS.name, "sha256": sha256(POSTTRAINING_LINEAGE_PREDICTIONS)},
            "frontierSharedBaseSensitivity": {"name": FRONTIER_SHARED_BASE_SENSITIVITY.name, "sha256": sha256(FRONTIER_SHARED_BASE_SENSITIVITY)},
            "frontierLineageEvidence": {"name": FRONTIER_LINEAGE_EVIDENCE.name, "sha256": sha256(FRONTIER_LINEAGE_EVIDENCE)},
            "aaExpandedAudit": {"name": AA_EXPANDED_RESULT.name, "sha256": sha256(AA_EXPANDED_RESULT)},
            "aaExpandedPanel": {"name": AA_EXPANDED_PANEL.name, "sha256": sha256(AA_EXPANDED_PANEL)},
            "aaExpandedPredictions": {"name": AA_EXPANDED_PREDICTIONS.name, "sha256": sha256(AA_EXPANDED_PREDICTIONS)},
            "aaExpandedOverlaps": {"name": AA_EXPANDED_OVERLAPS.name, "sha256": sha256(AA_EXPANDED_OVERLAPS)},
            "aaDetailedRaw": {"name": AA_DETAILED_RAW.name, "sha256": sha256(AA_DETAILED_RAW)},
            "aaDetailedModels": {"name": AA_DETAILED_MODELS.name, "sha256": sha256(AA_DETAILED_MODELS)},
            "aaInferenceAudit": {"name": AA_INFERENCE_RESULT.name, "sha256": sha256(AA_INFERENCE_RESULT)},
            "aaDetailedPanel": {"name": AA_DETAILED_PANEL.name, "sha256": sha256(AA_DETAILED_PANEL)},
            "aaReasoningPairs": {"name": AA_REASONING_PAIRS.name, "sha256": sha256(AA_REASONING_PAIRS)},
            "aaDetailedCrosscheck": {"name": AA_DETAILED_CROSSCHECK.name, "sha256": sha256(AA_DETAILED_CROSSCHECK)},
            "aaInferencePredictions": {"name": AA_INFERENCE_PREDICTIONS.name, "sha256": sha256(AA_INFERENCE_PREDICTIONS)},
            "aaOperationalAudit": {"name": AA_OPERATIONAL_RESULT.name, "sha256": sha256(AA_OPERATIONAL_RESULT)},
            "aaOperationalPanel": {"name": AA_OPERATIONAL_PANEL.name, "sha256": sha256(AA_OPERATIONAL_PANEL)},
            "aaOperationalPredictions": {"name": AA_OPERATIONAL_PREDICTIONS.name, "sha256": sha256(AA_OPERATIONAL_PREDICTIONS)},
            "aaOpenRouterCrosscheck": {"name": AA_OPENROUTER_CROSSCHECK.name, "sha256": sha256(AA_OPENROUTER_CROSSCHECK)},
            "activeTransportAudit": {"name": ACTIVE_TRANSPORT_RESULT.name, "sha256": sha256(ACTIVE_TRANSPORT_RESULT)},
            "activeTransportPredictions": {"name": ACTIVE_TRANSPORT_PREDICTIONS.name, "sha256": sha256(ACTIVE_TRANSPORT_PREDICTIONS)},
            "activeTransportTargets": {"name": ACTIVE_TRANSPORT_TARGETS.name, "sha256": sha256(ACTIVE_TRANSPORT_TARGETS)},
            "k3ArchitectureFacts": {"name": K3_RELEASE_EVIDENCE.name, "sha256": sha256(K3_RELEASE_EVIDENCE)},
            "epochSnapshotManifest": {"name": EPOCH_SNAPSHOT_MANIFEST.name, "sha256": sha256(EPOCH_SNAPSHOT_MANIFEST)},
            "aaDetailedMetadata": {"name": AA_DETAILED_METADATA.name, "sha256": sha256(AA_DETAILED_METADATA)},
            "openRouterActivePriceAudit": {"name": OPENROUTER_ACTIVE_PRICE_RESULT.name, "sha256": sha256(OPENROUTER_ACTIVE_PRICE_RESULT)},
            "openRouterActivePriceMatches": {"name": OPENROUTER_ACTIVE_PRICE_MATCHES.name, "sha256": sha256(OPENROUTER_ACTIVE_PRICE_MATCHES)},
            "openRouterActivePricePredictions": {"name": OPENROUTER_ACTIVE_PRICE_PREDICTIONS.name, "sha256": sha256(OPENROUTER_ACTIVE_PRICE_PREDICTIONS)},
            "openRouterActivePriceTargets": {"name": OPENROUTER_ACTIVE_PRICE_TARGETS.name, "sha256": sha256(OPENROUTER_ACTIVE_PRICE_TARGETS)},
            "openRouterHistoricalPriceRaw": {"name": OPENROUTER_HISTORICAL_PRICE_RAW.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_RAW)},
            "openRouterHistoricalPricePoints": {"name": OPENROUTER_HISTORICAL_PRICE_POINTS.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_POINTS)},
            "openRouterHistoricalPriceMetadata": {"name": OPENROUTER_HISTORICAL_PRICE_METADATA.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_METADATA)},
            "openRouterHistoricalPriceAudit": {"name": OPENROUTER_HISTORICAL_PRICE_RESULT.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_RESULT)},
            "openRouterHistoricalPriceMatches": {"name": OPENROUTER_HISTORICAL_PRICE_MATCHES.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_MATCHES)},
            "openRouterHistoricalPricePredictions": {"name": OPENROUTER_HISTORICAL_PRICE_PREDICTIONS.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_PREDICTIONS)},
            "openRouterHistoricalPriceTargets": {"name": OPENROUTER_HISTORICAL_PRICE_TARGETS.name, "sha256": sha256(OPENROUTER_HISTORICAL_PRICE_TARGETS)},
            "hfArchitectureRaw": {"name": HF_ARCHITECTURE_RAW.name, "sha256": sha256(HF_ARCHITECTURE_RAW)},
            "hfArchitectureSignals": {"name": HF_ARCHITECTURE_SIGNALS.name, "sha256": sha256(HF_ARCHITECTURE_SIGNALS)},
            "hfArchitectureAudit": {"name": HF_ARCHITECTURE_AUDIT.name, "sha256": sha256(HF_ARCHITECTURE_AUDIT)},
            "noCotExactDateAudit": {"name": NO_COT_EXACT_DATE_AUDIT.name, "sha256": sha256(NO_COT_EXACT_DATE_AUDIT)},
            "noCotExactDateModels": {"name": NO_COT_EXACT_DATE_MODELS.name, "sha256": sha256(NO_COT_EXACT_DATE_MODELS)},
            "noCotArchitectureAudit": {"name": NO_COT_ARCHITECTURE_AUDIT.name, "sha256": sha256(NO_COT_ARCHITECTURE_AUDIT)},
            "noCotArchitecturePredictions": {"name": NO_COT_ARCHITECTURE_PREDICTIONS.name, "sha256": sha256(NO_COT_ARCHITECTURE_PREDICTIONS)},
            "frontierPrimaryEvidence": {"name": FRONTIER_PRIMARY_EVIDENCE.name, "sha256": sha256(FRONTIER_PRIMARY_EVIDENCE)},
            "frontierPrimaryMetadata": {"name": FRONTIER_PRIMARY_METADATA.name, "sha256": sha256(FRONTIER_PRIMARY_METADATA)},
            "frontierPrimaryAudit": {"name": FRONTIER_PRIMARY_AUDIT.name, "sha256": sha256(FRONTIER_PRIMARY_AUDIT)},
            "frontierPrimaryControls": {"name": FRONTIER_PRIMARY_CONTROLS.name, "sha256": sha256(FRONTIER_PRIMARY_CONTROLS)},
            "metrPrimarySignals": {"name": METR_PRIMARY_SIGNALS.name, "sha256": sha256(METR_PRIMARY_SIGNALS)},
            "metrPrimaryRaw": {"name": METR_PRIMARY_RAW.name, "sha256": sha256(METR_PRIMARY_RAW)},
            "metrPrimaryMetadata": {"name": METR_PRIMARY_METADATA.name, "sha256": sha256(METR_PRIMARY_METADATA)},
            "metrPrimaryAudit": {"name": METR_PRIMARY_AUDIT.name, "sha256": sha256(METR_PRIMARY_AUDIT)},
            "ikpAudit": {"name": IKP_RESULT.name, "sha256": sha256(IKP_RESULT)},
            "ikpConditionalAudit": {"name": IKP_CONDITIONAL_RESULT.name, "sha256": sha256(IKP_CONDITIONAL_RESULT)},
            "claudeOpus5Evidence": {"name": OPUS_5_EVIDENCE.name, "sha256": sha256(OPUS_5_EVIDENCE)},
            "k3EfficiencyPrior": {"name": K3_EFFICIENCY_PRIOR.name, "sha256": sha256(K3_EFFICIENCY_PRIOR)},
        },
    }
    SITE_DATA.parent.mkdir(parents=True, exist_ok=True)
    SITE_DATA.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(SITE_DATA), "models": len(models), "active_forecasts": len(forecasts)}, indent=2))


if __name__ == "__main__":
    main()
