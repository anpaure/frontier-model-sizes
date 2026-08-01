#!/usr/bin/env python3
"""Audit AA's detailed parameter and inference-budget signals.

This analysis answers two separate questions without letting either one leak
into the live model automatically:

1. Does AA's much larger, de-duplicated open-weight parameter inventory improve
   strictly chronological developer-held-out parameter recovery?
2. Conditional on the identical token-covered checkpoints, do measured answer
   and reasoning tokens per Intelligence Index task improve recovery beyond AA
   score and exact release date?

Every configuration is retained in the raw source table.  For model-level
regression, configurations sharing the same weight URL, parameter counts, and
release date are collapsed to the highest AA score, matching the user's stated
duplicate policy.  Token features are admitted only when AA reports them.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import load_workbook

from aa_calibration_overrides import (
    OVERRIDES_PATH as AA_CALIBRATION_OVERRIDES_PATH,
    apply_calibration_overrides,
    load_calibration_overrides,
    parameter_label_available_before,
    parameter_training_eligibility_date,
)
from aa_parameter_label_availability import (
    LEDGER_PATH as AA_PARAMETER_LABEL_AVAILABILITY_PATH,
    load_parameter_label_availability,
    resolve_parameter_label_available_date,
)
from aa_score_availability import (
    LEDGER_PATH as AA_SCORE_AVAILABILITY_PATH,
    RAW_PATH as AA_CHANGELOG_PATH,
    aa_prediction_information_date,
    aa_score_availability_verified,
    aa_score_available_date,
)
from k3_primary_evidence import K3_EVIDENCE_PATH, K3_PARAMETER_SOURCE, K3_TOTAL_B, K3_TOTAL_T
from open_model_parameter_truth import (
    LEDGER_PATH as OPEN_MODEL_PARAMETER_TRUTH_PATH,
    apply_parameter_truth,
)
from run_aa_expanded_parameter_audit import load_current_panel


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "outputs/019f6c42-2d53-7743-ab07-6293e2618dd7"
DETAIL = ROOT / "sources/aa_detailed_model_signals_2026-07-31.csv"
RAW = ROOT / "sources/aa_detailed_snapshot_2026-07-31.html.gz"
COLLECTION = ROOT / "sources/aa_detailed_collection_metadata_2026-07-31.json"
EXACT = OUT / "aa_expanded_parameter_panel_2026-07-18.csv"
BRANCH_WORKBOOK = OUT / "k3_calibrated_frontier_parameter_crosscheck_2026-07-17.xlsx"

RESULT = OUT / "aa_inference_budget_audit_2026-07-18.json"
PANEL = OUT / "aa_detailed_parameter_panel_2026-07-18.csv"
PAIRS = OUT / "aa_reasoning_pair_audit_2026-07-18.csv"
EPOCH_CROSSCHECK = OUT / "aa_detailed_epoch_crosscheck_2026-07-18.csv"
PREDICTIONS = OUT / "aa_inference_budget_predictions_2026-07-18.csv"

DATE_ORIGIN = date(2023, 1, 1)
BOOTSTRAPS = 20_000
MIN_CURRENT_ROWS = 16
MIN_CURRENT_DEVELOPERS = 6
MIN_DETAILED_ROWS = 30
MIN_DETAILED_DEVELOPERS = 8
MIN_TOKEN_ROWS = 20
MIN_TOKEN_DEVELOPERS = 8
K3_SLUG = "kimi-k3"
RAW_SNAPSHOT_CHECKPOINT_GROUPS = 273

# These are identity crosswalks, not value overrides.  Each detailed AA slug is
# the same checkpoint named by Epoch, but at least one AA metadata field falls
# outside the conservative automatic gates.  The crosscheck preserves both
# source values and emits the disagreement explicitly.
MANUAL_EPOCH_DETAIL_SLUGS = {
    "checkpoint:epoch:phi-4-mini": {
        "slug": "phi-4-mini",
        "reason": "exact model identity; AA release year is inconsistent with the 2025 public release",
    },
    "checkpoint:epoch:granite-4-0-h-tiny": {
        "slug": "granite-4-0-h-nano-1b",
        "reason": "manual active-parameter alias; AA reports 1.5B while Epoch documents 7B total / 1B active",
    },
    "checkpoint:epoch:ring-flash-linear-2-0": {
        "slug": "ring-flash-2-0",
        "reason": "manual architecture-token alias; AA creator/date metadata differ from Epoch",
    },
}

DEVELOPER_TO_CREATOR = {
    "allenai": "ai2",
    "ant": "antgroup",
    "microsoft": "azure",
    "moonshot": "kimi",
    "nous": "nous-research",
    "prime_intellect": "prime-intellect",
    "reka": "reka-ai",
}

FRONTIER_SLUGS = (
    "claude-fable-5",
    "gpt-5-6-sol",
    "kimi-k3",
    "claude-opus-4-8",
    "gpt-5-5",
    "gpt-5-6-terra",
    "claude-sonnet-5",
    "gpt-5-6-luna",
    "grok-4-5",
)

FRONTIER_BRANCH_ALIASES = {
    "claude-fable-5": ("Claude Fable 5",),
    "gpt-5-6-sol": ("GPT-5.6 Sol",),
    "kimi-k3": ("Kimi K3",),
    "claude-opus-4-8": ("Claude Opus 4.7", "Claude Opus 4.8"),
    "gpt-5-5": ("GPT-5.5",),
    "gpt-5-6-terra": ("GPT-5.6 Terra",),
    "claude-sonnet-5": ("Claude Sonnet 5",),
    "gpt-5-6-luna": ("GPT-5.6 Luna",),
    "grok-4-5": ("Grok 4.5",),
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"Refusing to write empty output {path}")
    fields: list[str] = []
    for row in rows:
        for field in row:
            if field not in fields:
                fields.append(field)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fields, extrasaction="raise", lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)


def load_upstream_aa_factors() -> dict[str, float]:
    """Read AA factors from the upstream workbook, avoiding a site/audit cycle."""
    workbook = load_workbook(BRANCH_WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["Revised Estimates"]
    headers = [cell.value for cell in sheet[5]]
    columns = {header: index for index, header in enumerate(headers)}
    by_model: dict[str, float] = {}
    for values in sheet.iter_rows(min_row=6, max_col=len(headers), values_only=True):
        model = values[columns["Model"]]
        if model in (None, ""):
            break
        value = values[columns["AA direct, K3-anchored (B)"]]
        if value is not None:
            by_model[str(model)] = float(value) / 1000.0
    workbook.close()

    output: dict[str, float] = {}
    for slug, aliases in FRONTIER_BRANCH_ALIASES.items():
        values = [by_model[alias] for alias in aliases if alias in by_model]
        if not values:
            raise ValueError(f"Missing upstream AA factor for {slug}: {aliases}")
        output[slug] = math.exp(sum(math.log(value) for value in values) / len(values))
    return output


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def number(value: str | None) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def boolean(value: str) -> bool:
    if value not in {"True", "False"}:
        raise ValueError(f"Invalid serialized boolean {value!r}")
    return value == "True"


def years(value: str) -> float:
    return (date.fromisoformat(value) - DATE_ORIGIN).days / 365.25


def normalize_url(value: str) -> str:
    normalized = value.strip().lower().rstrip("/")
    return re.sub(r"/tree/(main|master)$", "", normalized)


def normalize_name(value: str) -> str:
    text = value.lower().replace("non-reasoning", "non reasoning")
    text = re.sub(
        r"\([^)]*\b(max|xhigh|high|medium|low|minimal|reasoning)\b[^)]*\)",
        "",
        text,
    )
    text = re.sub(
        r"\b(non reasoning|reasoning|thinking|instruct|vision|preview|it)\b", "", text
    )
    return re.sub(r"[^a-z0-9]+", "", text)


def checkpoint_group_key(row: dict[str, Any]) -> str:
    parameter_key = (
        f"{row['parameters_b']:.9g}|"
        f"{'' if row['active_parameters_b'] is None else f'{row['active_parameters_b']:.9g}'}|"
        f"{row['release_date']}"
    )
    weights_url = normalize_url(row["model_weights_source_url"])
    if weights_url:
        return f"weights:{weights_url}|{parameter_key}"
    return (
        f"fallback:{row['creator_slug']}|{parameter_key}|"
        f"{normalize_name(row['name'])}"
    )


def configuration_group_key(row: dict[str, Any]) -> str:
    """Group reasoning modes of one checkpoint, including closed models.

    Proprietary models have no weight URL or parameter count, so the strongest
    common identity available across reasoning modes is creator + exact release
    date + reported parameter fields + a mode-stripped normalized name.  Raw
    rows remain preserved and every selected pair records whether an exact
    weight URL independently confirms the identity.
    """

    parameters = "" if row["parameters_b"] is None else f"{row['parameters_b']:.9g}"
    active = (
        ""
        if row["active_parameters_b"] is None
        else f"{row['active_parameters_b']:.9g}"
    )
    return (
        f"configuration:{row['creator_slug']}|{row['release_date']}|"
        f"{parameters}|{active}|{normalize_name(row['name'])}"
    )


def parse_detail_rows() -> list[dict[str, Any]]:
    output = []
    for row in read_csv(DETAIL):
        output.append(
            {
                **row,
                "is_reasoning": boolean(row["is_reasoning"]),
                "is_open_weights": boolean(row["is_open_weights"]),
                "intelligence_index_estimated": boolean(
                    row["intelligence_index_estimated"]
                ),
                "parameters_b": number(row["parameters_b"]),
                "active_parameters_b": number(row["active_parameters_b"]),
                "intelligence_index": number(row["intelligence_index"]),
                "output_tokens_per_task": number(
                    row["intelligence_output_tokens_per_task"]
                ),
                "answer_tokens_per_task": number(
                    row["intelligence_answer_tokens_per_task"]
                ),
                "reasoning_tokens_per_task": number(
                    row["intelligence_reasoning_tokens_per_task"]
                ),
            }
        )
    if len(output) != 587:
        raise ValueError(f"Expected 587 detailed AA model rows, found {len(output)}")
    corrected, override_audit = apply_calibration_overrides(output)
    corrected = [
        apply_parameter_truth(
            row,
            name_fields=("name", "short_name"),
            total_fields=("parameters_b",),
            active_fields=("active_parameters_b",),
        )
        for row in corrected
    ]
    expected_override_ids = {
        row["override_id"] for row in load_calibration_overrides()["overrides"]
    }
    if {row["override_id"] for row in override_audit} != expected_override_ids:
        raise ValueError(f"Unexpected AA primary-source overrides: {override_audit}")
    return corrected


def build_panel(
    detailed: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    eligible = [
        row
        for row in detailed
        if row["is_open_weights"]
        and row["parameters_b"] is not None
        and row["intelligence_index"] is not None
        and row["release_date"]
    ]
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in eligible:
        groups[checkpoint_group_key(row)].append(row)

    panel = []
    for group_id, configurations in sorted(groups.items()):
        parameters = {row["parameters_b"] for row in configurations}
        active = {row["active_parameters_b"] for row in configurations}
        dates = {row["release_date"] for row in configurations}
        creators = {row["creator_slug"] for row in configurations}
        if any(len(values) != 1 for values in (parameters, active, dates, creators)):
            raise ValueError(f"Non-identical checkpoint group {group_id}")
        selected = max(
            configurations,
            key=lambda row: (row["intelligence_index"], row["slug"]),
        )
        panel.append(
            {
                "checkpoint_group_id": group_id,
                "selected_model_id": selected["model_id"],
                "selected_slug": selected["slug"],
                "selected_name": selected["name"],
                "creator_slug": selected["creator_slug"],
                "creator_name": selected["creator_name"],
                "release_date": selected["release_date"],
                "parameters_b": selected["parameters_b"],
                "active_parameters_b": selected["active_parameters_b"],
                "intelligence_index": selected["intelligence_index"],
                "intelligence_index_estimated": selected[
                    "intelligence_index_estimated"
                ],
                "is_reasoning": selected["is_reasoning"],
                "output_tokens_per_task": selected["output_tokens_per_task"],
                "answer_tokens_per_task": selected["answer_tokens_per_task"],
                "reasoning_tokens_per_task": selected[
                    "reasoning_tokens_per_task"
                ],
                "configuration_count": len(configurations),
                "lower_score_configurations_removed": len(configurations) - 1,
                "configuration_slugs": " | ".join(
                    row["slug"]
                    for row in sorted(
                        configurations,
                        key=lambda row: (-row["intelligence_index"], row["slug"]),
                    )
                ),
                "model_weights_source_url": selected["model_weights_source_url"],
                "source_page_url": selected["source_page_url"],
                "parameter_source": selected.get("calibration_override_source_url")
                or "Artificial Analysis public model metadata",
                "calibration_override_id": selected.get(
                    "calibration_override_id", ""
                ),
                "calibration_override_fields": selected.get(
                    "calibration_override_fields", ""
                ),
                "calibration_override_source_url": selected.get(
                    "calibration_override_source_url", ""
                ),
                "calibration_override_source_record_sha256": selected.get(
                    "calibration_override_source_record_sha256", ""
                ),
                "parameter_label_available_date": resolve_parameter_label_available_date(
                    selected
                ),
                "parameter_training_eligibility_date": parameter_training_eligibility_date(
                    selected
                ),
                "aa_score_available_date": aa_score_available_date(selected),
                "aa_score_availability_verified": aa_score_availability_verified(
                    selected
                ),
                "architecture_class": selected.get("architecture_class", ""),
                "exact_tensor_parameters": selected.get("exact_tensor_parameters", ""),
                "lineage_class": selected.get("lineage_class", ""),
                "base_model_id": selected.get("base_model_id", ""),
                "lineage_family_id": selected.get("lineage_family_id", ""),
                "parameter_truth_id": selected.get("parameter_truth_id", ""),
                "parameter_truth_basis": selected.get("parameter_truth_basis", ""),
                "raw_parameter_total_b": selected.get("raw_parameter_total_b", ""),
                "raw_parameter_active_b": selected.get("raw_parameter_active_b", ""),
            }
        )
    collection = json.loads(COLLECTION.read_text(encoding="utf-8"))
    override_count = len(load_calibration_overrides()["overrides"])
    expected_eligible = (
        int(collection["open_weight_parameter_score_date_rows"]) + override_count
    )
    expected_panel = RAW_SNAPSHOT_CHECKPOINT_GROUPS + override_count
    if len(eligible) != expected_eligible or len(panel) != expected_panel:
        raise ValueError(
            f"Unexpected detailed AA grouping: {len(eligible)} rows -> {len(panel)} groups"
        )
    if len({row["checkpoint_group_id"] for row in panel}) != len(panel):
        raise ValueError("Detailed AA panel group IDs are not unique")
    return panel, groups


def build_configuration_groups(
    detailed: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in detailed:
        if (
            row["intelligence_index"] is not None
            and row["release_date"]
            and row["creator_slug"]
        ):
            groups[configuration_group_key(row)].append(row)
    return groups


def build_reasoning_pairs(
    groups: dict[str, list[dict[str, Any]]], *, pair_basis: str
) -> list[dict[str, Any]]:
    output = []
    for group_id, configurations in sorted(groups.items()):
        reasoning = [row for row in configurations if row["is_reasoning"]]
        nonreasoning = [row for row in configurations if not row["is_reasoning"]]
        if not reasoning or not nonreasoning:
            continue
        high = max(reasoning, key=lambda row: row["intelligence_index"])
        low = max(nonreasoning, key=lambda row: row["intelligence_index"])
        high_url = normalize_url(high["model_weights_source_url"])
        low_url = normalize_url(low["model_weights_source_url"])
        if high["is_open_weights"] != low["is_open_weights"]:
            raise ValueError(f"Open-weight status differs within pair {group_id}")
        if high["parameters_b"] != low["parameters_b"]:
            raise ValueError(f"Parameter counts differ within pair {group_id}")
        output.append(
            {
                "checkpoint_group_id": group_id,
                "creator_slug": high["creator_slug"],
                "release_date": high["release_date"],
                "is_open_weights": high["is_open_weights"],
                "parameters_known": high["parameters_b"] is not None,
                "parameters_b": high["parameters_b"],
                "active_parameters_b": high["active_parameters_b"],
                "reasoning_slug": high["slug"],
                "reasoning_name": high["name"],
                "reasoning_aa": high["intelligence_index"],
                "reasoning_answer_tokens_per_task": high[
                    "answer_tokens_per_task"
                ],
                "reasoning_reasoning_tokens_per_task": high[
                    "reasoning_tokens_per_task"
                ],
                "reasoning_output_tokens_per_task": high["output_tokens_per_task"],
                "nonreasoning_slug": low["slug"],
                "nonreasoning_name": low["name"],
                "nonreasoning_aa": low["intelligence_index"],
                "nonreasoning_answer_tokens_per_task": low[
                    "answer_tokens_per_task"
                ],
                "nonreasoning_reasoning_tokens_per_task": low[
                    "reasoning_tokens_per_task"
                ],
                "nonreasoning_output_tokens_per_task": low[
                    "output_tokens_per_task"
                ],
                "aa_uplift": high["intelligence_index"] - low["intelligence_index"],
                "same_weights_url": bool(high_url and high_url == low_url),
                "pair_basis": pair_basis,
                "selection": "highest-score reasoning versus highest-score non-reasoning configuration",
            }
        )
    return output


def match_epoch_exact(
    panel: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    exact = [
        row
        for row in read_csv(EXACT)
        if row["panel_source"] == "exact_epoch_checkpoint"
    ]
    output = []
    unmatched = []
    used: set[str] = set()
    for epoch in exact:
        expected_creator = DEVELOPER_TO_CREATOR.get(
            epoch["developer"], epoch["developer"]
        )
        manual = MANUAL_EPOCH_DETAIL_SLUGS.get(epoch["canonical_checkpoint_id"])
        if manual:
            candidates = [
                row for row in panel if row["selected_slug"] == manual["slug"]
            ]
            if len(candidates) != 1:
                raise ValueError(
                    f"Manual detailed AA match does not resolve uniquely: {epoch['canonical_checkpoint_id']}"
                )
            detailed = candidates[0]
            ratio = detailed["parameters_b"] / float(epoch["total_parameters_b"])
            date_delta = (
                date.fromisoformat(detailed["release_date"])
                - date.fromisoformat(epoch["release_date"])
            ).days
            similarity = max(
                SequenceMatcher(
                    None,
                    normalize_name(epoch[name_field]),
                    normalize_name(detailed["selected_name"]),
                ).ratio()
                for name_field in ("model", "matched_epoch_model")
            )
            match_method = f"manual explicit identity crosswalk: {manual['reason']}"
        else:
            candidates = []
            for detailed in panel:
                if detailed["creator_slug"] != expected_creator:
                    continue
                ratio = detailed["parameters_b"] / float(epoch["total_parameters_b"])
                if not 0.80 <= ratio <= 1.20:
                    continue
                date_delta = (
                    date.fromisoformat(detailed["release_date"])
                    - date.fromisoformat(epoch["release_date"])
                ).days
                if abs(date_delta) > 45:
                    continue
                name_similarity = max(
                    SequenceMatcher(
                        None,
                        normalize_name(epoch[name_field]),
                        normalize_name(detailed["selected_name"]),
                    ).ratio()
                    for name_field in ("model", "matched_epoch_model")
                )
                if name_similarity < 0.50:
                    continue
                score = (
                    name_similarity,
                    -abs(math.log(ratio)),
                    -abs(date_delta),
                    detailed["intelligence_index"],
                )
                candidates.append(
                    (score, detailed, ratio, date_delta, name_similarity)
                )
            candidates.sort(key=lambda item: item[0], reverse=True)
            if not candidates:
                unmatched.append(epoch["model"])
                continue
            _, detailed, ratio, date_delta, similarity = candidates[0]
            match_method = "creator + name similarity + date + parameter scale"
        if detailed["checkpoint_group_id"] in used:
            raise ValueError(
                f"Detailed AA checkpoint reused in Epoch match: {detailed['selected_name']}"
            )
        used.add(detailed["checkpoint_group_id"])
        output.append(
            {
                "epoch_checkpoint_id": epoch["canonical_checkpoint_id"],
                "epoch_model": epoch["matched_epoch_model"],
                "aa_attachment_model": epoch["model"],
                "aa_detailed_group_id": detailed["checkpoint_group_id"],
                "aa_detailed_model": detailed["selected_name"],
                "creator_slug": detailed["creator_slug"],
                "epoch_release_date": epoch["release_date"],
                "aa_release_date": detailed["release_date"],
                "date_delta_days": date_delta,
                "epoch_parameters_b": float(epoch["total_parameters_b"]),
                "aa_parameters_b": detailed["parameters_b"],
                "aa_over_epoch_parameters": ratio,
                "absolute_log_parameter_difference": abs(math.log(ratio)),
                "name_similarity": similarity,
                "creator_agreement": detailed["creator_slug"] == expected_creator,
                "date_within_45_days": abs(date_delta) <= 45,
                "parameters_within_20_percent": 0.80 <= ratio <= 1.20,
                "match_method": match_method,
            }
        )
    return output, unmatched


def developer_weights(rows: list[dict[str, Any]], field: str) -> np.ndarray:
    counts = Counter(row[field] for row in rows)
    weights = np.asarray(
        [
            (0.5 if row.get("estimated", False) else 1.0) / counts[row[field]]
            for row in rows
        ],
        dtype=float,
    )
    return weights / weights.mean()


def feature(row: dict[str, Any], name: str) -> float:
    if name == "date":
        return years(row["release_date"])
    if name == "log_answer_tokens":
        return math.log10(1.0 + row["answer_tokens_per_task"])
    if name == "log_reasoning_tokens":
        return math.log10(1.0 + row["reasoning_tokens_per_task"])
    return float(row[name])


def fit(
    rows: list[dict[str, Any]], features: tuple[str, ...], developer_field: str
) -> np.ndarray:
    x = np.asarray(
        [[1.0, *[feature(row, name) for name in features]] for row in rows],
        dtype=float,
    )
    y = np.log10(np.asarray([row["parameters_b"] for row in rows], dtype=float))
    root = np.sqrt(developer_weights(rows, developer_field))
    beta, *_ = np.linalg.lstsq(x * root[:, None], y * root, rcond=None)
    return beta


def predict(
    rows: list[dict[str, Any]],
    test: dict[str, Any],
    features: tuple[str, ...],
    developer_field: str,
) -> tuple[float, list[float]]:
    beta = fit(rows, features, developer_field)
    values = np.asarray([1.0, *[feature(test, name) for name in features]])
    return float(values @ beta), [float(value) for value in beta]


def parameter_metrics(errors: Iterable[float]) -> dict[str, Any]:
    values = np.asarray(list(errors), dtype=float)
    absolute = np.abs(values)
    return {
        "n": int(len(values)),
        "median_multiplicative_error": float(10 ** np.median(absolute)),
        "mean_absolute_log10_error": float(np.mean(absolute)),
        "geomean_multiplicative_error": float(10 ** np.mean(absolute)),
        "rmse_log10": float(np.sqrt(np.mean(values**2))),
        "p80_multiplicative_error": float(10 ** np.quantile(absolute, 0.80)),
        "within_2x": float(np.mean(absolute <= math.log10(2.0))),
        "signed_bias_factor": float(10 ** np.mean(values)),
    }


def paired_cluster_bootstrap(
    rows: list[dict[str, Any]],
    cluster_field: str,
    candidate_error: str,
    baseline_error: str,
    seed: int,
) -> dict[str, Any]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row[cluster_field]].append(row)
    effects = np.asarray(
        [
            np.mean(
                [
                    abs(item[candidate_error]) - abs(item[baseline_error])
                    for item in items
                ]
            )
            for _, items in sorted(grouped.items())
        ],
        dtype=float,
    )
    rng = np.random.default_rng(seed)
    draws = np.mean(
        effects[
            rng.integers(0, len(effects), size=(BOOTSTRAPS, len(effects)))
        ],
        axis=1,
    )
    return {
        "metric": "equal-cluster mean absolute log10 error; candidate minus baseline",
        "observed_delta": float(np.mean(effects)),
        "ci_90": [float(value) for value in np.quantile(draws, [0.05, 0.95])],
        "bootstrap_probability_candidate_better": float(np.mean(draws < 0)),
        "samples": BOOTSTRAPS,
        "clusters": len(effects),
    }


def current_comparison(
    panel: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    current = [
        {
            "model": row["model"],
            "developer": row["developer"],
            "release_date": row["release_date"],
            "score": row["aa_score"],
            "parameters_b": row["total_parameters_b"],
            "raw_parameter_total_b": row.get("raw_parameter_total_b", ""),
            "estimated": bool(row["estimated_score"]),
            "parameter_label_available_date": row[
                "parameter_label_available_date"
            ],
            "aa_slug": row.get("aa_slug", ""),
            "aa_score_available_date": row.get(
                "aa_score_available_date", row["release_date"]
            ),
            "aa_score_availability_verified": row.get(
                "aa_score_availability_verified", False
            ),
        }
        for row in load_current_panel()
    ]
    detailed = [
        {
            "model": row["selected_name"],
            "developer": row["creator_slug"],
            "release_date": row["release_date"],
            "score": row["intelligence_index"],
            "parameters_b": row["parameters_b"],
            "raw_parameter_total_b": row.get("raw_parameter_total_b", ""),
            "estimated": row["intelligence_index_estimated"],
            "parameter_label_available_date": row[
                "parameter_label_available_date"
            ],
            "selected_slug": row["selected_slug"],
            "aa_score_available_date": row["aa_score_available_date"],
            "aa_score_availability_verified": row[
                "aa_score_availability_verified"
            ],
        }
        for row in panel
    ]
    output = []
    features = ("score", "date")
    for test in sorted(
        current,
        key=lambda row: (
            aa_prediction_information_date(row),
            row["release_date"],
            row["model"],
        ),
    ):
        prediction_date = aa_prediction_information_date(test)
        current_train = [
            row
            for row in current
            if parameter_label_available_before(row, prediction_date)
            and row["developer"] != test["developer"]
        ]
        detailed_creator = DEVELOPER_TO_CREATOR.get(
            test["developer"], test["developer"]
        )
        detailed_train = [
            row
            for row in detailed
            if parameter_label_available_before(row, prediction_date)
            and row["developer"] != detailed_creator
        ]
        if (
            len(current_train) < MIN_CURRENT_ROWS
            or len({row["developer"] for row in current_train})
            < MIN_CURRENT_DEVELOPERS
            or len(detailed_train) < MIN_DETAILED_ROWS
            or len({row["developer"] for row in detailed_train})
            < MIN_DETAILED_DEVELOPERS
        ):
            continue
        current_prediction, current_beta = predict(
            current_train, test, features, "developer"
        )
        detailed_prediction, detailed_beta = predict(
            detailed_train, test, features, "developer"
        )
        actual = math.log10(test["parameters_b"])
        output.append(
            {
                "comparison": "current_50_vs_detailed",
                "release_date": test["release_date"],
                "prediction_information_date": prediction_date,
                "model": test["model"],
                "developer": test["developer"],
                "aa_score": test["score"],
                "actual_parameters_b": test["parameters_b"],
                "frontier_score_rank": sum(
                    row["score"] <= test["score"] for row in current_train
                )
                / len(current_train),
                "baseline_log10_error": current_prediction - actual,
                "candidate_log10_error": detailed_prediction - actual,
                "baseline_predicted_b": 10**current_prediction,
                "candidate_predicted_b": 10**detailed_prediction,
                "baseline_train_n": len(current_train),
                "candidate_train_n": len(detailed_train),
                "baseline_train_developers": len(
                    {row["developer"] for row in current_train}
                ),
                "candidate_train_developers": len(
                    {row["developer"] for row in detailed_train}
                ),
                "baseline_train_max_date": max(
                    parameter_training_eligibility_date(row)
                    for row in current_train
                ),
                "candidate_train_max_date": max(
                    parameter_training_eligibility_date(row)
                    for row in detailed_train
                ),
                "test_developer_excluded": True,
                "baseline_coefficients": json.dumps(current_beta, separators=(",", ":")),
                "candidate_coefficients": json.dumps(
                    detailed_beta, separators=(",", ":")
                ),
            }
        )
    return output


def token_comparison(panel: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = [
        {
            "model": row["selected_name"],
            "developer": row["creator_slug"],
            "release_date": row["release_date"],
            "score": row["intelligence_index"],
            "parameters_b": row["parameters_b"],
            "raw_parameter_total_b": row.get("raw_parameter_total_b", ""),
            "estimated": row["intelligence_index_estimated"],
            "answer_tokens_per_task": row["answer_tokens_per_task"],
            "reasoning_tokens_per_task": row["reasoning_tokens_per_task"],
            "parameter_label_available_date": row[
                "parameter_label_available_date"
            ],
            "selected_slug": row["selected_slug"],
            "aa_score_available_date": row["aa_score_available_date"],
            "aa_score_availability_verified": row[
                "aa_score_availability_verified"
            ],
        }
        for row in panel
        if row["output_tokens_per_task"] is not None
        and row["answer_tokens_per_task"] is not None
        and row["reasoning_tokens_per_task"] is not None
    ]
    baseline_features = ("score", "date")
    candidate_features = (
        "score",
        "date",
        "log_answer_tokens",
        "log_reasoning_tokens",
    )
    output = []
    for test in sorted(
        rows,
        key=lambda row: (
            aa_prediction_information_date(row),
            row["release_date"],
            row["model"],
        ),
    ):
        prediction_date = aa_prediction_information_date(test)
        train = [
            row
            for row in rows
            if parameter_label_available_before(row, prediction_date)
            and row["developer"] != test["developer"]
        ]
        if (
            len(train) < MIN_TOKEN_ROWS
            or len({row["developer"] for row in train}) < MIN_TOKEN_DEVELOPERS
        ):
            continue
        baseline_prediction, baseline_beta = predict(
            train, test, baseline_features, "developer"
        )
        candidate_prediction, candidate_beta = predict(
            train, test, candidate_features, "developer"
        )
        actual = math.log10(test["parameters_b"])
        output.append(
            {
                "comparison": "token_budget_incremental",
                "release_date": test["release_date"],
                "prediction_information_date": prediction_date,
                "model": test["model"],
                "developer": test["developer"],
                "aa_score": test["score"],
                "actual_parameters_b": test["parameters_b"],
                "answer_tokens_per_task": test["answer_tokens_per_task"],
                "reasoning_tokens_per_task": test["reasoning_tokens_per_task"],
                "frontier_score_rank": sum(
                    row["score"] <= test["score"] for row in train
                )
                / len(train),
                "baseline_log10_error": baseline_prediction - actual,
                "candidate_log10_error": candidate_prediction - actual,
                "baseline_predicted_b": 10**baseline_prediction,
                "candidate_predicted_b": 10**candidate_prediction,
                "baseline_train_n": len(train),
                "candidate_train_n": len(train),
                "baseline_train_developers": len(
                    {row["developer"] for row in train}
                ),
                "candidate_train_developers": len(
                    {row["developer"] for row in train}
                ),
                "baseline_train_max_date": max(
                    parameter_training_eligibility_date(row) for row in train
                ),
                "candidate_train_max_date": max(
                    parameter_training_eligibility_date(row) for row in train
                ),
                "test_developer_excluded": True,
                "baseline_coefficients": json.dumps(
                    baseline_beta, separators=(",", ":")
                ),
                "candidate_coefficients": json.dumps(
                    candidate_beta, separators=(",", ":")
                ),
            }
        )
    return output


def equal_creator_median_uplift(pairs: list[dict[str, Any]]) -> float | None:
    by_creator: dict[str, list[float]] = defaultdict(list)
    for row in pairs:
        by_creator[row["creator_slug"]].append(row["aa_uplift"])
    if not by_creator:
        return None
    return float(
        np.median(
            [np.median(values) for _, values in sorted(by_creator.items())]
        )
    )


def standardize_reasoning_score(
    row: dict[str, Any],
    pairs: list[dict[str, Any]],
    *,
    creator_aware: bool,
) -> tuple[float | None, float | None, str, int, int, str]:
    """Return a no-reasoning-equivalent AA score using only available history.

    A directly measured non-reasoning configuration of the identical checkpoint
    is highest priority.  Otherwise the adjustment uses only pairs released
    strictly before the row.  The portable variant excludes the row's creator;
    the creator-aware variant uses a same-creator median only after two prior
    pairs, falling back to the portable equal-creator median.
    """

    if not row["is_reasoning"]:
        return row["intelligence_index"], 0.0, "observed non-reasoning", 0, 0, ""

    direct = [
        pair
        for pair in pairs
        if pair["reasoning_slug"] == row["selected_slug"]
        and pair["release_date"] == row["release_date"]
    ]
    if len(direct) > 1:
        raise ValueError(f"Ambiguous direct reasoning pair for {row['selected_slug']}")
    if direct:
        pair = direct[0]
        return (
            pair["nonreasoning_aa"],
            pair["aa_uplift"],
            "direct same-checkpoint non-reasoning score",
            1,
            1,
            row["release_date"],
        )

    prior = [pair for pair in pairs if pair["release_date"] < row["release_date"]]
    portable = [
        pair for pair in prior if pair["creator_slug"] != row["creator_slug"]
    ]
    portable_uplift = equal_creator_median_uplift(portable)
    portable_creators = len({pair["creator_slug"] for pair in portable})
    if creator_aware:
        same_creator = [
            pair for pair in prior if pair["creator_slug"] == row["creator_slug"]
        ]
        if len(same_creator) >= 2:
            uplift = float(np.median([pair["aa_uplift"] for pair in same_creator]))
            return (
                row["intelligence_index"] - uplift,
                uplift,
                "median of >=2 strictly earlier same-creator pairs",
                len(same_creator),
                1,
                max(pair["release_date"] for pair in same_creator),
            )
    if portable_uplift is None or portable_creators < 3:
        return (
            None,
            None,
            "insufficient strictly earlier pair history",
            len(portable),
            portable_creators,
            max((pair["release_date"] for pair in portable), default=""),
        )
    return (
        row["intelligence_index"] - portable_uplift,
        portable_uplift,
        "equal-creator median of strictly earlier other-creator pairs",
        len(portable),
        portable_creators,
        max(pair["release_date"] for pair in portable),
    )


def reasoning_standardization_comparison(
    panel: list[dict[str, Any]],
    pairs: list[dict[str, Any]],
    *,
    creator_aware: bool,
) -> list[dict[str, Any]]:
    rows = []
    for row in panel:
        standardized, uplift, method, pair_n, pair_creators, history_max_date = (
            standardize_reasoning_score(
                row, pairs, creator_aware=creator_aware
            )
        )
        if standardized is None:
            continue
        rows.append(
            {
                "model": row["selected_name"],
                "slug": row["selected_slug"],
                "developer": row["creator_slug"],
                "release_date": row["release_date"],
                "score": row["intelligence_index"],
                "standardized_score": standardized,
                "reasoning_uplift_removed": uplift,
                "standardization_method": method,
                "standardization_pair_n": pair_n,
                "standardization_pair_creators": pair_creators,
                "standardization_history_max_date": history_max_date,
                "parameters_b": row["parameters_b"],
                "raw_parameter_total_b": row.get("raw_parameter_total_b", ""),
                "estimated": row["intelligence_index_estimated"],
                "parameter_label_available_date": row[
                    "parameter_label_available_date"
                ],
                "selected_slug": row["selected_slug"],
                "aa_score_available_date": row["aa_score_available_date"],
                "aa_score_availability_verified": row[
                    "aa_score_availability_verified"
                ],
            }
        )

    output = []
    baseline_features = ("score", "date")
    candidate_features = ("standardized_score", "date")
    label = "creator_aware" if creator_aware else "portable"
    for test in sorted(
        rows,
        key=lambda row: (
            aa_prediction_information_date(row),
            row["release_date"],
            row["model"],
        ),
    ):
        prediction_date = aa_prediction_information_date(test)
        train = [
            row
            for row in rows
            if parameter_label_available_before(row, prediction_date)
            and row["developer"] != test["developer"]
        ]
        if (
            len(train) < MIN_DETAILED_ROWS
            or len({row["developer"] for row in train}) < MIN_DETAILED_DEVELOPERS
        ):
            continue
        baseline_prediction, baseline_beta = predict(
            train, test, baseline_features, "developer"
        )
        candidate_prediction, candidate_beta = predict(
            train, test, candidate_features, "developer"
        )
        actual = math.log10(test["parameters_b"])
        output.append(
            {
                "comparison": f"reasoning_standardization_{label}",
                "release_date": test["release_date"],
                "prediction_information_date": prediction_date,
                "model": test["model"],
                "developer": test["developer"],
                "aa_score": test["score"],
                "standardized_aa_score": test["standardized_score"],
                "reasoning_uplift_removed": test["reasoning_uplift_removed"],
                "standardization_method": test["standardization_method"],
                "standardization_pair_n": test["standardization_pair_n"],
                "standardization_pair_creators": test[
                    "standardization_pair_creators"
                ],
                "standardization_history_max_date": test[
                    "standardization_history_max_date"
                ],
                "actual_parameters_b": test["parameters_b"],
                "frontier_score_rank": sum(
                    row["score"] <= test["score"] for row in train
                )
                / len(train),
                "baseline_log10_error": baseline_prediction - actual,
                "candidate_log10_error": candidate_prediction - actual,
                "baseline_predicted_b": 10**baseline_prediction,
                "candidate_predicted_b": 10**candidate_prediction,
                "baseline_train_n": len(train),
                "candidate_train_n": len(train),
                "baseline_train_developers": len(
                    {row["developer"] for row in train}
                ),
                "candidate_train_developers": len(
                    {row["developer"] for row in train}
                ),
                "baseline_train_max_date": max(
                    parameter_training_eligibility_date(row) for row in train
                ),
                "candidate_train_max_date": max(
                    parameter_training_eligibility_date(row) for row in train
                ),
                "test_developer_excluded": True,
                "baseline_coefficients": json.dumps(
                    baseline_beta, separators=(",", ":")
                ),
                "candidate_coefficients": json.dumps(
                    candidate_beta, separators=(",", ":")
                ),
            }
        )
    return output


def comparison_summary(
    rows: list[dict[str, Any]], cluster_field: str, seed: int
) -> dict[str, Any]:
    scopes = {}
    for index, (scope, selected) in enumerate(
        (
            ("all", rows),
            (
                "frontier_like",
                [row for row in rows if row["frontier_score_rank"] >= 0.90],
            ),
        )
    ):
        scopes[scope] = {
            "n": len(selected),
            "developers": len({row[cluster_field] for row in selected}),
            "baseline": parameter_metrics(
                row["baseline_log10_error"] for row in selected
            ),
            "candidate": parameter_metrics(
                row["candidate_log10_error"] for row in selected
            ),
            "paired_cluster_bootstrap": paired_cluster_bootstrap(
                selected,
                cluster_field,
                "candidate_log10_error",
                "baseline_log10_error",
                seed + index,
            ),
        }
    return scopes


def pair_summary(pairs: list[dict[str, Any]]) -> dict[str, Any]:
    by_creator: dict[str, list[float]] = defaultdict(list)
    for row in pairs:
        by_creator[row["creator_slug"]].append(row["aa_uplift"])
    creator_medians = np.asarray(
        [np.median(values) for _, values in sorted(by_creator.items())], dtype=float
    )
    rng = np.random.default_rng(20260724)
    draws = np.median(
        creator_medians[
            rng.integers(
                0,
                len(creator_medians),
                size=(BOOTSTRAPS, len(creator_medians)),
            )
        ],
        axis=1,
    )
    return {
        "pairs": len(pairs),
        "creators": len(by_creator),
        "checkpoint_median_aa_uplift": float(
            np.median([row["aa_uplift"] for row in pairs])
        ),
        "equal_creator_median_aa_uplift": float(np.median(creator_medians)),
        "equal_creator_median_bootstrap_90_ci": [
            float(value) for value in np.quantile(draws, [0.05, 0.95])
        ],
        "creator_medians": {
            creator: float(np.median(values))
            for creator, values in sorted(by_creator.items())
        },
    }


def anchored_frontier(
    detailed: list[dict[str, Any]], panel: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    by_slug = {row["slug"]: row for row in detailed}
    token_panel = [
        {
            "developer": row["creator_slug"],
            "release_date": row["release_date"],
            "score": row["intelligence_index"],
            "parameters_b": row["parameters_b"],
            "raw_parameter_total_b": row.get("raw_parameter_total_b", ""),
            "estimated": row["intelligence_index_estimated"],
            "answer_tokens_per_task": row["answer_tokens_per_task"],
            "reasoning_tokens_per_task": row["reasoning_tokens_per_task"],
        }
        for row in panel
        if row["output_tokens_per_task"] is not None
        and row["answer_tokens_per_task"] is not None
        and row["reasoning_tokens_per_task"] is not None
    ]
    baseline_features = ("score", "date")
    candidate_features = (
        "score",
        "date",
        "log_answer_tokens",
        "log_reasoning_tokens",
    )
    baseline_beta = fit(token_panel, baseline_features, "developer")
    candidate_beta = fit(token_panel, candidate_features, "developer")

    def target_row(source: dict[str, Any]) -> dict[str, Any]:
        return {
            "release_date": source["release_date"],
            "score": source["intelligence_index"],
            "answer_tokens_per_task": source["answer_tokens_per_task"],
            "reasoning_tokens_per_task": source["reasoning_tokens_per_task"],
        }

    k3 = target_row(by_slug[K3_SLUG])
    baseline_shift = math.log10(K3_TOTAL_B) - float(
        np.asarray([1.0, *[feature(k3, name) for name in baseline_features]])
        @ baseline_beta
    )
    candidate_shift = math.log10(K3_TOTAL_B) - float(
        np.asarray([1.0, *[feature(k3, name) for name in candidate_features]])
        @ candidate_beta
    )
    live = load_upstream_aa_factors()
    output = []
    for slug in FRONTIER_SLUGS:
        source = by_slug[slug]
        target = target_row(source)
        baseline_t = 10 ** (
            float(
                np.asarray(
                    [1.0, *[feature(target, name) for name in baseline_features]]
                )
                @ baseline_beta
            )
            + baseline_shift
        ) / 1000
        candidate_t = 10 ** (
            float(
                np.asarray(
                    [1.0, *[feature(target, name) for name in candidate_features]]
                )
                @ candidate_beta
            )
            + candidate_shift
        ) / 1000
        output.append(
            {
                "slug": slug,
                "model": source["name"],
                "release_date": source["release_date"],
                "aa_score": source["intelligence_index"],
                "answer_tokens_per_task": source["answer_tokens_per_task"],
                "reasoning_tokens_per_task": source["reasoning_tokens_per_task"],
                "current_live_aa_t": K3_TOTAL_T
                if slug == K3_SLUG
                else live[slug],
                "detailed_token_subset_baseline_t": baseline_t,
                "inference_budget_adjusted_t": candidate_t,
                "candidate_over_baseline": candidate_t / baseline_t,
                "live_use": "No — diagnostic only",
            }
        )
    return output


def main() -> None:
    detailed = parse_detail_rows()
    panel, parameter_groups = build_panel(detailed)
    exact_pairs = build_reasoning_pairs(
        parameter_groups, pair_basis="exact open-weight parameter checkpoint group"
    )
    configuration_groups = build_configuration_groups(detailed)
    pairs = build_reasoning_pairs(
        configuration_groups,
        pair_basis="creator + exact date + reported parameter fields + mode-stripped name",
    )
    pair_scope_counts = Counter(
        "open_weights" if row["is_open_weights"] else "proprietary"
        for row in pairs
    )
    if (
        len(pairs) != 100
        or pair_scope_counts != Counter({"open_weights": 61, "proprietary": 39})
        or len(exact_pairs) != 53
    ):
        raise ValueError(
            f"Unexpected reasoning-pair inventory: {len(pairs)}, {pair_scope_counts}, exact={len(exact_pairs)}"
        )
    epoch_crosscheck, unmatched_epoch = match_epoch_exact(panel)
    current_predictions = current_comparison(panel)
    token_predictions = token_comparison(panel)
    portable_reasoning_predictions = reasoning_standardization_comparison(
        panel, pairs, creator_aware=False
    )
    creator_reasoning_predictions = reasoning_standardization_comparison(
        panel, pairs, creator_aware=True
    )
    all_predictions = (
        current_predictions
        + token_predictions
        + portable_reasoning_predictions
        + creator_reasoning_predictions
    )
    frontier = anchored_frontier(detailed, panel)

    current_summary = comparison_summary(current_predictions, "developer", 20260725)
    token_summary = comparison_summary(token_predictions, "developer", 20260727)
    portable_reasoning_summary = comparison_summary(
        portable_reasoning_predictions, "developer", 20260729
    )
    creator_reasoning_summary = comparison_summary(
        creator_reasoning_predictions, "developer", 20260731
    )
    parameter_ratios = np.asarray(
        [row["aa_over_epoch_parameters"] for row in epoch_crosscheck], dtype=float
    )
    epoch_disagreements = [
        row
        for row in epoch_crosscheck
        if not row["creator_agreement"]
        or not row["date_within_45_days"]
        or not row["parameters_within_20_percent"]
    ]
    primary_overrides = [
        {
            "override_id": row["calibration_override_id"],
            "slug": row["selected_slug"],
            "model": row["selected_name"],
            "changed_fields": row["calibration_override_fields"].split(" | "),
            "source_url": row["calibration_override_source_url"],
            "source_record_sha256": row[
                "calibration_override_source_record_sha256"
            ],
            "parameter_label_available_date": row[
                "parameter_label_available_date"
            ],
            "parameter_training_eligibility_date": row[
                "parameter_training_eligibility_date"
            ],
            "architecture_class": row["architecture_class"],
            "exact_tensor_parameters": row["exact_tensor_parameters"],
            "lineage_class": row["lineage_class"],
            "base_model_id": row["base_model_id"] or None,
            "lineage_family_id": row["lineage_family_id"],
        }
        for row in panel
        if row["calibration_override_id"]
    ]
    expected_override_ids = {
        row["override_id"] for row in load_calibration_overrides()["overrides"]
    }
    if {row["override_id"] for row in primary_overrides} != expected_override_ids:
        raise ValueError(
            f"Primary metadata overrides disagree with the source ledger: {primary_overrides}"
        )
    calibration_evidence_paths = tuple(
        ROOT / evidence["path"]
        for override in load_calibration_overrides()["overrides"]
        for evidence in override["primary_source"]["local_evidence"]
    )
    label_timing_ledger = load_parameter_label_availability()
    label_timing_evidence_paths = tuple(
        ROOT / evidence["path"]
        for record in label_timing_ledger["records"]
        for evidence in record["local_evidence"]
    )
    result = {
        "generated_on": "2026-07-31",
        "question": "Do AA's detailed open-weight panel and measured inference-token budgets improve frontier parameter inference?",
        "data_audit": {
            "raw_models": len(detailed),
            "open_weight_parameter_score_date_configurations": sum(
                row["is_open_weights"]
                and row["parameters_b"] is not None
                and row["intelligence_index"] is not None
                and bool(row["release_date"])
                for row in detailed
            ),
            "unique_checkpoint_groups": len(panel),
            "lower_score_configurations_removed": sum(
                row["lower_score_configurations_removed"] for row in panel
            ),
            "creators": len({row["creator_slug"] for row in panel}),
            "token_covered_checkpoint_groups": sum(
                row["output_tokens_per_task"] is not None for row in panel
            ),
            "primary_metadata_overrides": primary_overrides,
            "parameter_label_timing_records": len(
                label_timing_ledger["records"]
            ),
            "parameter_label_timing_lags": [
                {
                    "model": record["identity"]["canonical_model_name"],
                    "release_date": record["identity"]["aa_release_date"],
                    "parameter_label_available_date": record["timing"][
                        "parameter_label_available_date"
                    ],
                    "weights_available_date": record["timing"][
                        "weights_available_date"
                    ],
                    "basis": record["timing"]["parameter_label_basis"],
                }
                for record in label_timing_ledger["records"]
            ],
            "epoch_exact_crosschecks": len(epoch_crosscheck),
            "epoch_exact_unmatched": unmatched_epoch,
            "epoch_crosschecks_with_metadata_disagreement": len(
                epoch_disagreements
            ),
            "epoch_metadata_disagreements": [
                {
                    "epoch_checkpoint_id": row["epoch_checkpoint_id"],
                    "aa_detailed_model": row["aa_detailed_model"],
                    "creator_agreement": row["creator_agreement"],
                    "date_within_45_days": row["date_within_45_days"],
                    "parameters_within_20_percent": row[
                        "parameters_within_20_percent"
                    ],
                    "match_method": row["match_method"],
                }
                for row in epoch_disagreements
            ],
            "epoch_parameter_ratio_median": float(np.median(parameter_ratios)),
            "epoch_parameter_ratio_p05_p95": [
                float(value) for value in np.quantile(parameter_ratios, [0.05, 0.95])
            ],
            "deduplication_policy": "Same normalized weight URL, total/active parameter counts, and release date; otherwise creator + normalized name + counts + date. Highest AA score retained.",
            "chronological_parameter_label_policy": "A row can train a fold only when max(checkpoint release date, parameter-label availability date) is strictly earlier than the test checkpoint date.",
            "source_preservation": "All 587 raw records and every configuration remain unchanged in the source CSV and compressed HTML. A separate hash-pinned primary-source overlay corrects stale Motif 3 Beta and Motif-2-12.7B-Reasoning openness metadata only in the calibration view.",
            "family_balance_policy": "Regression weights are allocated within creator/developer and every held-out fold removes the entire creator. Motif-2-12.7B-Reasoning is additionally tagged same-base-posttrain to Motif-2-12.7B-Base, so it is not treated as a new independent developer or pretraining family.",
        },
        "same_weight_reasoning_pairs": pair_summary(exact_pairs),
        "reasoning_configuration_pairs": {
            "all": pair_summary(pairs),
            "open_weights": pair_summary(
                [row for row in pairs if row["is_open_weights"]]
            ),
            "proprietary": pair_summary(
                [row for row in pairs if not row["is_open_weights"]]
            ),
            "pairing_policy": "All model configurations are preserved. Reasoning and non-reasoning variants are paired only when creator, exact release date, reported parameter fields, and mode-stripped name agree; highest score per mode is retained.",
            "interpretation": "The equal-creator median validates the existing global six-point reasoning correction, while large creator variation argues against a universal lab-specific adjustment without chronological validation.",
        },
        "detailed_panel_backtest": {
            "target": "log10 total parameters in billions",
            "baseline": "current 50-model AA panel: AA score + exact date",
            "candidate": f"{len(panel)}-checkpoint detailed AA panel: AA score + exact date",
            "outer_split": "strictly earlier parameter-training eligibility date; entire test developer removed",
            "training_weights": "equal total weight per developer; estimated AA scores receive 0.5 weight",
            "identical_test_rows": len(current_predictions),
            "scopes": current_summary,
        },
        "inference_budget_backtest": {
            "target": "log10 total parameters in billions",
            "baseline": "AA score + exact date",
            "candidate": "AA score + exact date + log10(1 + answer tokens/task) + log10(1 + reasoning tokens/task)",
            "outer_split": "strictly earlier parameter-training eligibility date; entire test developer removed",
            "training_weights": "equal total weight per developer; estimated AA scores receive 0.5 weight",
            "identical_test_rows": len(token_predictions),
            "scopes": token_summary,
        },
        "reasoning_standardization_backtest": {
            "target": "log10 total parameters in billions",
            "baseline": "raw highest-configuration AA score + exact date",
            "candidate": "no-reasoning-equivalent AA score + exact date",
            "standardization_priority": "direct same-checkpoint non-reasoning score; otherwise strictly earlier pair history only",
            "outer_split": "strictly earlier parameter-training eligibility date; entire test developer removed",
            "portable": {
                "history_rule": "equal-creator median of strictly earlier pairs from other creators",
                "identical_test_rows": len(portable_reasoning_predictions),
                "scopes": portable_reasoning_summary,
            },
            "creator_aware": {
                "history_rule": "median of at least two strictly earlier same-creator pairs, otherwise portable fallback",
                "identical_test_rows": len(creator_reasoning_predictions),
                "scopes": creator_reasoning_summary,
            },
        },
        "frontier_inference_budget_diagnostic": frontier,
        "k3_anchor": {
            "total_parameters_b": K3_TOTAL_B,
            "source": K3_PARAMETER_SOURCE,
        },
        "decision": {
            "change_live_aa_branch": False,
            "incremental_detailed_panel_weight": 0.0,
            "incremental_inference_budget_weight": 0.0,
            "incremental_reasoning_standardization_weight": 0.0,
            "reason": "The detailed panel improves broad and mid-frontier recovery, but its frontier-like interval is not decisive. Measured token budgets and chronological reasoning standardization improve some mean/tail metrics while failing to improve all primary median/error criteria with a decisive cluster interval. Preserve all three as prospective diagnostics until new high-score parameter disclosures validate the direction.",
        },
        "limitations": [
            "AA's current benchmark snapshot is not a historical benchmark vintage.",
            "AA parameter counts are public metadata and can be rounded; Epoch remains the higher-priority exact source on matched checkpoints.",
            "Motif 3 Beta is admitted through an explicit official-repository override because the frozen AA row predates its public weight release; the original AA row is retained unchanged.",
            "Motif's AA checkpoint date remains 2026-07-14, but its parameter label cannot train a chronological fold until after the official weights appeared on 2026-07-20.",
            "Motif-2-12.7B-Reasoning is admitted through an exact checkpoint match to the official Apache-2.0 repository. Its weights appeared on 2025-12-01, before AA's 2025-12-04 checkpoint date, so chronology uses the later AA date.",
            "Motif-2-12.7B-Reasoning shares Motif-2-12.7B-Base lineage and the Motif Technologies developer allocation; it is not counted as an independent family-level observation.",
            "Six recent AA checkpoints have a separate, primary-source parameter-label timing ledger. MiMo-V2.5 and LongCat 2.0 retain launch-article label dates earlier than their weight drops; the other rows use the first public technical report or official repository evidence.",
            "Output tokens are an inference-budget proxy, not FLOPs: model architecture, speculative decoding, tool calls, and cascades are not fully observed.",
            "Only token-covered rows can enter the incremental comparison, creating non-random availability.",
            "K3 has no non-reasoning AA configuration, so the inference-budget branch is anchored to its observed max-thinking configuration rather than a measured base-only score.",
        ],
        "source_manifest": {
            str(path.relative_to(ROOT)): sha256(path)
            for path in (
                DETAIL,
                RAW,
                COLLECTION,
                AA_CALIBRATION_OVERRIDES_PATH,
                AA_PARAMETER_LABEL_AVAILABILITY_PATH,
                AA_SCORE_AVAILABILITY_PATH,
                AA_CHANGELOG_PATH,
                OPEN_MODEL_PARAMETER_TRUTH_PATH,
                EXACT,
                BRANCH_WORKBOOK,
                K3_EVIDENCE_PATH,
                *calibration_evidence_paths,
                *label_timing_evidence_paths,
            )
        },
        "outputs": {
            "panel": str(PANEL.relative_to(ROOT)),
            "reasoning_pairs": str(PAIRS.relative_to(ROOT)),
            "epoch_crosscheck": str(EPOCH_CROSSCHECK.relative_to(ROOT)),
            "predictions": str(PREDICTIONS.relative_to(ROOT)),
        },
    }
    write_csv(PANEL, panel)
    write_csv(PAIRS, pairs)
    write_csv(EPOCH_CROSSCHECK, epoch_crosscheck)
    write_csv(PREDICTIONS, all_predictions)
    RESULT.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(RESULT),
                "models": len(detailed),
                "panel": len(panel),
                "reasoning_pairs": len(pairs),
                "exact_weight_reasoning_pairs": len(exact_pairs),
                "epoch_crosschecks": len(epoch_crosscheck),
                "token_predictions": len(token_predictions),
                "portable_reasoning_predictions": len(
                    portable_reasoning_predictions
                ),
                "creator_reasoning_predictions": len(
                    creator_reasoning_predictions
                ),
                "change_live_aa_branch": False,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
